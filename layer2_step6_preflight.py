"""
Layer 2 Step 6 PREFLIGHT — ABS ERP at SA2 ingest
─────────────────────────────────────────────────
READ-ONLY. No DB mutations. No file writes except recon/.

Per project status doc 2026-04-26b, this preflight covers:
  A. Confirm Table 1 column structure (header row 6 confirmed)
  B. Sample first 5 SA2 rows to verify code/name/year layout
  C. Project ingest volume (~25K rows expected)
  D. Check overlap with services.sa2_code (post-step-1)

Outputs:
  - Console summary
  - recon/layer2_step6_preflight.md
"""

from __future__ import annotations

import os
import sqlite3
import sys
from pathlib import Path

import openpyxl

# ──────────────────────────────────────────────────────────────────────
# Paths (relative to project root — script must be run from there)
# ──────────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path.cwd()
WORKBOOK = PROJECT_ROOT / "abs_data" / "Population and People Database.xlsx"
DB = PROJECT_ROOT / "data" / "kintell.db"
OUT_DIR = PROJECT_ROOT / "recon"
OUT_FILE = OUT_DIR / "layer2_step6_preflight.md"

HEADER_ROW = 6  # confirmed in Layer 1 recon

# Under-5 cohort column indices documented in the status doc
UNDER_5_COL_INDICES = [12, 48, 84]


def banner(title: str) -> str:
    return f"\n{'─' * 70}\n{title}\n{'─' * 70}"


def main() -> int:
    findings: list[str] = []

    findings.append("# Layer 2 Step 6 — Preflight Findings")
    findings.append("")
    findings.append("Read-only inventory of ABS ERP workbook before ingest.")
    findings.append("")

    # ── A. File presence ─────────────────────────────────────────────
    print(banner("A. FILE PRESENCE"))
    findings.append("## A. File presence\n")

    if not WORKBOOK.exists():
        msg = f"FAIL: workbook not found at {WORKBOOK}"
        print(msg)
        findings.append(f"- **FAIL**: workbook not found at `{WORKBOOK}`")
        Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
        OUT_FILE.write_text("\n".join(findings), encoding="utf-8")
        return 1

    size_mb = WORKBOOK.stat().st_size / (1024 * 1024)
    msg = f"OK: workbook present at {WORKBOOK} ({size_mb:.2f} MB)"
    print(msg)
    findings.append(f"- OK: `{WORKBOOK.name}` present ({size_mb:.2f} MB)")

    if not DB.exists():
        msg = f"FAIL: DB not found at {DB}"
        print(msg)
        findings.append(f"- **FAIL**: DB not found at `{DB}`")
        Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
        OUT_FILE.write_text("\n".join(findings), encoding="utf-8")
        return 1
    print(f"OK: DB present at {DB}")
    findings.append(f"- OK: DB present at `{DB}`\n")

    # ── B. Open Table 1 ──────────────────────────────────────────────
    print(banner("B. OPEN TABLE 1"))
    findings.append("## B. Open workbook Table 1\n")

    try:
        wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    except Exception as exc:
        msg = f"FAIL: openpyxl error: {exc}"
        print(msg)
        findings.append(f"- **FAIL**: openpyxl error: `{exc}`")
        Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
        OUT_FILE.write_text("\n".join(findings), encoding="utf-8")
        return 1

    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")
    findings.append(f"- Sheets in workbook: `{sheet_names}`")

    # Table 1 — locate by name pattern (handles "Table 1", "Table 1.", etc.)
    table1_name = None
    for name in sheet_names:
        norm = name.lower().strip().replace(" ", "")
        if norm.startswith("table1"):
            table1_name = name
            break
    if table1_name is None:
        msg = "FAIL: cannot locate a 'Table 1' sheet"
        print(msg)
        findings.append(f"- **FAIL**: {msg}")
        wb.close()
        Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
        OUT_FILE.write_text("\n".join(findings), encoding="utf-8")
        return 1
    ws = wb[table1_name]
    print(f"OK: opened sheet '{table1_name}'")
    findings.append(f"- OK: opened sheet `{table1_name}`\n")

    # ── C. Header row confirmation + columns of interest ─────────────
    print(banner("C. HEADER ROW & COLUMNS OF INTEREST"))
    findings.append("## C. Header row & under-5 cohort columns\n")

    # Read first HEADER_ROW + 6 rows so we can show context + sample data
    rows: list[tuple] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(row)
        if row_idx >= HEADER_ROW + 5:
            break

    if len(rows) < HEADER_ROW:
        msg = f"FAIL: sheet has fewer than {HEADER_ROW} rows"
        print(msg)
        findings.append(f"- **FAIL**: {msg}")
        wb.close()
        Path(OUT_DIR).mkdir(parents=True, exist_ok=True)
        OUT_FILE.write_text("\n".join(findings), encoding="utf-8")
        return 1

    header = rows[HEADER_ROW - 1]
    total_cols = sum(1 for c in header if c is not None)
    print(f"Header row {HEADER_ROW}: {total_cols} non-null columns "
          f"(workbook has {len(header)} total)")
    findings.append(f"- Header row {HEADER_ROW}: {total_cols} non-null "
                    f"columns of {len(header)} total\n")

    # Show first 6 columns (key cols) + the 3 under-5 cohort columns
    findings.append("### Key columns (first 6)\n")
    findings.append("| idx | header |")
    findings.append("|---:|---|")
    for i in range(min(6, len(header))):
        findings.append(f"| {i} | `{header[i]}` |")
    findings.append("")

    findings.append("### Under-5 cohort columns "
                    "(per status doc: indices 12, 48, 84)\n")
    findings.append("| idx | header |")
    findings.append("|---:|---|")
    print("Under-5 cohort columns:")
    for i in UNDER_5_COL_INDICES:
        if i < len(header):
            print(f"  col[{i}]: {header[i]}")
            findings.append(f"| {i} | `{header[i]}` |")
        else:
            print(f"  col[{i}]: OUT OF RANGE (only {len(header)} columns)")
            findings.append(f"| {i} | **OUT OF RANGE** |")
    findings.append("")

    # ── D. Sample 5 SA2 data rows ────────────────────────────────────
    print(banner("D. FIRST 5 SA2 SAMPLE ROWS"))
    findings.append("## D. First 5 sample data rows (after header)\n")
    findings.append("Showing column 0 (likely region code), column 1 "
                    "(likely region name), and the 3 under-5 cohort columns.\n")
    findings.append("| code | name | u5_a | u5_b | u5_c |")
    findings.append("|---|---|---:|---:|---:|")

    sample_rows = rows[HEADER_ROW:HEADER_ROW + 5]
    for r in sample_rows:
        code = r[0] if 0 < len(r) else None
        name = r[1] if 1 < len(r) else None
        u5_vals = []
        for i in UNDER_5_COL_INDICES:
            u5_vals.append(r[i] if i < len(r) else "OOR")
        line = (f"  {code} | {name} | "
                f"{u5_vals[0]} | {u5_vals[1]} | {u5_vals[2]}")
        print(line)
        findings.append(f"| `{code}` | `{name}` | "
                        f"{u5_vals[0]} | {u5_vals[1]} | {u5_vals[2]} |")
    findings.append("")

    # ── E. Project ingest volume — count SA2-shaped rows ─────────────
    print(banner("E. PROJECT INGEST VOLUME"))
    findings.append("## E. Projected ingest volume\n")

    # Stream the entire sheet, tally rows whose col[0] looks like a 9-digit SA2
    sa2_rows = 0
    non_sa2_rows = 0
    blank_rows = 0
    total_data_rows = 0

    # Re-iterate from the start; openpyxl read-only is forward-only
    wb2 = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws2 = wb2[table1_name]
    for row_idx, row in enumerate(ws2.iter_rows(values_only=True), start=1):
        if row_idx <= HEADER_ROW:
            continue
        total_data_rows += 1
        first_cell = row[0] if len(row) > 0 else None
        if first_cell is None or str(first_cell).strip() == "":
            blank_rows += 1
            continue
        s = str(first_cell).strip()
        if s.isdigit() and len(s) == 9:
            sa2_rows += 1
        else:
            non_sa2_rows += 1
    wb2.close()

    print(f"Total data rows:     {total_data_rows:,}")
    print(f"  SA2 rows (9-digit): {sa2_rows:,}")
    print(f"  Non-SA2 rows:       {non_sa2_rows:,}")
    print(f"  Blank rows:         {blank_rows:,}")
    findings.append(f"- Total data rows: **{total_data_rows:,}**")
    findings.append(f"- SA2 rows (9-digit code in col 0): **{sa2_rows:,}**")
    findings.append(f"- Non-SA2 rows (Australia/State/GCCSA aggregates): "
                    f"**{non_sa2_rows:,}**")
    findings.append(f"- Blank rows: **{blank_rows:,}**")
    findings.append("")
    findings.append("Per status doc, expected ingest volume ~25K rows after "
                    "filter to 9-digit SA2 codes.\n")

    # ── F. DB overlap with services.sa2_code ─────────────────────────
    print(banner("F. OVERLAP WITH services.sa2_code"))
    findings.append("## F. Overlap with `services.sa2_code` "
                    "(post Step 1)\n")

    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    cur = conn.cursor()

    # Total services with non-null sa2_code
    cur.execute("SELECT COUNT(DISTINCT sa2_code) "
                "FROM services WHERE sa2_code IS NOT NULL")
    distinct_service_sa2 = cur.fetchone()[0]
    print(f"Distinct services.sa2_code values: {distinct_service_sa2:,}")
    findings.append(f"- Distinct `services.sa2_code` values "
                    f"(non-null): **{distinct_service_sa2:,}**")

    # Build set of SA2 codes from workbook (we already counted them above
    # but didn't store them — re-stream for memory safety)
    wb3 = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws3 = wb3[table1_name]
    workbook_sa2: set[str] = set()
    for row_idx, row in enumerate(ws3.iter_rows(values_only=True), start=1):
        if row_idx <= HEADER_ROW:
            continue
        if not row:
            continue
        first_cell = row[0]
        if first_cell is None:
            continue
        s = str(first_cell).strip()
        if s.isdigit() and len(s) == 9:
            workbook_sa2.add(s)
    wb3.close()

    print(f"Workbook distinct SA2 codes: {len(workbook_sa2):,}")
    findings.append(f"- Workbook distinct SA2 codes: "
                    f"**{len(workbook_sa2):,}**")

    # Overlap
    cur.execute("SELECT DISTINCT sa2_code FROM services "
                "WHERE sa2_code IS NOT NULL")
    service_sa2_set = {str(r[0]).strip() for r in cur.fetchall()
                       if r[0] is not None}

    matched = service_sa2_set & workbook_sa2
    unmatched = service_sa2_set - workbook_sa2
    print(f"  matched in workbook:   {len(matched):,}")
    print(f"  unmatched in workbook: {len(unmatched):,}")
    findings.append(f"- Service SA2 codes matched in workbook: "
                    f"**{len(matched):,}**")
    findings.append(f"- Service SA2 codes UNMATCHED in workbook: "
                    f"**{len(unmatched):,}**")
    if unmatched:
        sample = sorted(unmatched)[:10]
        findings.append(f"  - first 10 unmatched: `{sample}`")
    findings.append("")

    conn.close()
    wb.close()

    # ── G. Recommended schema ────────────────────────────────────────
    findings.append("## G. Recommended schema (for review)\n")
    findings.append("```")
    findings.append("CREATE TABLE abs_sa2_erp_annual (")
    findings.append("    sa2_code     TEXT    NOT NULL,")
    findings.append("    year         INTEGER NOT NULL,")
    findings.append("    age_group    TEXT    NOT NULL, "
                    "-- 'persons' / 'males' / 'females'")
    findings.append("    persons      INTEGER,")
    findings.append("    PRIMARY KEY (sa2_code, year, age_group)")
    findings.append(");")
    findings.append("```")
    findings.append("")
    findings.append("(Long format. Wide-format alternative deferred to ingest "
                    "step decision.)")
    findings.append("")

    # ── Write findings ───────────────────────────────────────────────
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text("\n".join(findings), encoding="utf-8")
    print(banner("DONE"))
    print(f"Findings written to: {OUT_FILE}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
