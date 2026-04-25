"""
ingest_ncver_completions.py  v2
════════════════════════════════════════════════════════════════
Phase 1.5 Step 2 — ingest NCVER program-completions XLSX into the
training_completions staging tables created by
migrate_training_completions_tables.py.

v2 changes (2026-04-25):
  - Validation tolerance loosened to ±25 per state/year aggregate.
    NCVER rounds every cell independently to the nearest 5, so leaf
    sums drift from state totals by small amounts that are NOT
    parser bugs. v1 enforced exact equality and aborted on these
    rounding artefacts. Errors beyond ±25 still abort (those would
    indicate a real structural problem like dropped rows).
  - Caveats string written to ingest_run now explicitly notes the
    rounding-to-5 behaviour and per-state-year drift range.
  - Validation summary prints number of mismatches within tolerance
    so the rounding artefact is visible in the log.

SOURCE
──────
  abs_data/NCVER_ECEC_Completions_2019-2024.xlsx
  (override with --source-file)

  NCVER VOCSTATS Program Completions DataBuilder export, with
  these dimensions:
    Total / State/territory of residence / Remoteness region /
    Program name × Year (2019–2024)
  Filtered to four ECEC qualification codes:
    CHC30113 / CHC30121 / CHC50113 / CHC50121

WHAT IT DOES
────────────
  1. Reads the XLSX top-to-bottom, walking the four-level hierarchy
     (national total → state total → state+remoteness total →
      state+remoteness+program leaf).
  2. Skips aggregate rows; ingests only leaf rows.
  3. Splits each leaf row into per-year records (6 years → 6 rows
     per leaf).
  4. Stores '-' as 0 per source's footer note.
  5. Captures the file's footer (filter description, caveats) into
     a new training_completions_ingest_run row.
  6. Validates leaf sums against the file's own state-level
     aggregates and rolls back on mismatch.

DESIGN
──────
  • Wrapped in a single transaction with pre/post-state validation.
  • Takes a DB backup before mutation.
  • Re-ingest behaviour:
      default — aborts if any rows from this source_file already
                exist in training_completions_ingest_run.
      --force — deletes prior rows from this source_file's prior
                runs (cascades through training_completions),
                then ingests fresh. Always backed up beforehand.
  • Post-ingest validation: sum of leaf rows per (state, year) is
    compared to the state-total aggregate row in the source file.
    Mismatch = rollback.

USAGE
─────
    python ingest_ncver_completions.py --dry-run
    python ingest_ncver_completions.py
    python ingest_ncver_completions.py --force      (re-ingest)
    python ingest_ncver_completions.py --source-file path/to.xlsx

VALIDATION (post-run)
─────────────────────
    python -c "import sqlite3; c=sqlite3.connect('data/kintell.db'); print(c.execute('SELECT COUNT(*) FROM training_completions').fetchone()[0])"
    → expects: ~1,080 (180 leaf rows × 6 years), exact count varies
                with how source publishes Offshore/Not known.
"""

import argparse
import json
import os
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

os.environ.setdefault("PYTHONIOENCODING", "utf-8")
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    sys.exit("✗ Missing 'openpyxl'. Run: pip install openpyxl --break-system-packages")

DB_PATH = Path("data") / "kintell.db"
DEFAULT_SOURCE = Path("abs_data") / "NCVER_ECEC_Completions_2019-2024.xlsx"
BACKUP_PREFIX = "kintell.db.backup_pre_ncver_ingest_"
ACTOR = "ingest_ncver_completions.py"

SOURCE_LABEL = "NCVER VOCSTATS Total VET students and courses 2024 — Program Completions"
PUBLICATION_DATE = "2025-09-22"  # from the publication header

# v2: validation tolerance.
# NCVER rounds every cell independently to the nearest 5, so leaf sums
# drift from the source's state totals by small amounts. With ~20 leaf
# cells per state/year (4 programs × 5 remoteness bands), maximum
# theoretical drift is ±50 if every cell rounds in the same direction;
# in practice it's well under that, with most observed drifts at ±5
# and the worst at ±15. We treat anything within ±25 as a normal
# rounding artefact; beyond that suggests a real structural problem
# (dropped rows, parser bug, source format change).
VALIDATION_TOLERANCE = 25

# State-name → state-code mapping. NCVER uses full names; schema wants codes.
STATE_CODE = {
    "New South Wales":            "NSW",
    "Victoria":                   "VIC",
    "Queensland":                 "QLD",
    "South Australia":            "SA",
    "Western Australia":          "WA",
    "Tasmania":                   "TAS",
    "Northern Territory":         "NT",
    "Australian Capital Territory": "ACT",
    "Offshore":                   "OFFSHORE",
    "Not known":                  "UNKNOWN",
}

# Qualification code → (level, era) lookup. Era = 'old' for pre-2021
# codes (ending 0113), 'new' for current codes (ending 0121).
QUAL_META = {
    "CHC30113": ("cert3",   "old"),
    "CHC30121": ("cert3",   "new"),
    "CHC50113": ("diploma", "old"),
    "CHC50121": ("diploma", "new"),
}

# Marker strings used by NCVER for aggregate rows. Parser skips these.
TOTAL_MARKER_STATE      = "State/territory of residence total"
TOTAL_MARKER_REMOTENESS = "Remoteness region total"
TOTAL_MARKER_PROGRAM    = "Program name total"

# Source uses '-' to mean zero (per footer: "A dash represents a zero")
DASH = "-"


def take_backup(db_path):
    if not db_path.exists():
        raise SystemExit(f"  ✗ DB not found: {db_path}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = db_path.parent / f"{BACKUP_PREFIX}{ts}"
    shutil.copy2(db_path, backup)
    return backup


def parse_completions_value(raw):
    """Convert a source cell to an integer count.
    '-' → 0, None → 0, integer → integer, string-numeric → integer.
    """
    if raw is None:
        return 0
    if isinstance(raw, str):
        s = raw.strip()
        if s == DASH or s == "":
            return 0
        try:
            return int(s)
        except ValueError:
            raise ValueError(f"unparseable completion cell: {raw!r}")
    if isinstance(raw, (int, float)):
        return int(raw)
    raise ValueError(f"unexpected cell type for completions: {type(raw).__name__}")


def extract_qualification_code(program_label):
    """Pull the leading code from 'CHC30121 - Certificate III ...'."""
    if not program_label:
        return None
    parts = program_label.split(" - ", 1)
    code = parts[0].strip()
    return code if code in QUAL_META else None


def is_aggregate_row(row):
    """True if any column carries an aggregate-marker string."""
    state = row[1]
    remoteness = row[2]
    program = row[3]
    if state == TOTAL_MARKER_STATE:
        return True, "national"
    if remoteness == TOTAL_MARKER_REMOTENESS:
        return True, "state"
    if program == TOTAL_MARKER_PROGRAM:
        return True, "state_remoteness"
    return False, None


def parse_workbook(path):
    """
    Walk the workbook top-to-bottom. Returns:
      leaf_records: list of dicts (state, remoteness, qual_code, year, completions)
      state_year_aggregates: dict {(state_name, year): total_from_source}
                             (used for post-ingest validation)
      footer_text: list of footer strings for caveat capture
      year_columns: list of years detected in the header
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Header row at R1: ['Total', 'State/...', 'Remoteness ...', 'Program name', 'Year', 'Year total', 2019, 2020, ...]
    header = [c.value for c in ws[1]]
    year_columns = []
    year_col_indices = []  # column indices (0-based) for each year
    for idx, val in enumerate(header):
        if isinstance(val, int) and 2000 <= val <= 2100:
            year_columns.append(val)
            year_col_indices.append(idx)
    if not year_columns:
        raise SystemExit("  ✗ No year columns detected in header. File structure unexpected.")

    leaf_records = []
    state_year_aggregates = {}
    footer_text = []

    current_state = None
    current_remoteness = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Detect footer rows: they tend to have content only in column A,
        # and that content is prose, not numeric.
        first_cell = row[0]
        rest_empty = all((c is None or c == "") for c in row[1:])
        if rest_empty and isinstance(first_cell, str) and first_cell.strip():
            footer_text.append(first_cell.strip())
            continue
        # Fully empty row — skip
        if all((c is None or c == "") for c in row):
            continue

        state_cell = row[1]
        remoteness_cell = row[2]
        program_cell = row[3]

        # Update state context when a new state appears
        if state_cell and state_cell != TOTAL_MARKER_STATE:
            if state_cell in STATE_CODE:
                current_state = state_cell
                current_remoteness = None  # reset remoteness when state changes
        # National total row → capture nothing (skip)
        if state_cell == TOTAL_MARKER_STATE:
            continue

        # State-total row (remoteness column says "Remoteness region total")
        # Capture state-year aggregates for post-ingest validation
        if remoteness_cell == TOTAL_MARKER_REMOTENESS:
            if current_state:
                for year, col_idx in zip(year_columns, year_col_indices):
                    if col_idx < len(row):
                        try:
                            state_year_aggregates[(current_state, year)] = parse_completions_value(row[col_idx])
                        except ValueError:
                            pass
            continue

        # Remoteness in this column → it's a real remoteness band, update context
        if remoteness_cell and remoteness_cell != TOTAL_MARKER_REMOTENESS:
            current_remoteness = remoteness_cell
        # State-remoteness total row → skip but state-remoteness already captured
        if program_cell == TOTAL_MARKER_PROGRAM:
            continue

        # Program leaf row: extract code and emit one record per year
        qual_code = extract_qualification_code(program_cell)
        if qual_code is None:
            continue  # not one of our four codes
        if not current_state:
            continue  # safety: shouldn't happen if file structure is intact

        # 'remoteness_band' is None when source has no remoteness layer for
        # this state (e.g. Offshore / Unknown sometimes lack one)
        remoteness_band = current_remoteness

        for year, col_idx in zip(year_columns, year_col_indices):
            if col_idx >= len(row):
                continue
            try:
                completions = parse_completions_value(row[col_idx])
            except ValueError as e:
                raise SystemExit(f"  ✗ Row {row_idx}, year {year}: {e}")
            qual_level, qual_era = QUAL_META[qual_code]
            leaf_records.append({
                "state_name":          current_state,
                "state_code":          STATE_CODE[current_state],
                "remoteness_band":     remoteness_band,
                "qualification_code":  qual_code,
                "qualification_name":  program_cell.split(" - ", 1)[1].strip()
                                        if " - " in program_cell else program_cell,
                "qualification_level": qual_level,
                "qualification_era":   qual_era,
                "year":                year,
                "completions":         completions,
            })

    return leaf_records, state_year_aggregates, footer_text, year_columns


def validate_against_aggregates(leaf_records, state_year_aggregates):
    """Sum leaf records per (state, year) and compare to source's state totals.
    Returns (ok, mismatches, within_tolerance) where:
      mismatches      = list of (state, year, leaf_sum, source_total)
                        for differences EXCEEDING the rounding tolerance
                        (these are real errors that should abort)
      within_tolerance = list of (state, year, leaf_sum, source_total)
                        for differences within ±VALIDATION_TOLERANCE
                        (these are NCVER rounding artefacts, not bugs)
    """
    sums = {}
    for r in leaf_records:
        key = (r["state_name"], r["year"])
        sums[key] = sums.get(key, 0) + r["completions"]

    mismatches = []
    within_tolerance = []
    for key, source_total in state_year_aggregates.items():
        leaf_sum = sums.get(key, 0)
        diff = leaf_sum - source_total
        if diff == 0:
            continue
        if abs(diff) <= VALIDATION_TOLERANCE:
            within_tolerance.append((key[0], key[1], leaf_sum, source_total))
        else:
            mismatches.append((key[0], key[1], leaf_sum, source_total))

    return len(mismatches) == 0, mismatches, within_tolerance


def existing_runs_for_source(conn, source_file):
    cur = conn.execute(
        "SELECT run_id FROM training_completions_ingest_run WHERE source_file = ?",
        (str(source_file),),
    )
    return [r[0] for r in cur.fetchall()]


def delete_prior_runs(conn, run_ids):
    if not run_ids:
        return 0
    placeholders = ",".join("?" * len(run_ids))
    cur = conn.execute(
        f"DELETE FROM training_completions WHERE ingest_run_id IN ({placeholders})",
        run_ids,
    )
    deleted_leaves = cur.rowcount
    conn.execute(
        f"DELETE FROM training_completions_ingest_run WHERE run_id IN ({placeholders})",
        run_ids,
    )
    return deleted_leaves


def write_audit_row(conn, source_file, run_id, rows_ingested, deleted_leaves):
    before_json = {"deleted_prior_leaves": deleted_leaves}
    after_json = {
        "run_id": run_id,
        "rows_ingested": rows_ingested,
        "source_file": str(source_file),
    }
    reason = (
        "NCVER training-completions ingest. Phase 1.5 Step 2. "
        "Leaf rows only; aggregate rows in source skipped and "
        "re-derived at query time."
    )
    conn.execute(
        """INSERT INTO audit_log
              (actor, action, subject_type, subject_id,
               before_json, after_json, reason, occurred_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            ACTOR,
            "ingest_ncver_completions",
            "table",
            "training_completions",
            json.dumps(before_json),
            json.dumps(after_json),
            reason,
            datetime.now().isoformat(timespec="seconds"),
        ),
    )


def run(dry_run, force, source_file):
    print("─" * 64)
    print("  NCVER training-completions ingest")
    print("  (Phase 1.5 Step 2)")
    print("─" * 64)

    if not DB_PATH.exists():
        print(f"  ✗ {DB_PATH} not found.")
        return 1
    if not source_file.exists():
        print(f"  ✗ Source file not found: {source_file}")
        return 1

    print(f"  DB         : {DB_PATH}")
    print(f"  Source     : {source_file}  ({source_file.stat().st_size:,} bytes)")
    print(f"  Mode       : {'DRY RUN' if dry_run else 'APPLY'}{'  (--force)' if force else ''}")

    # Parse the workbook (read-only — safe before any DB mutation)
    leaf_records, state_year_aggregates, footer_text, year_columns = parse_workbook(source_file)

    print(f"  Years      : {year_columns}")
    print(f"  Leaf rows  : {len(leaf_records)}")
    print(f"  State/year aggregates captured : {len(state_year_aggregates)}")
    print(f"  Footer lines : {len(footer_text)}")

    # Validate against source aggregates BEFORE touching DB
    ok, mismatches, within_tolerance = validate_against_aggregates(leaf_records, state_year_aggregates)
    if not ok:
        print(f"\n  ✗ Validation failed. {len(mismatches)} state/year aggregates differ by more than ±{VALIDATION_TOLERANCE}:")
        for state, year, leaf_sum, src_total in mismatches[:10]:
            print(f"    {state} {year}: leaf-sum={leaf_sum}, source-total={src_total}, diff={leaf_sum - src_total}")
        if len(mismatches) > 10:
            print(f"    ... and {len(mismatches) - 10} more.")
        print("  Aborting. Real structural error suspected — do not ingest.")
        return 1
    print(f"  Validation : ✓ all {len(state_year_aggregates)} state/year aggregates within ±{VALIDATION_TOLERANCE} of leaf sums")
    if within_tolerance:
        max_drift = max(abs(t[2] - t[3]) for t in within_tolerance)
        print(f"               ({len(within_tolerance)} aggregates show small drift ≤ ±{max_drift}, attributed to NCVER rounding-to-5 per cell)")

    # Brief preview of what would be inserted
    by_state = {}
    for r in leaf_records:
        by_state[r["state_code"]] = by_state.get(r["state_code"], 0) + 1
    print(f"\n  Leaf rows by state code:")
    for code in sorted(by_state):
        print(f"    {code:10s}  {by_state[code]:3d} rows")

    if dry_run:
        print("\n  ✓ Dry run complete. No changes made.")
        return 0

    # Connect, check for prior runs, handle --force
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    prior_runs = existing_runs_for_source(conn, source_file)

    if prior_runs and not force:
        print(f"\n  ✗ Source file already ingested in run(s) {prior_runs}.")
        print(f"  Use --force to delete prior rows and re-ingest.")
        conn.close()
        return 1

    # Backup before any mutation
    backup = take_backup(DB_PATH)
    print(f"\n  Backup : {backup}")

    deleted_leaves = 0
    try:
        if prior_runs and force:
            deleted_leaves = delete_prior_runs(conn, prior_runs)
            print(f"  Deleted {deleted_leaves} prior leaf rows from runs {prior_runs}")

        # Build caveats string from footer plus an explicit rounding note
        rounding_note = (
            "Per NCVER footer: 'Numbers are rounded to the nearest 5. "
            "A dash represents a zero.' Each cell is rounded independently, "
            "so leaf-row sums may drift from source state-totals by up to "
            f"±{VALIDATION_TOLERANCE} per state/year. Drifts within tolerance "
            "are accepted; drifts beyond it would have aborted this ingest."
        )
        caveats_lines = [rounding_note] + [line for line in footer_text if line]
        caveats = "\n".join(caveats_lines)

        # Find filter description specifically (it's the longest "Filters applied:" line)
        filter_desc = next((l for l in footer_text if l.startswith("Filters applied:")), "")

        # Insert ingest run
        cur = conn.execute(
            """INSERT INTO training_completions_ingest_run
                  (source_file, source_label, publication_date,
                   filter_description, caveats, rows_ingested, ingested_by)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                str(source_file),
                SOURCE_LABEL,
                PUBLICATION_DATE,
                filter_desc,
                caveats,
                0,  # placeholder — updated after leaf insert
                ACTOR,
            ),
        )
        run_id = cur.lastrowid

        # Insert leaf rows
        for r in leaf_records:
            conn.execute(
                """INSERT INTO training_completions
                      (state_code, state_name, remoteness_band,
                       qualification_code, qualification_name,
                       qualification_level, qualification_era,
                       year, completions, ingest_run_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    r["state_code"], r["state_name"], r["remoteness_band"],
                    r["qualification_code"], r["qualification_name"],
                    r["qualification_level"], r["qualification_era"],
                    r["year"], r["completions"], run_id,
                ),
            )

        # Update rows_ingested on the run row
        conn.execute(
            "UPDATE training_completions_ingest_run SET rows_ingested = ? WHERE run_id = ?",
            (len(leaf_records), run_id),
        )

        # Audit log
        write_audit_row(conn, source_file, run_id, len(leaf_records), deleted_leaves)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"\n  ✗ Ingest failed: {e}")
        print(f"  Recovery: restore from {backup}")
        return 1

    # Post-ingest validation: re-query DB and compare against source aggregates
    print(f"\n  Inserted   : {len(leaf_records)} leaf rows under run_id {run_id}")

    cur = conn.execute(
        "SELECT state_name, year, SUM(completions) FROM training_completions "
        "WHERE ingest_run_id = ? GROUP BY state_name, year",
        (run_id,),
    )
    db_sums = {(r[0], r[1]): r[2] for r in cur.fetchall()}

    db_mismatches = []
    db_within_tolerance = []
    for key, src_total in state_year_aggregates.items():
        db_sum = db_sums.get(key, 0)
        diff = db_sum - src_total
        if diff == 0:
            continue
        if abs(diff) <= VALIDATION_TOLERANCE:
            db_within_tolerance.append((key[0], key[1], db_sum, src_total))
        else:
            db_mismatches.append((key[0], key[1], db_sum, src_total))

    if db_mismatches:
        print(f"\n  ✗ Post-ingest validation failed (differences exceed ±{VALIDATION_TOLERANCE}). Rolling back.")
        for state, year, db_sum, src_total in db_mismatches[:10]:
            print(f"    {state} {year}: db-sum={db_sum}, source-total={src_total}")
        # Rollback by deleting the just-inserted run
        conn.execute(
            "DELETE FROM training_completions WHERE ingest_run_id = ?", (run_id,)
        )
        conn.execute(
            "DELETE FROM training_completions_ingest_run WHERE run_id = ?", (run_id,)
        )
        conn.commit()
        conn.close()
        print(f"  Recovery: restore from {backup} if rollback insufficient.")
        return 1

    print(f"  Post-ingest validation : ✓ all DB sums within ±{VALIDATION_TOLERANCE} of source aggregates")
    if db_within_tolerance:
        max_drift = max(abs(t[2] - t[3]) for t in db_within_tolerance)
        print(f"                          ({len(db_within_tolerance)} aggregates show small drift ≤ ±{max_drift}, NCVER rounding-to-5)")
    conn.close()
    print(f"\n  ✓ Ingest complete. {len(leaf_records)} rows in training_completions (run_id {run_id}).")
    print(f"  ✓ Backup retained at {backup}")
    return 0


def main():
    p = argparse.ArgumentParser(description="Ingest NCVER training-completions XLSX into kintell.db")
    p.add_argument("--dry-run", action="store_true", help="Print plan without mutating DB")
    p.add_argument("--force", action="store_true", help="Delete prior rows from this source and re-ingest")
    p.add_argument("--source-file", type=Path, default=DEFAULT_SOURCE,
                   help=f"Source xlsx (default: {DEFAULT_SOURCE})")
    args = p.parse_args()
    sys.exit(run(dry_run=args.dry_run, force=args.force, source_file=args.source_file))


if __name__ == "__main__":
    main()
