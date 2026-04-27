"""
Layer 2 Step 5b DIAGNOSTIC v2 — Census/ABS subset layout discovery
─────────────────────────────────────────────────────────────────
v2 changes from v1:
  - Probes ALL tables (Table 1, 2, 3) in BOTH workbooks, not just Table 1
  - Confirmed both workbooks are TIME SERIES (2011-2025), not one-time
    Census 2021 snapshots
  - Search patterns expanded to cover median income variants
  - Search patterns for female-specific LFP added

READ-ONLY. No DB mutations. No file writes outside recon/.

Outputs:
  - Console summary
  - recon/layer2_step5b_diag_v2.md
  - recon/layer2_step5b_column_map.json
"""

from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
DB = PROJECT_ROOT / "data" / "kintell.db"
ABS = PROJECT_ROOT / "abs_data"
OUT_DIR = PROJECT_ROOT / "recon"
OUT_MD = OUT_DIR / "layer2_step5b_diag_v2.md"
OUT_JSON = OUT_DIR / "layer2_step5b_column_map.json"

INCOME_FILE = "Income Database.xlsx"
EMPL_FILE = "Education and employment database.xlsx"

# Search patterns — broader than v1
INCOME_PATTERNS = [
    "median",
    "median tot hhd",
    "median total household",
    "median household",
    "median weekly household",
    "median tot fam",
    "median family",
    "median personal",
    "median tot prsnl",
]

LFP_PATTERNS_GENERAL = [
    "labour force participation",
    "participation rate",
    "in the labour force",
]

LFP_PATTERNS_FEMALE = [
    "female",
    "females",
    "women",
]


def banner(title: str) -> str:
    return f"\n{'-' * 70}\n{title}\n{'-' * 70}"


def find_columns_matching(rows: list, n_cols: int,
                          patterns: list[str]) -> list:
    """Search rows 1-9 for column header text matching any pattern."""
    out = []
    for c_idx in range(n_cols):
        full_header = []
        for r_idx in range(min(9, len(rows))):
            v = (rows[r_idx][c_idx]
                 if c_idx < len(rows[r_idx]) else None)
            if v is not None:
                full_header.append(str(v).lower())
        joined = " | ".join(full_header)
        for pat in patterns:
            if pat in joined:
                out.append((c_idx, joined))
                break
    return out


def sample_column_values(wb_path: Path, sheet: str, col_idx: int,
                         data_start_row: int, n: int = 10) -> list:
    """Return up to n non-null sample values from data rows."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[sheet]
    values = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < data_start_row:
            continue
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s in ("-", "..", "np", "n.a.", "na", "NA", ""):
            continue
        try:
            values.append(float(s))
        except (ValueError, TypeError):
            pass
        if len(values) >= n:
            break
    wb.close()
    return values


def probe_table(wb_path: Path, sheet_name: str,
                target_patterns: list[str], label: str) -> dict:
    """Probe one sheet for matching columns + sample values."""
    print(f"\n  --- Sheet: '{sheet_name}' ({label}) ---")

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"    SKIP: sheet does not exist")
        wb.close()
        return {"matches": [], "n_cols": 0}
    ws = wb[sheet_name]

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        if i >= 12:
            break
    n_cols = max(len(r) for r in rows) if rows else 0

    # What's in row 4? (the table title row)
    if len(rows) >= 4:
        title_cell = rows[3][0] if rows[3] else None
        print(f"    Title row 4: {str(title_cell)[:120] if title_cell else '(empty)'}")

    # What's in row 6 col 0-3? (anchor info — the spanning header start)
    print(f"    Row 6 cols 0-3: {[str(c)[:40] if c else '·' for c in rows[5][:4]]}")
    print(f"    Total cols: {n_cols}")

    matches = find_columns_matching(rows, n_cols, target_patterns)
    print(f"    Patterns matched in {len(matches)} column(s)")

    sampled_matches = []
    for c_idx, header in matches[:15]:
        samples = sample_column_values(wb_path, sheet_name, c_idx, 8, n=10)
        result = {
            "col": c_idx,
            "header": header[:200],
            "samples": samples,
            "stats": None,
        }
        if samples:
            avg = sum(samples) / len(samples)
            result["stats"] = {
                "n": len(samples),
                "mean": round(avg, 2),
                "min": round(min(samples), 2),
                "max": round(max(samples), 2),
            }
            print(f"      col[{c_idx:3d}]: n={len(samples)} "
                  f"mean={avg:.2f} min={min(samples):.2f} "
                  f"max={max(samples):.2f}")
            print(f"             samples: {[round(s, 2) for s in samples[:5]]}")
            print(f"             header: {header[:100]}")
        else:
            print(f"      col[{c_idx:3d}]: no numeric samples "
                  f"({header[:80]})")
        sampled_matches.append(result)

    wb.close()
    return {"n_cols": n_cols, "matches": sampled_matches}


def probe_workbook_all_tables(wb_path: Path, label: str,
                              patterns: list[str]) -> dict:
    """Probe every Table N sheet in the workbook."""
    print(banner(f"PROBE: {label}"))
    print(f"  File: {wb_path.name}")

    if not wb_path.exists():
        print(f"  FAIL: file not found")
        return {}

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    print(f"  All sheets: {sheet_names}")

    table_sheets = [sn for sn in sheet_names
                    if sn.lower().strip().startswith("table")]

    out = {"file": wb_path.name, "tables": {}}
    for sn in table_sheets:
        info = probe_table(wb_path, sn, patterns, label)
        out["tables"][sn] = info

    return out


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Probe Income — all tables, broad pattern set
    income_info = probe_workbook_all_tables(
        ABS / INCOME_FILE,
        "INCOME (target: any 'median' income)",
        target_patterns=INCOME_PATTERNS,
    )

    # Probe E&E — first pass: any LFP-related column
    print("\n" + "=" * 70)
    print("EDUCATION & EMPLOYMENT — pass 1 (any LFP/participation column)")
    print("=" * 70)
    empl_info_general = probe_workbook_all_tables(
        ABS / EMPL_FILE,
        "EDUCATION & EMPLOYMENT (general LFP probe)",
        target_patterns=LFP_PATTERNS_GENERAL,
    )

    # Probe E&E — second pass: female-specific
    print("\n" + "=" * 70)
    print("EDUCATION & EMPLOYMENT — pass 2 (female/women columns)")
    print("=" * 70)
    empl_info_female = probe_workbook_all_tables(
        ABS / EMPL_FILE,
        "EDUCATION & EMPLOYMENT (female-specific probe)",
        target_patterns=LFP_PATTERNS_FEMALE,
    )

    # audit_log
    print(banner("AUDIT_LOG + TARGET TABLE"))
    if not DB.exists():
        print(f"FAIL: DB not found at {DB}")
        return 1
    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    cur = conn.cursor()
    cur.execute("SELECT COALESCE(MAX(audit_id), 0) FROM audit_log")
    last_audit = cur.fetchone()[0]
    print(f"Last audit_id: {last_audit} (next: {last_audit + 1})")

    # Check existing target tables
    for tn in ("census2021_sa2", "abs_sa2_socioeconomic_annual"):
        cur.execute("SELECT name FROM sqlite_master "
                    "WHERE type='table' AND name=?", (tn,))
        exists = cur.fetchone() is not None
        if exists:
            cur.execute(f"SELECT COUNT(*) FROM {tn}")
            n = cur.fetchone()[0]
            print(f"  {tn}: EXISTS ({n:,} rows)")
        else:
            print(f"  {tn}: does not exist")
    conn.close()

    # Save findings
    column_map = {
        "schema_note": (
            "Both workbooks are TIME SERIES 2011-2025, not one-shot "
            "Census snapshots. Schema target: long format "
            "(sa2_code, year, metric_name, value)."
        ),
        "income": income_info,
        "empl_general": empl_info_general,
        "empl_female": empl_info_female,
        "null_markers": ["-", "..", "np", "n.a.", "na", "NA", "NP", ""],
        "next_step": (
            "Manual review of probe output to select ONE column for "
            "median income and ONE column for female LFP. If columns "
            "don't exist, source from ABS Cat 6524.0.55.002 instead."
        ),
    }
    OUT_JSON.write_text(json.dumps(column_map, indent=2, default=str),
                        encoding="utf-8")

    findings = ["# Layer 2 Step 5b Diagnostic v2 — All Tables", ""]
    findings.append("```json")
    findings.append(json.dumps(column_map, indent=2, default=str))
    findings.append("```\n")
    OUT_MD.write_text("\n".join(findings), encoding="utf-8")

    print(banner("DONE"))
    print(f"  Markdown: {OUT_MD}")
    print(f"  JSON:     {OUT_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
