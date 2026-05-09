r"""
layer4_4_step_a4_schools_apply.py — A4 Schools at SA2 ingest (v1).

Ingests eight SA2-level school-infrastructure metrics into
`abs_sa2_education_employment_annual` (long-format), aggregating
ACARA's per-school records up to SA2 × year.

Sources (per DEC-82 — direct primary source, not derivative ABS DBR):
  - abs_data/School Location 2025.xlsx (sheet "SchoolLocations 2025")
      - Single 2025 snapshot
      - Provides: lat/lng, ACARA-pre-computed SA2, sector, school_type
      - Used as the canonical (acara_id -> sa2) map for currently-
        operating schools.
  - abs_data/School Profile 2008-2025.xlsx (sheet "SchoolProfile 2008-2025")
      - 18-year time series 2008-2025
      - Provides: sector, type, total enrolments, ICSEA per (school, year)
      - Joined to Location on ACARA SML ID to get SA2.

Caveat (banked in metric copy):
  Pre-2025 closed schools are absent from Location 2025 -> not mapped to
  SA2 -> their historical enrolment is excluded. The 18-year trajectory
  is "currently-operating-schools' enrolment over time" rather than
  "historical actual enrolment in this catchment". Defensible V1
  simplification.

Cross-validation: ACARA's "Statistical Area 2" column is checked against
this codebase's spatial-join (geo_helpers.points_to_sa2) during dry-run.
Halt if mismatch rate >1% per Patrick (sanity-check ratification).

Eight metrics (all banded state_x_remoteness per DEC-67 NES precedent;
3 rendered in V1 via the new "Education infrastructure" card; the other
5 sit in DB ready for V2 / Excel export / group page / downstream):

  1. acara_school_count_total           (count, render-V1)
  2. acara_school_enrolment_total       (count, render-V1)
  3. acara_school_govt_share_pct        (banded share, V2 banked)
  4. acara_school_catholic_share_pct    (banded share, V2 banked)
  5. acara_school_independent_share_pct (banded share, V2 banked)
  6. acara_school_enrolment_govt_share_pct          (banded share, render-V1)
  7. acara_school_enrolment_catholic_share_pct      (banded share, V2 banked)
  8. acara_school_enrolment_independent_share_pct   (banded share, V2 banked)

National 2024-25 references (ACARA published):
  ~10,000 schools nationally
  ~4 million students
  ~64% Govt share by enrolment, ~20% Catholic, ~16% Independent

Usage:
    python layer4_4_step_a4_schools_apply.py             # dry-run (default)
    python layer4_4_step_a4_schools_apply.py --apply     # write
    python layer4_4_step_a4_schools_apply.py --replace --apply

Run from repo root.
"""

import json
import shutil
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

DB_PATH = Path("data") / "kintell.db"
LOCATION_WB = Path("abs_data") / "School Location 2025.xlsx"
PROFILE_WB = Path("abs_data") / "School Profile 2008-2025.xlsx"
TARGET_TABLE = "abs_sa2_education_employment_annual"
ACTOR = "layer4_4_step_a4_schools_apply"

# Sheet + column maps (from probe_a4_acara_files.py 2026-05-10)
LOCATION_SHEET = "SchoolLocations 2025"
PROFILE_SHEET = "SchoolProfile 2008-2025"

LOC_COL = {
    "year": 0, "acara_id": 1, "school_name": 5,
    "state": 7, "sector": 9, "type": 10,
    "lat": 13, "lng": 14, "sa2": 19, "sa2_name": 20,
}
PROF_COL = {
    "year": 0, "acara_id": 1, "school_name": 4,
    "state": 6, "sector": 8, "type": 9, "icsea": 17,
    "enrolment": 27,
}

SECTORS = ("Government", "Catholic", "Independent")
SECTOR_KEY = {"Government": "govt", "Catholic": "catholic",
              "Independent": "independent"}

CROSS_VALIDATION_THRESHOLD = 0.01   # halt if >1% SA2 mismatch ACARA vs sjoin

VERIFY_SA2 = [
    ("211011251", "Bayswater Vic"),
    ("118011341", "Bondi Junction-Waverly NSW"),
    ("506031124", "Bentley-Wilson-St James WA"),
    ("702041063", "Outback NT"),
]

# Per-metric ingest specs
METRICS = [
    ("acara_school_count_total",                       "acara_school_count_total_ingest_v1",
        "School count per SA2 per year. Source: ACARA School Profile 2008-2025."
        " Currently-operating schools only (Profile filtered to ACARA SML IDs"
        " present in School Location 2025)."),
    ("acara_school_enrolment_total",                   "acara_school_enrolment_total_ingest_v1",
        "Total student enrolment per SA2 per year, summed across all schools"
        " in this catchment. Source: ACARA School Profile 'Total Enrolments'."),
    ("acara_school_govt_share_pct",                    "acara_school_govt_share_ingest_v1",
        "Government-sector share of school count per SA2 per year."
        " Numerator: count of Government-sector schools; Denominator: total"
        " school count. Source: ACARA School Profile."),
    ("acara_school_catholic_share_pct",                "acara_school_catholic_share_ingest_v1",
        "Catholic-sector share of school count per SA2 per year."),
    ("acara_school_independent_share_pct",             "acara_school_independent_share_ingest_v1",
        "Independent-sector share of school count per SA2 per year."),
    ("acara_school_enrolment_govt_share_pct",          "acara_school_enrolment_govt_share_ingest_v1",
        "Government-sector share of student enrolment per SA2 per year."
        " Captures the public/private mix in a single banded metric."
        " Source: ACARA School Profile 'Total Enrolments' aggregated by sector."),
    ("acara_school_enrolment_catholic_share_pct",      "acara_school_enrolment_catholic_share_ingest_v1",
        "Catholic-sector share of student enrolment per SA2 per year."),
    ("acara_school_enrolment_independent_share_pct",   "acara_school_enrolment_independent_share_ingest_v1",
        "Independent-sector share of student enrolment per SA2 per year."),
]


def section(title):
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def parse_args():
    return {
        "apply": "--apply" in sys.argv,
        "replace": "--replace" in sys.argv,
        "skip_xval": "--skip-xval" in sys.argv,
    }


# =====================================================================
# Preflight
# =====================================================================
def preflight():
    for p in (DB_PATH, LOCATION_WB, PROFILE_WB):
        if not p.exists():
            print(f"ERROR: {p} not found. Run from repo root.")
            sys.exit(1)
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    tabs = {r[0] for r in cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()}
    for need in ("audit_log", TARGET_TABLE):
        if need not in tabs:
            print(f"ERROR: required table '{need}' missing.")
            con.close()
            sys.exit(1)
    con.close()


# =====================================================================
# Source readers
# =====================================================================
def read_locations():
    """Returns list of dicts, one per school. Filters to rows with valid
    SA2 + lat/lng."""
    import openpyxl
    print(f"  Reading {LOCATION_WB.name}...")
    wb = openpyxl.load_workbook(LOCATION_WB, read_only=True, data_only=True)
    ws = wb[LOCATION_SHEET]
    out = []
    skipped_no_sa2 = 0
    skipped_no_latlng = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) <= max(LOC_COL.values()):
            continue
        acara_id = row[LOC_COL["acara_id"]]
        if acara_id is None:
            continue
        sa2 = row[LOC_COL["sa2"]]
        lat = row[LOC_COL["lat"]]
        lng = row[LOC_COL["lng"]]
        # Filter: valid SA2 (9-digit) + valid lat/lng
        if sa2 is None or not str(sa2).strip().isdigit() or len(str(sa2).strip()) != 9:
            skipped_no_sa2 += 1
            continue
        if lat is None or lng is None:
            skipped_no_latlng += 1
            continue
        out.append({
            "acara_id": str(acara_id).strip(),
            "school_name": str(row[LOC_COL["school_name"]] or ""),
            "state": str(row[LOC_COL["state"]] or ""),
            "sector": str(row[LOC_COL["sector"]] or "").strip(),
            "type": str(row[LOC_COL["type"]] or "").strip(),
            "lat": float(lat), "lng": float(lng),
            "sa2_acara": str(sa2).strip(),
        })
    wb.close()
    print(f"  Locations: {len(out):,} schools  "
          f"(skipped {skipped_no_sa2} no-SA2, {skipped_no_latlng} no-lat/lng)")
    return out


def read_profiles(allowed_ids):
    """Returns list[(acara_id, year, sector, enrolment)] per (school, year).
    Filters to schools in `allowed_ids` (i.e. currently in Location 2025)."""
    import openpyxl
    print(f"  Reading {PROFILE_WB.name}...")
    wb = openpyxl.load_workbook(PROFILE_WB, read_only=True, data_only=True)
    ws = wb[PROFILE_SHEET]
    out = []
    skipped_no_id = 0
    skipped_not_in_loc = 0
    skipped_no_enrol = 0
    skipped_bad_year = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(PROF_COL.values()):
            continue
        acara_id = row[PROF_COL["acara_id"]]
        if acara_id is None:
            skipped_no_id += 1
            continue
        aid = str(acara_id).strip()
        if aid not in allowed_ids:
            skipped_not_in_loc += 1
            continue
        year = row[PROF_COL["year"]]
        try:
            year = int(year)
        except (ValueError, TypeError):
            skipped_bad_year += 1
            continue
        if year < 2008 or year > 2026:
            skipped_bad_year += 1
            continue
        sector = str(row[PROF_COL["sector"]] or "").strip()
        if sector not in SECTORS:
            continue
        enrolment_raw = row[PROF_COL["enrolment"]]
        try:
            enrolment = int(float(enrolment_raw)) if enrolment_raw is not None else None
        except (ValueError, TypeError):
            enrolment = None
        if enrolment is None or enrolment <= 0:
            skipped_no_enrol += 1
            # Note: still count the school in count metrics; just no enrolment contribution
            enrolment = 0
        out.append((aid, year, sector, enrolment))
    wb.close()
    print(f"  Profile rows: {len(out):,}  "
          f"(skipped {skipped_no_id} no-id, {skipped_not_in_loc} not-in-Location-2025, "
          f"{skipped_bad_year} bad-year, {skipped_no_enrol} no-enrolment-zero-filled)")
    years_seen = sorted({r[1] for r in out})
    print(f"  Years covered: {min(years_seen)}-{max(years_seen)} ({len(years_seen)} distinct)")
    return out


# =====================================================================
# Cross-validation: ACARA SA2 vs our sjoin
# =====================================================================
def cross_validate_sa2(locations):
    """Spatial-join lat/lng ourselves; compare to ACARA's pre-computed SA2.
    Halt if mismatch rate >1%."""
    section("Cross-validation: ACARA SA2 vs geo_helpers sjoin")
    try:
        import pandas as pd
        from geo_helpers import points_to_sa2
    except ImportError as e:
        print(f"  WARN: cannot import geo_helpers ({e}); skipping cross-check.")
        return

    df = pd.DataFrame(locations)
    print(f"  Running sjoin on {len(df):,} schools...")
    joined, hit = points_to_sa2(df, lat_col="lat", lng_col="lng",
                                id_col="acara_id")
    if hit < 0.99:
        print(f"  ERROR: sjoin hit-rate {hit:.2%} below 99% threshold.")
        sys.exit(2)

    sjoined_sa2 = joined.set_index("acara_id")["SA2_CODE_2021"]
    matched = 0
    mismatched = []
    no_sjoin = 0
    for loc in locations:
        aid = loc["acara_id"]
        acara_sa2 = loc["sa2_acara"]
        if aid not in sjoined_sa2.index:
            no_sjoin += 1
            continue
        my_sa2 = sjoined_sa2.loc[aid]
        if my_sa2 is None or str(my_sa2) == "nan":
            no_sjoin += 1
            continue
        my_sa2 = str(my_sa2)
        if my_sa2 == acara_sa2:
            matched += 1
        else:
            mismatched.append((aid, loc["school_name"], acara_sa2, my_sa2))

    total_compared = matched + len(mismatched)
    if total_compared == 0:
        print(f"  WARN: nothing to compare; exiting xval check.")
        return
    mismatch_rate = len(mismatched) / total_compared
    print(f"  Matched: {matched:,} / {total_compared:,} ({100*matched/total_compared:.2f}%)")
    print(f"  Mismatched: {len(mismatched):,} ({100*mismatch_rate:.2f}%)")
    print(f"  No sjoin result: {no_sjoin:,}")
    if mismatched:
        print(f"  First 10 mismatches:")
        for aid, name, acara_s, my_s in mismatched[:10]:
            print(f"    {aid}: {name[:40]:<40s} ACARA={acara_s} sjoin={my_s}")
    if mismatch_rate > CROSS_VALIDATION_THRESHOLD:
        print(f"\n  ERROR: mismatch rate {mismatch_rate:.2%} > {CROSS_VALIDATION_THRESHOLD:.2%} threshold.")
        print(f"  Likely cause: ACARA may use a different ASGS edition than ours (2021).")
        print(f"  Use --skip-xval to bypass after manual investigation.")
        sys.exit(3)
    print(f"  PASS: mismatch rate within {CROSS_VALIDATION_THRESHOLD:.2%} threshold.")


# =====================================================================
# Aggregation: per (sa2, year) -> 8 metrics
# =====================================================================
def aggregate(locations, profile_rows):
    """Returns dict[metric_name] -> list[(sa2, year, value)]."""
    # locations: acara_id -> sa2
    aid_to_sa2 = {loc["acara_id"]: loc["sa2_acara"] for loc in locations}

    # For per-(sa2, year) aggregation, we need:
    #   schools_per_sector_year[(sa2, year)] = {"govt": n, "catholic": n, "independent": n, "total": n}
    #   enrolment_per_sector_year[(sa2, year)] = {"govt": n, ...}
    schools = defaultdict(lambda: defaultdict(int))     # (sa2, year) -> sector -> count
    enrol = defaultdict(lambda: defaultdict(int))       # (sa2, year) -> sector -> sum

    for aid, year, sector, enrolment in profile_rows:
        sa2 = aid_to_sa2.get(aid)
        if sa2 is None:
            continue
        sk = SECTOR_KEY.get(sector)
        if sk is None:
            continue
        schools[(sa2, year)][sk] += 1
        schools[(sa2, year)]["total"] += 1
        enrol[(sa2, year)][sk] += enrolment
        enrol[(sa2, year)]["total"] += enrolment

    # Build metric rows
    rows = {m[0]: [] for m in METRICS}
    sa2_years = sorted(schools.keys())
    for (sa2, year) in sa2_years:
        s = schools[(sa2, year)]
        e = enrol[(sa2, year)]
        s_total = s["total"]
        e_total = e["total"]

        rows["acara_school_count_total"].append((sa2, year, s_total))
        rows["acara_school_enrolment_total"].append((sa2, year, e_total))

        if s_total > 0:
            for sector in ("govt", "catholic", "independent"):
                share = round(s.get(sector, 0) / s_total * 100.0, 4)
                metric = f"acara_school_{sector}_share_pct"
                rows[metric].append((sa2, year, share))

        if e_total > 0:
            for sector in ("govt", "catholic", "independent"):
                share = round(e.get(sector, 0) / e_total * 100.0, 4)
                metric = f"acara_school_enrolment_{sector}_share_pct"
                rows[metric].append((sa2, year, share))

    for m in METRICS:
        n = len(rows[m[0]])
        years = sorted({r[1] for r in rows[m[0]]})
        sa2s = len({r[0] for r in rows[m[0]]})
        print(f"  {m[0]:48s}: {n:>7,} rows  {sa2s:>4} SA2s  years {min(years)}-{max(years)}")
    return rows


# =====================================================================
# Sanity checks
# =====================================================================
def sanity_checks(metric_rows, profile_rows):
    section("Sanity checks (national, 2024)")
    target_year = 2024  # 2025 enrolment may be partial — use 2024 as canonical reference
    # National school count + enrolment in 2024
    schools_2024 = sum(1 for r in profile_rows if r[1] == target_year)
    schools_govt = sum(1 for r in profile_rows if r[1] == target_year and r[2] == "Government")
    schools_cath = sum(1 for r in profile_rows if r[1] == target_year and r[2] == "Catholic")
    schools_indp = sum(1 for r in profile_rows if r[1] == target_year and r[2] == "Independent")
    enrol_2024 = sum(r[3] for r in profile_rows if r[1] == target_year)
    enrol_govt = sum(r[3] for r in profile_rows if r[1] == target_year and r[2] == "Government")
    enrol_cath = sum(r[3] for r in profile_rows if r[1] == target_year and r[2] == "Catholic")
    enrol_indp = sum(r[3] for r in profile_rows if r[1] == target_year and r[2] == "Independent")

    print(f"  {target_year} schools: total={schools_2024:,}  "
          f"govt={schools_govt:,} cath={schools_cath:,} indp={schools_indp:,}")
    print(f"  {target_year} enrolment: total={enrol_2024:,}  "
          f"govt={enrol_govt:,} ({100*enrol_govt/enrol_2024 if enrol_2024 else 0:.1f}%)  "
          f"cath={enrol_cath:,} ({100*enrol_cath/enrol_2024 if enrol_2024 else 0:.1f}%)  "
          f"indp={enrol_indp:,} ({100*enrol_indp/enrol_2024 if enrol_2024 else 0:.1f}%)")

    # ACARA-published reference bands
    sane = True
    if not (8000 <= schools_2024 <= 12000):
        print(f"  WARN: schools count {schools_2024:,} outside expected 8k-12k band.")
        sane = False
    if not (3500000 <= enrol_2024 <= 4500000):
        print(f"  WARN: total enrolment {enrol_2024:,} outside expected 3.5M-4.5M band.")
        sane = False
    govt_share_e = 100 * enrol_govt / enrol_2024 if enrol_2024 else 0
    if not (60.0 <= govt_share_e <= 68.0):
        print(f"  WARN: govt enrolment share {govt_share_e:.1f}% outside expected 60-68% band.")
        sane = False
    if sane:
        print(f"  All national totals within expected ABS-published bands.")

    # Verify SA2 spotcheck
    print(f"\n  Verification SA2 spotcheck ({target_year}):")
    by_metric_sa2 = {}
    for metric_name, rows in metric_rows.items():
        for sa2, yr, val in rows:
            if yr == target_year:
                by_metric_sa2[(metric_name, sa2)] = val
    for sa2, name in VERIFY_SA2:
        bits = [f"{sa2} {name}"]
        for m_short, lbl in [
            ("acara_school_count_total", "schools"),
            ("acara_school_enrolment_total", "students"),
            ("acara_school_enrolment_govt_share_pct", "govt%"),
        ]:
            v = by_metric_sa2.get((m_short, sa2))
            if v is None:
                bits.append(f"{lbl}=NA")
            elif "_share_" in m_short:
                bits.append(f"{lbl}={v:.1f}%")
            else:
                bits.append(f"{lbl}={int(v):,}")
        print(f"    " + " | ".join(bits))


# =====================================================================
# DB write helpers (clone of A3+SC pattern)
# =====================================================================
def pre_state_for_metric(con, metric_name):
    cur = con.cursor()
    return cur.execute(
        f'SELECT COUNT(*) FROM "{TARGET_TABLE}" WHERE metric_name = ?',
        (metric_name,),
    ).fetchone()[0]


def write_long_metric(con, metric_name, rows, replace=False):
    cur = con.cursor()
    if replace:
        n = cur.execute(
            f'DELETE FROM "{TARGET_TABLE}" WHERE metric_name = ?',
            (metric_name,),
        ).rowcount
        print(f"  Deleted {n:,} existing rows for metric_name='{metric_name}'")
    cur.executemany(
        f'INSERT INTO "{TARGET_TABLE}" (sa2_code, year, metric_name, value) '
        f"VALUES (?, ?, ?, ?)",
        [(sa2, yr, metric_name, val) for sa2, yr, val in rows],
    )
    return cur.rowcount


def write_audit_log(con, action, before, after, reason):
    cur = con.cursor()
    cur.execute(
        "INSERT INTO audit_log "
        "(actor, action, subject_type, subject_id, before_json, after_json, reason, occurred_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))",
        (ACTOR, action, TARGET_TABLE, 0,
         json.dumps(before), json.dumps(after), reason),
    )
    return cur.execute("SELECT MAX(audit_id) FROM audit_log").fetchone()[0]


def backup_db():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = DB_PATH.parent / f"pre_layer4_4_step_a4_schools_{timestamp}.db"
    print(f"  Backup: copying {DB_PATH} -> {backup_path}")
    shutil.copy2(DB_PATH, backup_path)
    print(f"  Backup size: {backup_path.stat().st_size / (1024 * 1024):.1f} MB")
    return backup_path


# =====================================================================
# Main
# =====================================================================
def main():
    args = parse_args()
    apply_mode = args["apply"]
    replace = args["replace"]
    skip_xval = args["skip_xval"]

    print(f"DB:      {DB_PATH}")
    print(f"LocWB:   {LOCATION_WB}")
    print(f"ProfWB:  {PROFILE_WB}")
    print(f"Mode:    {'APPLY' if apply_mode else 'DRY-RUN'}"
          f"{'  REPLACE' if replace else ''}"
          f"{'  SKIP-XVAL' if skip_xval else ''}")

    preflight()

    section("Reading sources")
    locations = read_locations()
    allowed_ids = {loc["acara_id"] for loc in locations}
    profile_rows = read_profiles(allowed_ids)

    if not skip_xval:
        cross_validate_sa2(locations)

    section("Aggregation: per (sa2, year)")
    metric_rows = aggregate(locations, profile_rows)

    sanity_checks(metric_rows, profile_rows)

    section("Pre-state inventory")
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    max_audit = cur.execute("SELECT COALESCE(MAX(audit_id),0) FROM audit_log").fetchone()[0]
    target_total = cur.execute(f'SELECT COUNT(*) FROM "{TARGET_TABLE}"').fetchone()[0]
    print(f"  audit_log max audit_id: {max_audit:,}")
    print(f"  {TARGET_TABLE} rows:    {target_total:,}")
    pre_existing = {}
    for spec in METRICS:
        m_name = spec[0]
        n = pre_state_for_metric(con, m_name)
        pre_existing[m_name] = n
        print(f"  {m_name}: {n:,} existing rows")

    blockers = [m for m, n in pre_existing.items() if n > 0]
    if blockers and not replace and apply_mode:
        print()
        print("  ERROR: existing rows present and --replace not set:")
        for m in blockers:
            print(f"    - {m}: {pre_existing[m]:,} rows")
        print("  Re-run with --replace --apply to overwrite.")
        con.close()
        sys.exit(4)

    if not apply_mode:
        section("DRY-RUN — no DB mutation")
        total_rows = sum(len(rows) for rows in metric_rows.values())
        for m_name, rows in metric_rows.items():
            print(f"  Would insert {len(rows):,} rows for {m_name}")
        print(f"\n  Total: {total_rows:,} rows; 8 audit_log rows "
              f"(next audit_id starts at {max_audit + 1})")
        print(f"  Would back up DB before write")
        print()
        print("  To proceed: re-run with --apply"
              f"{'' if not replace else ' (replace mode active)'}")
        con.close()
        return

    section("Backup")
    backup_path = backup_db()

    section("Writing")
    audit_ids = {}
    for spec in METRICS:
        m_name, action, reason = spec[0], spec[1], spec[2]
        rows = metric_rows[m_name]
        inserted = write_long_metric(con, m_name, rows, replace=replace)
        print(f"  Wrote {inserted:,} rows for {m_name}")
        years = sorted({r[1] for r in rows})
        sa2_count = len({r[0] for r in rows})
        before = {"existing_rows_for_metric": pre_existing[m_name]}
        after = {
            "rows_inserted": inserted,
            "sa2_count": sa2_count,
            "years": years,
            "year_min": min(years) if years else None,
            "year_max": max(years) if years else None,
            "backup": str(backup_path),
            "replaced": replace,
            "unit": "percentage_0_to_100" if "_pct" in m_name else "count",
        }
        aid = write_audit_log(con, action, before, after, reason)
        audit_ids[m_name] = aid
        print(f"    audit_id={aid}")

    con.commit()

    section("Post-state")
    cur = con.cursor()
    for spec in METRICS:
        m_name = spec[0]
        c, lo, hi, mean = cur.execute(
            f'SELECT COUNT(*), MIN(value), MAX(value), AVG(value) '
            f'FROM "{TARGET_TABLE}" WHERE metric_name = ?',
            (m_name,),
        ).fetchone()
        if c == 0:
            print(f"  {m_name}: 0 rows")
            continue
        print(f"  {m_name}: {c:,} rows  range [{lo:.2f}, {hi:.2f}]  mean {mean:.2f}")

    con.close()
    print()
    print("APPLIED.")
    print()
    print("Next:")
    print("  1. python patch_b2_layer3_add_a4_schools.py --apply")
    print("  2. python layer3_apply.py --apply")
    print("  3. python layer3_x_catchment_metric_banding.py --apply  # OI-35 workaround")


if __name__ == "__main__":
    main()
