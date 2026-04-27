"""
Layer 2 Step 5c APPLY — JSA IVI ingest (ANZSCO 4211 + 2411)
─────────────────────────────────────────────────────────────
Pattern matches Layer 2 steps 1, 2, 4, 5, 6:
  backup -> BEGIN -> CREATE TABLE -> ETL -> audit_log -> invariants -> COMMIT

Two modes:
  (default)  --dry-run : extract + write changeset CSV; no DB mutation
  --apply              : backup + transaction + INSERT + audit + commit

Three tables created (per Step 5c diagnostic findings):

  jsa_ivi_state_monthly (
    state_code     TEXT,       -- 'NSW', 'VIC', ..., 'AUST' (total)
    year_month     TEXT,       -- 'YYYY-MM'
    anzsco_code    TEXT,       -- '4211' or '2411'
    vacancy_count  INTEGER,    -- rounded from source float
    PRIMARY KEY (state_code, year_month, anzsco_code)
  )

  jsa_ivi_remoteness_monthly (
    remoteness     TEXT,       -- 'Major City' / 'Regional' / 'Remote'
    year_month     TEXT,
    anzsco_code    TEXT,
    vacancy_count  INTEGER,
    PRIMARY KEY (remoteness, year_month, anzsco_code)
  )

  jsa_sa4_remoteness_concordance (
    sa4_code       TEXT,
    sa4_name       TEXT,
    remoteness     TEXT,
    northern_aust  INTEGER,    -- 0 or 1
    PRIMARY KEY (sa4_code)
  )

Targets ANZSCO 4211 (Child Carers) and 2411 (Early Childhood
(Pre-primary School) Teachers) only. Other codes excluded — keeps
ingest focused on ECEC-relevant signal.

Source file structure (verified by Step 5c diagnostic):
  STATE FILE: '4 digit 3 month average' sheet
    Header row 1: ANZSCO_CODE | ANZSCO_TITLE | state | <241 months>
    Data starts row 2; 241 monthly cols 2006-03 -> 2026-03
  REMOTENESS FILE: 'JSA Remoteness' sheet
    Header row 1: Level | ANZSCO_CODE | ANZSCO_TITLE | JSA Remoteness | <87 months>
    Data starts row 2; 87 monthly cols 2019-01 -> 2026-03
  REMOTENESS FILE: 'Concordance' sheet
    Header row 1: SA4 Code | SA4 Name | JSA Remoteness | Northern Australia
    Data starts row 2; 1 row per SA4
  Skip 'JSA Northern Australia' sheet (not on V1 path).
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sqlite3
import sys
import time
from datetime import datetime, date
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
ABS = PROJECT_ROOT / "abs_data"
DB = PROJECT_ROOT / "data" / "kintell.db"
OUT_DIR = PROJECT_ROOT / "recon"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
BACKUP_PATH = PROJECT_ROOT / "data" / f"kintell.db.backup_pre_step5c_{TS}"
CHANGESET_STATE = OUT_DIR / f"layer2_step5c_state_changeset_{TS}.csv"
CHANGESET_REMOTE = OUT_DIR / f"layer2_step5c_remote_changeset_{TS}.csv"
CHANGESET_CONC = OUT_DIR / f"layer2_step5c_concordance_changeset_{TS}.csv"

STATE_PREFIX = "internet_vacancies_anzsco4_occupations_states"
REMOTE_PREFIX = "internet_vacancies_anzsco4_occupations_jsa_remoteness"
TARGET_ANZSCO = {"4211", "2411"}

ACTOR = "layer2_step5c_apply"
ACTION = "jsa_ivi_ingest_v1"
SUBJECT_TYPE = "jsa_ivi_state_monthly"
SUBJECT_ID = 0


def banner(title: str) -> str:
    return f"\n{'-' * 70}\n{title}\n{'-' * 70}"


def find_workbook(prefix: str) -> Path | None:
    matches = [p for p in ABS.glob("*.xlsx")
               if p.name.lower().startswith(prefix.lower())]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def to_year_month(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    if isinstance(v, date):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    return None


def to_int_or_none(v) -> int | None:
    """JSA stores some cells as floats, some as '.' for suppressed."""
    if v is None:
        return None
    s = str(v).strip()
    if s in (".", "", "-", "..", "np", "n.a.", "na", "NA", "NP"):
        return None
    try:
        return int(round(float(s)))
    except (ValueError, TypeError):
        return None


def extract_state(wb_path: Path) -> tuple[list, dict]:
    """Returns (rows, meta) for the state IVI file."""
    print(f"  Reading: {wb_path.name}")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["4 digit 3 month average"]

    rows_out = []
    month_cols: dict[int, str] = {}

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Header row 1
        if r_idx == 1:
            for c_idx, cell in enumerate(row):
                if c_idx < 3:
                    continue
                ym = to_year_month(cell)
                if ym:
                    month_cols[c_idx] = ym
            continue

        # Data rows
        if len(row) < 4:
            continue
        anzsco = str(row[0]).strip() if row[0] is not None else ""
        if anzsco not in TARGET_ANZSCO:
            continue
        state = str(row[2]).strip() if row[2] is not None else None
        if state is None:
            continue

        for col_idx, ym in month_cols.items():
            if col_idx >= len(row):
                continue
            val = to_int_or_none(row[col_idx])
            rows_out.append((state, ym, anzsco, val))

    wb.close()
    return rows_out, {
        "month_count": len(month_cols),
        "first_month": min(month_cols.values()) if month_cols else None,
        "last_month": max(month_cols.values()) if month_cols else None,
    }


def extract_remoteness(wb_path: Path) -> tuple[list, dict]:
    print(f"  Reading: {wb_path.name}")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["JSA Remoteness"]

    rows_out = []
    month_cols: dict[int, str] = {}

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            for c_idx, cell in enumerate(row):
                if c_idx < 4:
                    continue
                ym = to_year_month(cell)
                if ym:
                    month_cols[c_idx] = ym
            continue

        if len(row) < 5:
            continue
        anzsco = str(row[1]).strip() if row[1] is not None else ""
        if anzsco not in TARGET_ANZSCO:
            continue
        remoteness = str(row[3]).strip() if row[3] is not None else None
        if remoteness is None:
            continue

        for col_idx, ym in month_cols.items():
            if col_idx >= len(row):
                continue
            val = to_int_or_none(row[col_idx])
            rows_out.append((remoteness, ym, anzsco, val))

    wb.close()
    return rows_out, {
        "month_count": len(month_cols),
        "first_month": min(month_cols.values()) if month_cols else None,
        "last_month": max(month_cols.values()) if month_cols else None,
    }


def extract_concordance(wb_path: Path) -> list:
    print(f"  Reading: {wb_path.name} (Concordance sheet)")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["Concordance"]

    rows_out = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx == 1:
            continue  # header
        if len(row) < 4:
            continue
        sa4_code = str(row[0]).strip() if row[0] is not None else None
        sa4_name = str(row[1]).strip() if row[1] is not None else None
        remoteness = str(row[2]).strip() if row[2] is not None else None
        northern = str(row[3]).strip().lower() if row[3] is not None else ""
        northern_int = 1 if northern in ("yes", "y", "1", "true") else 0
        if sa4_code:
            rows_out.append((sa4_code, sa4_name, remoteness, northern_int))

    wb.close()
    return rows_out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true",
                        help="Mutate the DB. Without flag, dry-run only.")
    args = parser.parse_args()

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(banner(f"LAYER 2 STEP 5C — JSA IVI INGEST ({mode})"))
    print(f"DB:        {DB}")
    print(f"Timestamp: {TS}")

    if not DB.exists():
        print("FAIL: DB not found")
        return 1

    state_path = find_workbook(STATE_PREFIX)
    remote_path = find_workbook(REMOTE_PREFIX)
    if not state_path:
        print(f"FAIL: state IVI workbook not found")
        return 1
    if not remote_path:
        print(f"FAIL: remoteness IVI workbook not found")
        return 1
    print(f"State:      {state_path.name}")
    print(f"Remoteness: {remote_path.name}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ---- EXTRACT ---------------------------------------------------
    print(banner("EXTRACT"))
    t0 = time.time()
    state_rows, state_meta = extract_state(state_path)
    remote_rows, remote_meta = extract_remoteness(remote_path)
    conc_rows = extract_concordance(remote_path)
    elapsed = time.time() - t0

    print(f"\nStats:")
    print(f"  State rows:        {len(state_rows):,}")
    print(f"    months covered:  {state_meta['month_count']} "
          f"({state_meta['first_month']} -> {state_meta['last_month']})")
    print(f"    distinct states: {len(set(r[0] for r in state_rows))}")
    print(f"    non-null:        "
          f"{sum(1 for r in state_rows if r[3] is not None):,}")

    print(f"  Remoteness rows:   {len(remote_rows):,}")
    print(f"    months covered:  {remote_meta['month_count']} "
          f"({remote_meta['first_month']} -> {remote_meta['last_month']})")
    print(f"    distinct cats:   "
          f"{len(set(r[0] for r in remote_rows))}")
    print(f"    non-null:        "
          f"{sum(1 for r in remote_rows if r[3] is not None):,}")

    print(f"  Concordance rows:  {len(conc_rows):,}")
    print(f"  Extract time:      {elapsed:.1f}s")

    # ---- Sample ---------------------------------------------------
    print("\nState sample (first 5):")
    for r in state_rows[:5]:
        print(f"  {r}")

    print("\nRemoteness sample (first 5):")
    for r in remote_rows[:5]:
        print(f"  {r}")

    print("\nConcordance sample (first 5):")
    for r in conc_rows[:5]:
        print(f"  {r}")

    # ---- Sanity: latest month per category for ANZSCO 4211 -------
    print(banner("SANITY CHECK: latest-month vacancy snapshot for 4211"))
    latest_month = state_meta["last_month"]
    print(f"  Latest month: {latest_month}")
    print(f"  State vacancies for 4211 in {latest_month}:")
    for r in state_rows:
        if r[1] == latest_month and r[2] == "4211":
            print(f"    {r[0]:6s}  {r[3]}")
    print(f"  Remoteness vacancies for 4211 in {latest_month}:")
    for r in remote_rows:
        if r[1] == latest_month and r[2] == "4211":
            print(f"    {r[0]:25s}  {r[3]}")

    # ---- Changesets ----------------------------------------------
    print(banner("CHANGESETS"))
    with open(CHANGESET_STATE, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["state_code", "year_month", "anzsco_code", "vacancy_count"])
        w.writerows(state_rows)
    with open(CHANGESET_REMOTE, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["remoteness", "year_month", "anzsco_code", "vacancy_count"])
        w.writerows(remote_rows)
    with open(CHANGESET_CONC, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sa4_code", "sa4_name", "remoteness", "northern_aust"])
        w.writerows(conc_rows)
    print(f"  State:       {CHANGESET_STATE.name}")
    print(f"  Remoteness:  {CHANGESET_REMOTE.name}")
    print(f"  Concordance: {CHANGESET_CONC.name}")

    if not args.apply:
        print(banner("DRY-RUN COMPLETE"))
        print("No DB mutation. Review sanity check above.")
        print("Re-run with --apply when ready:")
        print("  python layer2_step5c_apply.py --apply")
        return 0

    # ================================================================
    # APPLY MODE
    # ================================================================
    print(banner("BACKUP"))
    shutil.copy2(DB, BACKUP_PATH)
    print(f"OK: backed up to {BACKUP_PATH}")
    print(f"    size: {BACKUP_PATH.stat().st_size:,} bytes")

    conn = sqlite3.connect(DB)
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    try:
        print(banner("BEGIN TRANSACTION"))
        cur.execute("BEGIN TRANSACTION")

        # CREATE TABLES
        cur.execute("""
            CREATE TABLE IF NOT EXISTS jsa_ivi_state_monthly (
                state_code     TEXT NOT NULL,
                year_month     TEXT NOT NULL,
                anzsco_code    TEXT NOT NULL,
                vacancy_count  INTEGER,
                PRIMARY KEY (state_code, year_month, anzsco_code)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS ix_ivi_state_anzsco "
                    "ON jsa_ivi_state_monthly(anzsco_code)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_ivi_state_ym "
                    "ON jsa_ivi_state_monthly(year_month)")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS jsa_ivi_remoteness_monthly (
                remoteness     TEXT NOT NULL,
                year_month     TEXT NOT NULL,
                anzsco_code    TEXT NOT NULL,
                vacancy_count  INTEGER,
                PRIMARY KEY (remoteness, year_month, anzsco_code)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS ix_ivi_remote_anzsco "
                    "ON jsa_ivi_remoteness_monthly(anzsco_code)")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS jsa_sa4_remoteness_concordance (
                sa4_code       TEXT NOT NULL,
                sa4_name       TEXT,
                remoteness     TEXT,
                northern_aust  INTEGER,
                PRIMARY KEY (sa4_code)
            )
        """)
        print("CREATE TABLES / INDEXES: OK")

        # Pre-invariants
        for table in ("jsa_ivi_state_monthly",
                      "jsa_ivi_remoteness_monthly",
                      "jsa_sa4_remoteness_concordance"):
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            n = cur.fetchone()[0]
            if n > 0:
                print(f"ABORT: {table} is not empty ({n} rows).")
                conn.execute("ROLLBACK")
                return 1

        # INSERT all three
        print(banner("INSERT"))
        cur.executemany(
            "INSERT INTO jsa_ivi_state_monthly "
            "(state_code, year_month, anzsco_code, vacancy_count) "
            "VALUES (?, ?, ?, ?)", state_rows
        )
        print(f"  state:       {len(state_rows):,} rows")

        cur.executemany(
            "INSERT INTO jsa_ivi_remoteness_monthly "
            "(remoteness, year_month, anzsco_code, vacancy_count) "
            "VALUES (?, ?, ?, ?)", remote_rows
        )
        print(f"  remoteness:  {len(remote_rows):,} rows")

        cur.executemany(
            "INSERT INTO jsa_sa4_remoteness_concordance "
            "(sa4_code, sa4_name, remoteness, northern_aust) "
            "VALUES (?, ?, ?, ?)", conc_rows
        )
        print(f"  concordance: {len(conc_rows):,} rows")

        # Post-invariants
        cur.execute("SELECT COUNT(*) FROM jsa_ivi_state_monthly")
        post_state = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM jsa_ivi_remoteness_monthly")
        post_remote = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM jsa_sa4_remoteness_concordance")
        post_conc = cur.fetchone()[0]

        print(banner("POST-INVARIANTS"))
        print(f"  state rows:       {post_state:,}")
        print(f"  remoteness rows:  {post_remote:,}")
        print(f"  concordance rows: {post_conc:,}")

        if (post_state != len(state_rows) or
                post_remote != len(remote_rows) or
                post_conc != len(conc_rows)):
            print("ABORT: post-counts don't match extracted")
            conn.execute("ROLLBACK")
            return 1

        # Audit log
        print(banner("AUDIT LOG"))
        payload = {
            "state_rows_inserted": post_state,
            "remoteness_rows_inserted": post_remote,
            "concordance_rows_inserted": post_conc,
            "target_anzsco_codes": list(TARGET_ANZSCO),
            "state_months": state_meta["month_count"],
            "state_first_month": state_meta["first_month"],
            "state_last_month": state_meta["last_month"],
            "remote_months": remote_meta["month_count"],
            "remote_first_month": remote_meta["first_month"],
            "remote_last_month": remote_meta["last_month"],
            "state_source_file": state_path.name,
            "remote_source_file": remote_path.name,
            "backup_path": BACKUP_PATH.name,
        }
        cur.execute("""
            INSERT INTO audit_log (
                actor, action, subject_type, subject_id,
                before_json, after_json, reason, occurred_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            ACTOR, ACTION, SUBJECT_TYPE, SUBJECT_ID,
            json.dumps({"state": 0, "remoteness": 0, "concordance": 0}),
            json.dumps({"state": post_state, "remoteness": post_remote,
                        "concordance": post_conc, "payload": payload}),
            "Layer 2 Step 5c: JSA IVI ingest "
            "(ANZSCO 4211 + 2411, state + remoteness)",
            datetime.now().isoformat(),
        ))
        cur.execute("SELECT MAX(audit_id) FROM audit_log")
        new_audit_id = cur.fetchone()[0]
        print(f"audit_log row: id={new_audit_id} action='{ACTION}'")

        conn.commit()
        print(banner("COMMIT SUCCESSFUL"))
        print(f"  jsa_ivi_state_monthly:           {post_state:,} rows")
        print(f"  jsa_ivi_remoteness_monthly:      {post_remote:,} rows")
        print(f"  jsa_sa4_remoteness_concordance:  {post_conc:,} rows")
        print(f"  Backup:    {BACKUP_PATH.name}")
        print(f"  audit_id:  {new_audit_id}")
        return 0

    except Exception as exc:
        print(banner(f"FAIL: {type(exc).__name__}"))
        print(str(exc))
        try:
            conn.execute("ROLLBACK")
            print("ROLLBACK successful - DB unchanged.")
        except Exception:
            pass
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
