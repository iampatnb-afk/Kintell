# layer1_recheck.py
# Re-inspect Population and Economic-and-Industry workbooks after replacement.
# Looks for SA2-level data in Table 1, under-5 age columns, and confirms openpyxl can read.
import os, sys
from openpyxl import load_workbook

ABS = "abs_data"
TARGETS = [
    ("Population and People Database.xlsx", ["under 5", "0-4", "0 to 4", "aged 0"]),
    ("Economic and Industry Database.xlsx", []),
]

def dump_sheet(wb, sheet_name, max_rows=15, max_cols=18, narrow_text=True):
    ws = wb[sheet_name]
    print(f"\n--- sheet: `{sheet_name}` (dim: {ws.max_row} rows x {ws.max_column} cols) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = list(row)[:max_cols]
        if narrow_text:
            cells = [(str(c)[:60] if c is not None else "") for c in cells]
        print(f"  row {i:2d}: {cells}")

for fn, age_keywords in TARGETS:
    p = os.path.join(ABS, fn)
    print(f"\n=========================================")
    print(f"=== {fn}")
    print(f"=========================================")
    if not os.path.exists(p):
        print(f"  MISSING")
        continue
    print(f"  size: {os.path.getsize(p):,} bytes")
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        print(f"  sheet count: {len(wb.sheetnames)}")
        for s in wb.sheetnames:
            print(f"    - {s}")
        # Dump every non-Contents sheet
        for s in wb.sheetnames:
            if "content" in s.lower():
                continue
            dump_sheet(wb, s, max_rows=12, max_cols=20)
        # For Population: scan all sheets for under-5 age columns
        if age_keywords:
            print(f"\n  [searching for under-5 cohort columns]")
            for s in wb.sheetnames:
                if "content" in s.lower():
                    continue
                ws = wb[s]
                hits = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i > 20:
                        break
                    for c in row:
                        if c is None:
                            continue
                        cs = str(c).lower()
                        for kw in age_keywords:
                            if kw in cs:
                                hits.append((i, str(c)[:80]))
                                break
                if hits:
                    print(f"    sheet `{s}`: {len(hits)} hits")
                    for h in hits[:8]:
                        print(f"      row {h[0]}: {h[1]}")
        wb.close()
    except Exception as e:
        print(f"  ERROR: {e}")
