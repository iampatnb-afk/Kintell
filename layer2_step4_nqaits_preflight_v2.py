# layer2_step4_nqaits_preflight_v2.py — read-only
# v2: canonical column resolution covering all 50 sheets (incl 17 older sheets
#     using "Service Approval Number" instead of "Service ID").
import os, sys, sqlite3, datetime, re
from collections import Counter, defaultdict

NQAITS = os.path.join("abs_data", "NQAITS Quarterly Data Splits (Q3 2013 - Q4 2025).xlsx")
DB = os.path.join("data", "kintell.db")
OUT = os.path.join("recon", "layer2_step4_nqaits_preflight_v2.md")

COL_VARIANTS = {
    "service_id":               ["Service ID", "Service Approval Number"],
    "service_name":             ["Service Name"],
    "provider_id":              ["Provider ID"],
    "provider_name":            ["Provider Name"],
    "provider_management_type": ["Provider Management Type"],
    "managing_jurisdiction":    ["Managing Jurisdiction", "Managing Jurisdiction "],
    "service_type":             ["Service Type"],
    "service_sub_type":         ["Service sub-type (ordered counting method)", "Service Sub Type"],
    "approval_date":            ["Approval Date", "ApprovalDate"],
    "seifa":                    ["SEIFA"],
    "aria":                     ["ARIA", "ARIA+"],
    "max_places":               ["Maximum total places"],
    "nqs_version":              ["NQS Version"],
    "final_report_sent_date":   ["Final Report Sent Date"],
    "overall_rating":           ["Overall Rating", "OverallRating"],
    "quality_area_1":           ["Quality Area 1", "Q1"],
    "quality_area_2":           ["Quality Area 2", "Q2"],
    "quality_area_3":           ["Quality Area 3", "Q3"],
    "quality_area_4":           ["Quality Area 4", "Q4"],
    "quality_area_5":           ["Quality Area 5", "Q5"],
    "quality_area_6":           ["Quality Area 6", "Q6"],
    "quality_area_7":           ["Quality Area 7", "Q7"],
    "long_day_care":            ["Long Day Care"],
    "preschool_standalone":     ["PreschoolKindergarten Stand Alone", "Preschool/Kindergarten Stand Alone", "Preschool/\nKindergarten Stand Alone"],
    "preschool_in_school":      ["PreschoolKindergarten Part of a School", "Preschool/\nKindergarten Part of a School"],
    "oshc_before_school":       ["OSHC BeforeSchool", "OSHC Before School"],
    "oshc_after_school":        ["OSHC After School"],
    "oshc_vacation_care":       ["OSHC Vacation Care"],
    "nature_care_other":        ["Nature Care Other"],
    "postcode":                 ["Postcode"],
    "latitude":                 ["Latitude"],
    "longitude":                ["Longitude"],
}

def resolve_columns(header_row):
    h = [str(c).strip() if c is not None else None for c in header_row]
    out = {}
    for canon, variants in COL_VARIANTS.items():
        for v in variants:
            if v in h:
                out[canon] = h.index(v)
                break
    return out, h

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("FATAL: pip install openpyxl")

lines = []
def emit(s=""):
    print(s, flush=True)
    lines.append(s)

emit("# Phase 2.5 Layer 2 Step 4 — NQAITS Preflight v2")
emit(f"_Generated: {datetime.datetime.now().isoformat(timespec='seconds')}_")
emit("")

print("[opening NQAITS — 138MB, ~30s ...]", flush=True)
wb = load_workbook(NQAITS, read_only=True, data_only=True)
data_sheets = [s for s in wb.sheetnames if re.match(r"Q\d20\d\ddata", s)]

# Per-sheet column resolution
print(f"[resolving columns in {len(data_sheets)} sheets ...]", flush=True)
sheet_cols = {}     # sheet -> {canon: idx}
sheet_headers = {}  # sheet -> raw header list
unresolved_per_sheet = {}
for s in data_sheets:
    ws = wb[s]
    rows_iter = ws.iter_rows(values_only=True)
    h = next(rows_iter)
    cols, hraw = resolve_columns(h)
    sheet_cols[s] = cols
    sheet_headers[s] = hraw
    missing = [c for c in COL_VARIANTS if c not in cols]
    if missing:
        unresolved_per_sheet[s] = missing

emit("## Canonical column resolution per sheet")
emit(f"  total canonical fields wanted: {len(COL_VARIANTS)}")
fully_resolved = sum(1 for s in data_sheets if not unresolved_per_sheet.get(s))
emit(f"  sheets with ALL canonical fields resolved: {fully_resolved}/{len(data_sheets)}")
if unresolved_per_sheet:
    emit("  sheets missing at least one canonical field:")
    miss_counter = Counter()
    for s, miss in unresolved_per_sheet.items():
        emit(f"    {s}: missing {miss}")
        for m in miss:
            miss_counter[m] += 1
    emit("")
    emit(f"  fields missing in N sheets:")
    for m, n in miss_counter.most_common():
        emit(f"    {m}: missing in {n} sheets")
emit("")

# Verify Service ID resolution covers all 50 sheets
sid_resolved = sum(1 for s in data_sheets if "service_id" in sheet_cols[s])
emit(f"## service_id resolution: {sid_resolved}/{len(data_sheets)} sheets")
if sid_resolved != len(data_sheets):
    emit("  WARN: not all sheets have service_id resolved — variant map incomplete")
emit("")

# Now scan all sheets for sid x provider_id
print("[scanning all 50 sheets for sid x provider_id ...]", flush=True)
sid_quarters = defaultdict(set)
sid_provider_per_quarter = defaultdict(dict)
for idx, s in enumerate(data_sheets):
    if idx % 10 == 0:
        print(f"  sheet {idx+1}/{len(data_sheets)}: {s}", flush=True)
    cols = sheet_cols[s]
    sid_idx = cols.get("service_id")
    pid_idx = cols.get("provider_id")
    if sid_idx is None:
        continue
    ws = wb[s]
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)
    for row in rows_iter:
        if not row or row[sid_idx] is None:
            continue
        sid = str(row[sid_idx]).strip()
        if not sid:
            continue
        sid_quarters[sid].add(s)
        if pid_idx is not None and pid_idx < len(row) and row[pid_idx] is not None:
            sid_provider_per_quarter[sid][s] = str(row[pid_idx]).strip()
wb.close()

emit("## Cross-quarter coverage (all 50 sheets)")
emit(f"  unique service_ids across all quarters: {len(sid_quarters):,}")
qcounts = Counter(len(qs) for qs in sid_quarters.values())
emit(f"  distribution of services by quarter-count:")
for q, n in sorted(qcounts.items(), reverse=True)[:15]:
    emit(f"    {q:>3} quarters: {n:>6} services")
all_q = set(data_sheets)
in_all = [sid for sid, qs in sid_quarters.items() if qs == all_q]
emit(f"  services present in ALL 50 quarters: {len(in_all):,}")
emit("")

# PA chain analysis
emit("## PA chain analysis (50-sheet coverage)")
chain_changes = 0
chain_2 = 0
chain_3plus = 0
for sid, qprov in sid_provider_per_quarter.items():
    n = len(set(qprov.values()))
    if n > 1:
        chain_changes += 1
        if n == 2:
            chain_2 += 1
        else:
            chain_3plus += 1
emit(f"  services with >1 distinct provider_id: {chain_changes:,} ({chain_changes/len(sid_quarters)*100:.1f}%)")
emit(f"    2 distinct providers: {chain_2:,}")
emit(f"    3+ distinct providers: {chain_3plus:,}")
emit("")

# Sample chain changers
emit("  sample chain changers (5):")
samples = []
for sid, qprov in sid_provider_per_quarter.items():
    if len(set(qprov.values())) > 1:
        samples.append((sid, qprov))
        if len(samples) >= 5:
            break
for sid, qprov in samples:
    pids = set(qprov.values())
    emit(f"    {sid} touched {len(pids)} provider(s)")
    for q, p in sorted(qprov.items(), reverse=True)[:8]:
        emit(f"      {q} -> {p}")
emit("")

# Join check vs DB
emit("## Join check: NQAITS service_id vs services.service_approval_number")
conn = sqlite3.connect(DB)
cur = conn.cursor()
cur.execute("SELECT service_approval_number, is_active FROM services WHERE service_approval_number IS NOT NULL;")
db_rows = cur.fetchall()
db_active = {r[0] for r in db_rows if r[1] == 1}
db_inactive = {r[0] for r in db_rows if r[1] != 1}
db_all = db_active | db_inactive
conn.close()

nq = set(sid_quarters.keys())
emit(f"  DB total                 : {len(db_all):,}")
emit(f"  DB active                : {len(db_active):,}")
emit(f"  DB inactive              : {len(db_inactive):,}")
emit(f"  NQAITS unique sids       : {len(nq):,}")
emit(f"  matched (DB ∩ NQAITS)    : {len(db_all & nq):,}  ({len(db_all & nq)/len(db_all)*100:.2f}% of DB)")
emit(f"    of DB active matched   : {len(db_active & nq):,}  ({len(db_active & nq)/len(db_active)*100:.2f}% of DB active)")
emit(f"    of DB inactive matched : {len(db_inactive & nq):,}")
emit(f"  in DB but not NQAITS     : {len(db_all - nq):,}")
emit(f"  in NQAITS but not DB     : {len(nq - db_all):,}")
emit("")

emit("  sample DB-active not in NQAITS (5):")
for s in sorted(db_active - nq)[:5]:
    emit(f"    {s!r}")
emit("  sample NQAITS not in DB (5):")
for s in sorted(nq - db_all)[:5]:
    emit(f"    {s!r}")
emit("")

# Volume estimate
total = sum(len(qs) for qs in sid_quarters.values())
emit("## Ingest volume estimate")
emit(f"  total (sid, quarter) pairs across all sheets: {total:,}")
emit(f"  avg quarters per service: {total / len(sid_quarters):.1f}")
emit("")

os.makedirs("recon", exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"\n[wrote {OUT}]")
