"""
abs_dbr_remoteness_probe.py

Read-only probe of all ABS Data by Region xlsx files in abs_data/
to determine if any of them carry an SA2-level remoteness/ARIA column
we didn't ingest in Layer 2.

For each Database.xlsx file:
  - List sheets
  - Inspect rows 1-10 for ARIA / Remoteness header text
  - Sample one SA2-level data row (~row 80, past AUS/State aggregates)
  - Report header positions + any matching column values

Output: stdout summary. No file writes.

Usage:
  python abs_dbr_remoteness_probe.py
"""
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.")
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
ABS_DIR = REPO_ROOT / "abs_data"

CANDIDATE_FILES = [
    "Population and People Database.xlsx",
    "Income Database.xlsx",
    "Education and employment database.xlsx",
    "Family and Community Database.xlsx",
    "Economic and Industry Database.xlsx",
]

KEYWORDS = ["aria", "remote", "remoteness"]
MAX_ROWS_TO_SCAN = 10
SAMPLE_DATA_ROW = 80  # SA2-level rows typically start ~row 75 (Std 25)


def is_match(cell):
    if cell is None:
        return False
    if not isinstance(cell, str):
        return False
    low = cell.lower()
    return any(kw in low for kw in KEYWORDS)


def probe_file(path: Path):
    print(f"=== {path.name} ===")
    if not path.exists():
        print("  [FILE NOT FOUND]")
        print()
        return None

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ERROR opening: {e}")
        print()
        return None

    findings_for_file = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_hits = []
        rows_buf = []
        for ridx, row in enumerate(ws.iter_rows(values_only=True)):
            rows_buf.append(row)
            if ridx >= MAX_ROWS_TO_SCAN - 1:
                break

        for ridx, row in enumerate(rows_buf):
            if not row:
                continue
            for cidx, cell in enumerate(row):
                if is_match(cell):
                    sheet_hits.append((ridx, cidx, cell))

        if sheet_hits:
            print(f"  Sheet '{sheet_name}': {len(sheet_hits)} keyword hit(s)")
            for ridx, cidx, cell in sheet_hits[:10]:
                snippet = str(cell)[:80]
                print(f"    row {ridx} col {cidx}: {snippet!r}")
            if len(sheet_hits) > 10:
                print(f"    ... +{len(sheet_hits)-10} more")

            # Try to sample one SA2 data row at the matching columns
            cols_to_sample = list({c for _, c, _ in sheet_hits})[:5]
            try:
                ws2 = wb[sheet_name]
                row_at = None
                for ridx, row in enumerate(ws2.iter_rows(values_only=True)):
                    if ridx == SAMPLE_DATA_ROW:
                        row_at = row
                        break
                if row_at is not None:
                    print(f"    Sample row {SAMPLE_DATA_ROW} at "
                          f"matching cols:")
                    for c in cols_to_sample:
                        if c < len(row_at):
                            v = str(row_at[c])[:60]
                            print(f"      col {c}: {v!r}")
            except Exception as e:
                print(f"    (sample row probe failed: {e})")

            findings_for_file[sheet_name] = sheet_hits

    if not findings_for_file:
        print("  No ARIA / Remote / Remoteness keywords in first "
              f"{MAX_ROWS_TO_SCAN} rows of any sheet.")

    print()
    wb.close()
    return findings_for_file


def main():
    print("ABS Data by Region remoteness probe - read-only")
    print(f"Scanning: {ABS_DIR}")
    print(f"Keywords: {KEYWORDS}")
    print()

    results = {}
    for fname in CANDIDATE_FILES:
        path = ABS_DIR / fname
        results[fname] = probe_file(path)

    print("=" * 64)
    any_hit = any(v for v in results.values() if v)
    if any_hit:
        print("OUTCOME: remoteness/ARIA keywords FOUND in:")
        for fname, found in results.items():
            if found:
                sheet_list = ", ".join(found.keys())
                print(f"  - {fname} -> sheets: {sheet_list}")
    else:
        print("OUTCOME: no remoteness/ARIA columns in any Database.xlsx.")
        print("Need separate ABS Remoteness Areas data source.")
    print("=" * 64)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
