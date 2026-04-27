"""
Layer 2 Step 5 APPLY — SALM SA2 unemployment ingest
─────────────────────────────────────────────────────────
Pattern matches Layer 2 steps 1, 2, 4, 6:
  backup -> BEGIN -> CREATE TABLE -> ETL -> audit_log -> invariants -> COMMIT

Two modes:
  (default)  --dry-run : extract + write changeset CSV; no DB mutation
  --apply              : backup + transaction + INSERT + audit + commit

Schema (long on time, wide on metric):
  abs_sa2_unemployment_quarterly (
    sa2_code      TEXT,
    year_qtr      TEXT,        -- e.g. "2025-Q4"
    rate          REAL,        -- smoothed_unemployment_rate (%)
    count         INTEGER,     -- smoothed_unemployment_count
    labour_force  INTEGER,     -- smoothed_labour_force
    PRIMARY KEY (sa2_code, year_qtr)
  )

Why this shape (not full long): rate/count/LF are tightly coupled
(rate ~= count/LF * 100). Storing per-row enables a one-row sanity
check and a single-row query for centre-page rendering.

Workbook structure (verified by Step 5 diagnostic):
  - 3 sheets: 'Smoothed SA2 unemployment rate',
              'Smoothed SA2 unemployment',
              'Smoothed SA2 labour force'
  - All sheets identical structure: 63 cols, header row 4, data
    starts row 5, SA2 name col 0, SA2 code col 1, quarters cols 2-62
  - Quarter headers are Excel datetime objects (yyyy-mm-01)
  - Suppression marker: '-'
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sqlite3
import sys
import time
from datetime import datetime, date
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
ABS = PROJECT_ROOT / "abs_data"
DB = PROJECT_ROOT / "data" / "kintell.db"
OUT_DIR = PROJECT_ROOT / "recon"
TS = datetime.now().strftime("%Y%m%d_%H%M%S")
BACKUP_PATH = PROJECT_ROOT / "data" / f"kintell.db.backup_pre_step5_{TS}"
CHANGESET_PATH = OUT_DIR / f"layer2_step5_changeset_{TS}.csv"

# Workbook structure (verified by diagnostic)
SA2_PREFIX = "SALM Smoothed SA2"
HEADER_ROW = 4
DATA_START_ROW = 5
SA2_NAME_COL = 0
SA2_CODE_COL = 1
FIRST_QUARTER_COL = 2
NULL_MARKERS = {"-", "..", "np", "n.a.", "na", "NA", "NP", "—", ""}

SHEET_FOR_METRIC = {
    "rate": "Smoothed SA2 unemployment rate",
    "count": "Smoothed SA2 unemployment",
    "labour_force": "Smoothed SA2 labour force",
}

ACTOR = "layer2_step5_apply"
ACTION = "salm_sa2_ingest_v1"
SUBJECT_TYPE = "abs_sa2_unemployment_quarterly"
SUBJECT_ID = 0


def banner(title: str) -> str:
    return f"\n{'-' * 70}\n{title}\n{'-' * 70}"


def find_workbook(prefix: str) -> Path | None:
    matches = [p for p in ABS.glob("*.xlsx")
               if p.name.startswith(prefix)]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def datetime_to_year_qtr(dt) -> str | None:
    """Convert an Excel datetime cell to 'YYYY-Q[1-4]'."""
    if dt is None:
        return None
    if isinstance(dt, str):
        # Try to parse 'YYYY-MM-DD' or similar
        try:
            dt = datetime.strptime(dt[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None
    if isinstance(dt, datetime):
        dt = dt.date()
    if isinstance(dt, date):
        q = (dt.month - 1) // 3 + 1
        return f"{dt.year}-Q{q}"
    return None


def to_int(v) -> int | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in NULL_MARKERS:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def to_float(v) -> float | None:
    if v is None:
        return None
    s = str(v).strip()
    if s in NULL_MARKERS:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def read_quarter_headers(ws) -> dict[int, str]:
    """Read row HEADER_ROW; return {col_idx: 'YYYY-Q[1-4]'}."""
    quarter_cols: dict[int, str] = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx != HEADER_ROW:
            continue
        for c_idx, cell in enumerate(row):
            if c_idx < FIRST_QUARTER_COL:
                continue
            yq = datetime_to_year_qtr(cell)
            if yq:
                quarter_cols[c_idx] = yq
        break
    return quarter_cols


def extract_metric(wb, sheet_name: str, parser):
    """Read a single sheet, return {(sa2_code, year_qtr): value}."""
    ws = wb[sheet_name]
    quarter_cols = read_quarter_headers(ws)
    out: dict[tuple, float | int | None] = {}

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < DATA_START_ROW:
            continue
        if SA2_CODE_COL >= len(row):
            continue
        sa2_cell = row[SA2_CODE_COL]
        if sa2_cell is None:
            continue
        sa2 = str(sa2_cell).strip()
        if not (sa2.isdigit() and len(sa2) == 9):
            continue

        for col_idx, year_qtr in quarter_cols.items():
            if col_idx >= len(row):
                continue
            val = parser(row[col_idx])
            out[(sa2, year_qtr)] = val

    return out, quarter_cols


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true",
                        help="Mutate the DB. Without this flag, dry-run only.")
    args = parser.parse_args()

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(banner(f"LAYER 2 STEP 5 — SALM SA2 INGEST ({mode})"))
    print(f"DB:        {DB}")
    print(f"Timestamp: {TS}")

    if not DB.exists():
        print("FAIL: DB not found")
        return 1

    wb_path = find_workbook(SA2_PREFIX)
    if wb_path is None:
        print(f"FAIL: no workbook in abs_data/ starting with '{SA2_PREFIX}'")
        return 1
    print(f"Workbook:  {wb_path.name}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ---- EXTRACT all three sheets ----------------------------------
    print(banner("EXTRACT"))
    t0 = time.time()
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    rate_data, rate_q = extract_metric(
        wb, SHEET_FOR_METRIC["rate"], to_float)
    print(f"  rate sheet:         {len(rate_data):,} (sa2,quarter) cells, "
          f"{len(rate_q)} quarter columns")

    count_data, count_q = extract_metric(
        wb, SHEET_FOR_METRIC["count"], to_int)
    print(f"  count sheet:        {len(count_data):,} (sa2,quarter) cells, "
          f"{len(count_q)} quarter columns")

    lf_data, lf_q = extract_metric(
        wb, SHEET_FOR_METRIC["labour_force"], to_int)
    print(f"  labour_force sheet: {len(lf_data):,} (sa2,quarter) cells, "
          f"{len(lf_q)} quarter columns")

    wb.close()
    elapsed = time.time() - t0
    print(f"  Extract time: {elapsed:.1f}s")

    # ---- Sanity: same quarter columns across all 3 sheets ----------
    if not (rate_q == count_q == lf_q):
        print("FAIL: quarter columns differ across sheets")
        print(f"  rate quarters:  {sorted(rate_q.values())[:3]} ... "
              f"{sorted(rate_q.values())[-3:]}")
        print(f"  count quarters: {sorted(count_q.values())[:3]} ... "
              f"{sorted(count_q.values())[-3:]}")
        return 1

    # Build merged rows
    all_keys = (set(rate_data.keys()) | set(count_data.keys())
                | set(lf_data.keys()))
    all_keys = sorted(all_keys)
    rows = []
    for key in all_keys:
        sa2, yq = key
        rows.append((
            sa2, yq,
            rate_data.get(key),
            count_data.get(key),
            lf_data.get(key),
        ))

    n_total = len(rows)
    n_rate_nn = sum(1 for r in rows if r[2] is not None)
    n_count_nn = sum(1 for r in rows if r[3] is not None)
    n_lf_nn = sum(1 for r in rows if r[4] is not None)
    distinct_sa2 = len({r[0] for r in rows})
    distinct_qtr = sorted({r[1] for r in rows})

    print()
    print(f"  Total rows:        {n_total:,}")
    print(f"  Distinct SA2:      {distinct_sa2:,}")
    print(f"  Distinct quarters: {len(distinct_qtr)} "
          f"({distinct_qtr[0]} -> {distinct_qtr[-1]})")
    print(f"  Non-null rate:         {n_rate_nn:,}")
    print(f"  Non-null count:        {n_count_nn:,}")
    print(f"  Non-null labour_force: {n_lf_nn:,}")

    # Sample
    print("\nFirst 5 rows:")
    for r in rows[:5]:
        print(f"  {r}")

    # Sample one SA2 across all quarters
    if rows:
        sample_sa2 = rows[0][0]
        sample = [r for r in rows if r[0] == sample_sa2]
        print(f"\nSA2 {sample_sa2} — first 5 quarters and last 5:")
        for r in sample[:5]:
            print(f"  {r}")
        if len(sample) > 5:
            print("  ...")
            for r in sample[-5:]:
                print(f"  {r}")

    # ---- Sanity check: rate ~= count/labour_force * 100 ------------
    print(banner("SANITY CHECK: rate ~= count/labour_force * 100"))
    checks = 0
    mismatches = 0
    sample_mismatches = []
    for r in rows:
        sa2, yq, rate, cnt, lf = r
        if rate is None or cnt is None or lf is None or lf == 0:
            continue
        checks += 1
        derived = (cnt / lf) * 100
        if abs(derived - rate) > 0.5:  # allow 0.5pp tolerance for rounding
            mismatches += 1
            if len(sample_mismatches) < 5:
                sample_mismatches.append(
                    (sa2, yq, rate, cnt, lf, derived))
    print(f"  Checks performed: {checks:,}")
    print(f"  Mismatches > 0.5pp: {mismatches}")
    if sample_mismatches:
        print("  Sample mismatches:")
        for s in sample_mismatches:
            print(f"    sa2={s[0]} yq={s[1]} stored_rate={s[2]} "
                  f"cnt={s[3]} lf={s[4]} derived={s[5]:.2f}")
    else:
        print("  All rows internally consistent.")

    # ---- Sanity check: count <= labour_force -----------------------
    print(banner("SANITY CHECK: count <= labour_force"))
    violations = 0
    sample_v = []
    for r in rows:
        sa2, yq, rate, cnt, lf = r
        if cnt is None or lf is None:
            continue
        if cnt > lf:
            violations += 1
            if len(sample_v) < 5:
                sample_v.append(r)
    print(f"  Violations: {violations}")
    if sample_v:
        for s in sample_v:
            print(f"    {s}")

    # ---- Write changeset -------------------------------------------
    print(banner("CHANGESET"))
    with open(CHANGESET_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sa2_code", "year_qtr", "rate", "count", "labour_force"])
        w.writerows(rows)
    print(f"Changeset written: {CHANGESET_PATH}")

    if not args.apply:
        print(banner("DRY-RUN COMPLETE"))
        print("No DB mutation. Review sanity checks above.")
        print("Re-run with --apply when ready:")
        print("  python layer2_step5_apply.py --apply")
        return 0

    # ================================================================
    # APPLY MODE
    # ================================================================
    print(banner("BACKUP"))
    shutil.copy2(DB, BACKUP_PATH)
    print(f"OK: backed up to {BACKUP_PATH}")
    print(f"    size: {BACKUP_PATH.stat().st_size:,} bytes")

    conn = sqlite3.connect(DB)
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    try:
        print(banner("BEGIN TRANSACTION"))
        cur.execute("BEGIN TRANSACTION")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS abs_sa2_unemployment_quarterly (
                sa2_code      TEXT NOT NULL,
                year_qtr      TEXT NOT NULL,
                rate          REAL,
                count         INTEGER,
                labour_force  INTEGER,
                PRIMARY KEY (sa2_code, year_qtr)
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS ix_salm_year_qtr "
                    "ON abs_sa2_unemployment_quarterly(year_qtr)")
        print("CREATE TABLE / INDEX: OK")

        cur.execute("SELECT COUNT(*) FROM abs_sa2_unemployment_quarterly")
        pre_count = cur.fetchone()[0]
        print(f"Pre-state: {pre_count:,} rows")
        if pre_count > 0:
            print("ABORT: table is not empty.")
            conn.execute("ROLLBACK")
            return 1

        # INSERT
        print(banner("INSERT"))
        chunk_size = 1000
        inserted = 0
        for i in range(0, n_total, chunk_size):
            chunk = rows[i:i + chunk_size]
            cur.executemany(
                "INSERT INTO abs_sa2_unemployment_quarterly "
                "(sa2_code, year_qtr, rate, count, labour_force) "
                "VALUES (?, ?, ?, ?, ?)",
                chunk
            )
            inserted += len(chunk)
        print(f"Inserted {inserted:,} rows")

        # Post-invariants
        cur.execute("SELECT COUNT(*) FROM abs_sa2_unemployment_quarterly")
        post_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(DISTINCT sa2_code) "
                    "FROM abs_sa2_unemployment_quarterly")
        post_sa2 = cur.fetchone()[0]
        cur.execute("SELECT COUNT(DISTINCT year_qtr) "
                    "FROM abs_sa2_unemployment_quarterly")
        post_qtr = cur.fetchone()[0]

        print(banner("POST-INVARIANTS"))
        print(f"  rows:              {post_count:,}")
        print(f"  distinct sa2:      {post_sa2:,}")
        print(f"  distinct quarters: {post_qtr}")

        if post_count != n_total:
            print(f"ABORT: post_count {post_count} != extracted {n_total}")
            conn.execute("ROLLBACK")
            return 1

        # Audit log
        print(banner("AUDIT LOG"))
        payload = {
            "rows_inserted": post_count,
            "distinct_sa2": post_sa2,
            "distinct_quarters": post_qtr,
            "first_quarter": distinct_qtr[0],
            "last_quarter": distinct_qtr[-1],
            "non_null_rate": n_rate_nn,
            "non_null_count": n_count_nn,
            "non_null_labour_force": n_lf_nn,
            "source_file": wb_path.name,
            "backup_path": BACKUP_PATH.name,
            "changeset_csv": CHANGESET_PATH.name,
        }
        cur.execute("""
            INSERT INTO audit_log (
                actor, action, subject_type, subject_id,
                before_json, after_json, reason, occurred_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            ACTOR, ACTION, SUBJECT_TYPE, SUBJECT_ID,
            json.dumps({"rows": pre_count}),
            json.dumps({"rows": post_count, "payload": payload}),
            "Layer 2 Step 5: SALM SA2 unemployment ingest",
            datetime.now().isoformat(),
        ))
        cur.execute("SELECT MAX(audit_id) FROM audit_log")
        new_audit_id = cur.fetchone()[0]
        print(f"audit_log row: id={new_audit_id} action='{ACTION}'")

        conn.commit()
        print(banner("COMMIT SUCCESSFUL"))
        print(f"  Rows in abs_sa2_unemployment_quarterly: {post_count:,}")
        print(f"  Backup file:    {BACKUP_PATH.name}")
        print(f"  Changeset CSV:  {CHANGESET_PATH.name}")
        print(f"  audit_id:       {new_audit_id}")
        return 0

    except Exception as exc:
        print(banner(f"FAIL: {type(exc).__name__}"))
        print(str(exc))
        try:
            conn.execute("ROLLBACK")
            print("ROLLBACK successful - DB unchanged.")
        except Exception:
            pass
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
