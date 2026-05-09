r"""
A4 — Schools at SA2 column probe (read-only).

Inspects all four candidate workbooks for any school / enrolment / education-
infrastructure columns at SA2 level:

  1. abs_data/Education and employment database.xlsx — columns we don't already
     ingest via layer2_step5b_prime_apply.py
  2. abs_data/Family and Community Database.xlsx — sometimes carries school stats
  3. abs_data/Population and People Database.xlsx — has age cohort cols already
     read for parent-cohort; may also carry schooling counts
  4. Any obvious ACARA file in abs_data/ (currently absent — flag as gap)

Run from repo root (set PYTHONIOENCODING=utf-8 first):
    python recon\probes\probe_a4_schools_columns.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ABS_DATA = Path("abs_data")

WORKBOOKS = [
    ("Education and employment database.xlsx", "Table 1"),
    ("Family and Community Database.xlsx", None),  # sheet detected at runtime
    ("Population and People Database.xlsx", "Table 1"),
    ("Economic and Industry Database.xlsx", None),
    ("Income Database.xlsx", None),
]

# Existing EE_METRICS columns (1-based) already ingested by layer2_step5b_prime
EE_INGESTED_COLS = {4, 9, 11, 12, 22, 54, 55, 60, 62, 72, 73, 75, 76, 83, 84, 87}

SCHOOL_KEYWORDS = [
    "school", "enrol", "primary", "secondary", "kindergar",
    "pupil", "student", "year level", "year_level",
]


def section(title):
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def probe_workbook(filename, preferred_sheet):
    path = ABS_DATA / filename
    if not path.exists():
        print(f"  SKIP: {path} not found.")
        return

    section(f"{filename}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    print(f"  Sheets: {sheets}")

    sheet_name = preferred_sheet if preferred_sheet in sheets else None
    if sheet_name is None:
        # Scan all sheets that look data-bearing
        for sn in sheets:
            if sn.lower().startswith("table"):
                sheet_name = sn
                break
    if sheet_name is None:
        sheet_name = sheets[1] if len(sheets) > 1 else sheets[0]

    print(f"  Probing sheet: '{sheet_name}'")
    ws = wb[sheet_name]

    # Read the first 8 rows so we get the header rows (row 6 spanning, row 7 col labels)
    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
    if len(rows) < 7:
        print(f"  ERROR: <7 rows; cannot find header row.")
        return

    row6 = rows[5] if len(rows) >= 6 else ()
    row7 = rows[6] if len(rows) >= 7 else ()

    # Find school-related columns in row 7
    print(f"  Row 7 has {len(row7)} cells. Scanning for school keywords...")
    hits = []
    for idx, label in enumerate(row7):
        if label is None:
            continue
        s = str(label).lower()
        if any(k in s for k in SCHOOL_KEYWORDS):
            spanning = ""
            if idx < len(row6) and row6[idx]:
                spanning = str(row6[idx])
            already = "* INGESTED *" if (idx + 1) in EE_INGESTED_COLS and "Education and employment" in filename else ""
            hits.append((idx, label, spanning, already))

    if not hits:
        print("  No school-keyword columns found.")
        return

    # Find the spanning headers covering this sheet for context
    print(f"  Spanning headers (row 6) — anchors:")
    for idx, label in enumerate(row6):
        if label is not None:
            print(f"    [{idx:>3}] {str(label)[:90]}")

    print(f"\n  School-related row 7 columns:")
    for idx, label, spanning, already in hits:
        spanning_str = f"  span:'{spanning[:50]}'" if spanning else ""
        print(f"    [{idx:>3}] {str(label)[:80]}{spanning_str} {already}")

    wb.close()


def main():
    if not ABS_DATA.exists():
        print(f"ERROR: {ABS_DATA} not found.")
        sys.exit(1)

    print("Existing preschool series in DB (ee_preschool_*_count):")
    print("  - ee_preschool_4yo_count        (col 4)")
    print("  - ee_preschool_total_count      (col 9)")
    print("  - ee_preschool_15h_plus_count   (col 11)")
    print("All from Education and employment database.xlsx Table 1.")

    for fn, sheet in WORKBOOKS:
        probe_workbook(fn, sheet)

    section("ACARA-direct file presence check")
    acara_candidates = [p for p in ABS_DATA.iterdir()
                        if "acara" in p.name.lower() or "school" in p.name.lower()]
    if acara_candidates:
        print("  Candidate files:")
        for p in acara_candidates:
            print(f"    {p}")
    else:
        print("  No ACARA-direct file in abs_data/. A4 must derive school")
        print("  enrolment from existing workbook columns or note the gap.")


if __name__ == "__main__":
    main()
