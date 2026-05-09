r"""
A4 ACARA file probe — read-only inspection of whatever ACARA workbooks
Patrick has dropped into abs_data/.

Auto-discovers any file matching `acara_*.xlsx`, `school_*.xlsx`, or
`*acara*.xlsx` (ACARA's own naming varies by release year). Reports:
  - Sheets in each file
  - Row 1 + first data row of the main sheet
  - Detected lat/lon columns
  - Detected sector / school-type columns
  - Detected enrolment columns
  - Year column if a time-series file
  - SA2 / state / postcode columns (for spatial-join cross-check)

Goal: confirm V1 fields are present + identify the right column map for
the layer4_4_step_a4_schools_apply.py ingest before any DB mutation.

Run from repo root:
    python recon\probes\probe_a4_acara_files.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ABS_DATA = Path("abs_data")

# Filename patterns that suggest an ACARA workbook
ACARA_PATTERNS = [
    re.compile(r"^acara[_\s-]", re.I),
    re.compile(r"school[_\s]?profile", re.I),
    re.compile(r"school[_\s]?location", re.I),
    re.compile(r"^school[_\s-]", re.I),
    re.compile(r"enrolment.*grade", re.I),
]

# Required V1 fields — what we need to confirm before ingest
V1_REQUIRED = {
    "lat":          ["lat", "latitude"],
    "lng":          ["lng", "lon", "long", "longitude"],
    "school_id":    ["acara_id", "acara id", "school_id", "school id", "school no"],
    "school_name":  ["school name", "school_name"],
    "sector":       ["school sector", "sector", "school_sector"],
    "type":         ["school type", "type", "school_type"],
    "enrolment":    ["enrol", "enrolment", "enrolments", "enrollment", "students",
                     "student count", "student_count", "total enrolments"],
    "state":        ["state", "state/territory"],
    "year":         ["year", "data year", "calendar year"],
    "icsea":        ["icsea"],
    "lbote":        ["lbote", "lbote percent", "lbote (%)"],
}

# Columns that signal it's the right kind of ACARA file
SECTOR_VALUES = {"Government", "Catholic", "Independent",
                 "Govt.", "Independent (Non-Government)"}


def section(title):
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def discover_acara_files():
    if not ABS_DATA.exists():
        return []
    out = []
    for p in sorted(ABS_DATA.iterdir()):
        if not p.is_file():
            continue
        if p.suffix.lower() not in (".xlsx", ".xls", ".csv"):
            continue
        name = p.name
        if any(pat.search(name) for pat in ACARA_PATTERNS):
            out.append(p)
    return out


def detect_field(headers_lower, candidates):
    """Find the first header column whose lowercase form contains any of
    the candidate substrings. Returns (idx, original_label) or None."""
    for idx, h in enumerate(headers_lower):
        if h is None:
            continue
        for cand in candidates:
            if cand in h:
                return idx, h
    return None


def find_header_row(ws, max_scan=30):
    """ACARA workbooks vary — sometimes header is row 1, sometimes row 5+.
    Scan first N rows; return the row index whose cells look most like
    a header (mostly text, few nulls)."""
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows):
        if not row:
            continue
        non_null = sum(1 for c in row if c is not None and str(c).strip())
        text_count = sum(1 for c in row
                         if c is not None
                         and isinstance(c, str)
                         and not c.replace(".", "").replace("-", "").isdigit())
        # Header heuristic: many non-null text cells, few numeric cells
        score = text_count - 0.5 * (non_null - text_count)
        if non_null >= 5 and score > best_score:
            best_score = score
            best_idx = idx
    return best_idx, rows[best_idx] if rows else ()


def probe_sheet(path, sheet_name):
    print(f"\n  --- Sheet: {sheet_name!r} ---")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    header_idx, header_row = find_header_row(ws)
    if not header_row:
        print(f"    (sheet appears empty)")
        return
    print(f"    Header row detected: row {header_idx + 1} ({len(header_row)} cells)")

    # Pretty-print first 25 columns
    print(f"    First 25 column labels:")
    for i, label in enumerate(header_row[:25]):
        s = str(label)[:60] if label is not None else "(blank)"
        print(f"      [{i:>3}] {s}")
    if len(header_row) > 25:
        print(f"      ... ({len(header_row) - 25} more cols)")

    # Detect V1 fields
    headers_lower = [str(h).lower().strip() if h is not None else None
                     for h in header_row]
    print(f"\n    V1 field detection:")
    found = {}
    missing = []
    for field, candidates in V1_REQUIRED.items():
        result = detect_field(headers_lower, candidates)
        if result:
            idx, label = result
            print(f"      [{idx:>3}] {field:<14s} <- '{label}'")
            found[field] = (idx, label)
        else:
            missing.append(field)

    if missing:
        print(f"\n    MISSING fields (will need manual map or different file):")
        for f in missing:
            print(f"      - {f}")

    # Sample first 2 data rows
    print(f"\n    Sample data rows (rows {header_idx + 2}, {header_idx + 3}):")
    sample_rows = list(ws.iter_rows(
        min_row=header_idx + 2, max_row=header_idx + 3, values_only=True))
    for r_idx, row in enumerate(sample_rows):
        print(f"      Row {header_idx + 2 + r_idx}:")
        for field, (idx, _) in list(found.items())[:8]:
            v = row[idx] if idx < len(row) else "(out of range)"
            v_str = str(v)[:50] if v is not None else "(null)"
            print(f"        {field:<14s}: {v_str}")

    # Sector value scan if sector column found
    if "sector" in found:
        idx = found["sector"][0]
        sectors = set()
        for row in ws.iter_rows(min_row=header_idx + 2,
                                max_row=header_idx + 100,
                                values_only=True):
            if idx < len(row) and row[idx] is not None:
                sectors.add(str(row[idx]).strip())
        print(f"\n    Distinct sector values (first 100 rows): {sorted(sectors)}")

    # Year scan if year column found
    if "year" in found:
        idx = found["year"][0]
        years = set()
        for row in ws.iter_rows(min_row=header_idx + 2,
                                max_row=header_idx + 500,
                                values_only=True):
            if idx < len(row) and row[idx] is not None:
                try:
                    years.add(int(float(str(row[idx]))))
                except (ValueError, TypeError):
                    pass
        if years:
            print(f"\n    Year values (first 500 rows): "
                  f"{min(years)}-{max(years)} ({len(years)} distinct)")

    wb.close()


def probe_file(path):
    section(f"{path.name}  ({path.stat().st_size / (1024*1024):.1f} MB)")
    if path.suffix.lower() == ".csv":
        # Quick CSV header peek
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            first = f.readline().rstrip()
            second = f.readline().rstrip()
        print(f"  CSV first line ({len(first)} chars):")
        print(f"    {first[:300]}")
        print(f"  CSV second line: {second[:300]}")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"  Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    wb.close()

    # Probe first non-meta sheet (skip "Notes", "Cover", "Contents")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    skip_keywords = ("note", "cover", "content", "table of", "metadata")
    target_sheets = []
    for sn in wb.sheetnames:
        if not any(kw in sn.lower() for kw in skip_keywords):
            target_sheets.append(sn)
    wb.close()

    if not target_sheets:
        print(f"  All sheets look like metadata; probing first sheet anyway.")
        target_sheets = wb.sheetnames[:1]

    for sn in target_sheets[:2]:  # probe up to 2 data sheets
        probe_sheet(path, sn)


def main():
    print(f"Scanning {ABS_DATA}/ for ACARA files...")
    files = discover_acara_files()
    if not files:
        print("\n  NO ACARA-NAMED FILES FOUND.")
        print("  Drop ACARA workbooks into abs_data/ with names like:")
        print("    - school_profile_2025.xlsx")
        print("    - school_location_2025.xlsx")
        print("    - school_profile_2008-2025.xlsx")
        print("  Then re-run this probe.")
        return

    print(f"  Found {len(files)} candidate file(s):")
    for p in files:
        print(f"    {p.name}  ({p.stat().st_size / (1024*1024):.1f} MB)")

    for p in files:
        probe_file(p)

    section("DONE")
    print("Read-only probe. No writes performed.")


if __name__ == "__main__":
    main()
