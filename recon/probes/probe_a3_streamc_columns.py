r"""
A3 + Stream C column probe — read-only inspection of:

  1) TSP zip T05 (marital status) column shapes
  2) TSP zip T07 (children-ever-born / fertility) column shapes
  3) ERP "Population and People Database.xlsx" age columns covering 25-44

Banked from A10 probe (recon/a10_c8_design.md §1):
  - T05 holds C{yr}_{age}_Mar_RegM_* columns (Married Registered, age x sex x year)
  - T07 holds C{yr}_AP_{age}_NCB_{count} columns (NCB = Number of Children Born)
  - Both live in 2021_TSP_SA2_for_AUS_short-header.zip alongside T06/T08/T14

This probe confirms exact column names and samples for the four
verification SA2s carried from A2/A10:
    211011251 Bayswater Vic
    118011341 Bondi Junction-Waverly NSW
    506031124 Bentley-Wilson-St James WA
    702041063 Outback NT (high-ATSI)

Run from repo root (set PYTHONIOENCODING=utf-8 first):
    python recon\probes\probe_a3_streamc_columns.py
"""

from __future__ import annotations

import io
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd

ZIP_PATH = Path("abs_data") / "2021_TSP_SA2_for_AUS_short-header.zip"
ERP_WORKBOOK = Path("abs_data") / "Population and People Database.xlsx"
ZIP_PREFIX = "2021 Census TSP Statistical Area 2 for AUS/"

VERIFY_SA2 = [
    ("211011251", "Bayswater Vic"),
    ("118011341", "Bondi Junction-Waverly NSW"),
    ("506031124", "Bentley-Wilson-St James WA"),
    ("702041063", "Outback NT"),
]


def section(title):
    print()
    print("=" * 72)
    print(title)
    print("=" * 72)


def read_csv_from_zip(z, name):
    if name not in z.namelist():
        return None
    df = pd.read_csv(io.BytesIO(z.read(name)), dtype=str, low_memory=False)
    df.columns = [str(c).strip() for c in df.columns]
    if "SA2_CODE_2021" in df.columns:
        df["SA2_CODE_2021"] = df["SA2_CODE_2021"].astype(str).str.strip()
        sa2_re = re.compile(r"^\d{9}$")
        df = df[df["SA2_CODE_2021"].str.match(sa2_re)].copy()
    return df


def list_t_files(z, t_code):
    """Return sorted list of CSV names matching 2021Census_T{t_code}{A,B,C,...}_AUST_SA2.csv."""
    pat = re.compile(rf"2021Census_T{t_code}([A-Z]?)_AUST_SA2\.csv$")
    out = []
    for n in z.namelist():
        if pat.search(n):
            out.append(n)
    return sorted(out)


def column_headers(name, df):
    print(f"  {name}")
    print(f"    rows={len(df):,} cols={len(df.columns)}")
    sample = list(df.columns)[:8]
    print(f"    first 8 cols: {sample}")


def group_columns(cols, regex):
    """Return columns matching regex, alphabetised."""
    pat = re.compile(regex)
    return sorted([c for c in cols if pat.search(c)])


# =====================================================================
# T05 marital status probe
# =====================================================================
def probe_t05(z):
    section("T05 — Marital status (TSP)")
    files = list_t_files(z, "05")
    print(f"  T05 files in zip: {len(files)}")
    for f in files:
        print(f"    {f}")
    if not files:
        print("  WARNING: no T05 files found.")
        return

    # Read each file and survey columns
    for f in files:
        df = read_csv_from_zip(z, f)
        column_headers(f.split("/")[-1], df)

        # All cols sample
        cols = list(df.columns)

        # Try to find Mar_RegM family / age bands / census-year markers
        mar_regm = group_columns(cols, r"Mar_RegM")
        mar_regs = group_columns(cols, r"Mar_RegS")
        nev_mar = group_columns(cols, r"Nev_Marr|Nev_marr|NeverMar|Never_Mar")
        sep_div = group_columns(cols, r"Sep|Div|Wid")
        tot_mar = group_columns(cols, r"_Tot_(Mar|Persons|P|Persons_Tot)")

        print(f"    Mar_RegM cols: {len(mar_regm)} (sample: {mar_regm[:3]})")
        print(f"    Mar_RegS cols: {len(mar_regs)} (sample: {mar_regs[:3]})")
        print(f"    Never-married cols: {len(nev_mar)} (sample: {nev_mar[:3]})")
        print(f"    Sep/Div/Wid cols: {len(sep_div)} (sample: {sep_div[:3]})")
        print(f"    Tot-bearing cols: {len(tot_mar)} (sample: {tot_mar[:3]})")

        # Detect age bands via regex `_(\d+_\d+|\d+_yrs)_` snippets
        age_bands = set()
        for c in cols:
            m = re.search(r"(\d{2}_\d{2}|85_ov|85ov|85_over)", c)
            if m:
                age_bands.add(m.group(1))
        print(f"    Distinct age bands: {sorted(age_bands)}")

        # Census-year prefixes
        cyr = set()
        for c in cols:
            m = re.match(r"C(\d{2})_", c)
            if m:
                cyr.add(m.group(1))
        print(f"    Census-year prefixes: {sorted(cyr)}")


# =====================================================================
# T07 children-ever-born (fertility) probe
# =====================================================================
def probe_t07(z):
    section("T07 — Children ever born / fertility (TSP)")
    files = list_t_files(z, "07")
    print(f"  T07 files in zip: {len(files)}")
    for f in files:
        print(f"    {f}")
    if not files:
        print("  WARNING: no T07 files found.")
        return

    for f in files:
        df = read_csv_from_zip(z, f)
        column_headers(f.split("/")[-1], df)
        cols = list(df.columns)

        ncb = group_columns(cols, r"_NCB_")
        ap = group_columns(cols, r"_AP_")
        cef_yes = group_columns(cols, r"CEFB|Children_ever_born")
        tot = group_columns(cols, r"_Tot$|_Tot_P|_T_NCB|_NCB_T")

        print(f"    NCB cols: {len(ncb)} (sample: {ncb[:3]})")
        print(f"    AP cols: {len(ap)} (sample: {ap[:3]})")
        print(f"    CEFB-direct cols: {len(cef_yes)} (sample: {cef_yes[:3]})")
        print(f"    Tot-bearing cols: {len(tot)} (sample: {tot[:3]})")

        age_bands = set()
        for c in cols:
            m = re.search(r"(\d{2}_\d{2}|85_ov|85ov|85_over)", c)
            if m:
                age_bands.add(m.group(1))
        print(f"    Distinct age bands: {sorted(age_bands)}")

        cyr = set()
        for c in cols:
            m = re.match(r"C(\d{2})_", c)
            if m:
                cyr.add(m.group(1))
        print(f"    Census-year prefixes: {sorted(cyr)}")

        ncb_counts = set()
        for c in ncb:
            m = re.search(r"_NCB_([0-9]+|6mo|6m|6_mo|None|Tot)", c)
            if m:
                ncb_counts.add(m.group(1))
        print(f"    NCB counts seen: {sorted(ncb_counts)}")


# =====================================================================
# Verification SA2 sample print
# =====================================================================
def sample_verify_t05(z):
    section("T05 — Verification SA2 spot sample")
    files = list_t_files(z, "05")
    if not files:
        return
    # Use the last file (likely 2021/T05C) for spot
    target = files[-1]
    df = read_csv_from_zip(z, target).set_index("SA2_CODE_2021")
    cols = list(df.columns)
    # Pick a tiny sample: first 6 columns + any C21_ Tot Persons column
    sample_cols = cols[:6]
    extra = [c for c in cols if "C21" in c and ("Tot_P" in c or "_Tot" in c)][:2]
    sample_cols = list(dict.fromkeys(sample_cols + extra))
    print(f"  Showing cols: {sample_cols}")
    for sa2, name in VERIFY_SA2:
        if sa2 not in df.index:
            print(f"    {sa2} {name}: NOT IN INDEX")
            continue
        row = df.loc[sa2]
        cells = " | ".join(f"{c}={row[c]}" for c in sample_cols)
        print(f"    {sa2} {name}: {cells[:250]}")


def sample_verify_t07(z):
    section("T07 — Verification SA2 spot sample")
    files = list_t_files(z, "07")
    if not files:
        return
    target = files[-1]
    df = read_csv_from_zip(z, target).set_index("SA2_CODE_2021")
    cols = list(df.columns)
    sample_cols = cols[:6]
    extra = [c for c in cols if "C21" in c and "_NCB_" in c][:3]
    sample_cols = list(dict.fromkeys(sample_cols + extra))
    print(f"  Showing cols: {sample_cols}")
    for sa2, name in VERIFY_SA2:
        if sa2 not in df.index:
            print(f"    {sa2} {name}: NOT IN INDEX")
            continue
        row = df.loc[sa2]
        cells = " | ".join(f"{c}={row[c]}" for c in sample_cols)
        print(f"    {sa2} {name}: {cells[:250]}")


# =====================================================================
# ERP age columns probe (Population and People Database.xlsx)
# =====================================================================
def probe_erp_age_columns():
    section("ERP — Population and People Database.xlsx column header survey")
    if not ERP_WORKBOOK.exists():
        print(f"  ERROR: {ERP_WORKBOOK} not found.")
        return

    wb = openpyxl.load_workbook(ERP_WORKBOOK, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheets: {sheet_names[:20]}")

    # We know from layer2_step6_diag.md that table sheet is "Table 1" (POPULATION)
    # but workbook may have other tables. List all sheets and surface the population one.
    for sn in sheet_names:
        ws = wb[sn]
        # Read row 6 (spanning headers) and row 7 (column labels) — first 165 cols
        try:
            rows = list(ws.iter_rows(min_row=6, max_row=7, values_only=True))
        except Exception as e:
            print(f"  Sheet '{sn}' read error: {e}")
            continue
        if not rows or not rows[0]:
            continue
        row6 = rows[0] if len(rows) >= 1 else ()
        row7 = rows[1] if len(rows) >= 2 else ()
        if not any("Estimated resident population" in str(c or "") for c in row6):
            continue

        # This is the population sheet
        print(f"\n  >> Population sheet: '{sn}'")
        print(f"     Total cols: {len(row7)}")

        # Find all "X-Y years" age-band column indices in row 7
        # Note: ABS workbook uses "(no.)" not "(n)" for count units.
        age_band_re = re.compile(r"(?P<sex>Males|Females|Persons)\s*-\s*(?P<lo>\d+)\s*-\s*(?P<hi>\d+)\s*years\s*\(no\.?\)", re.I)
        age_band_85_re = re.compile(r"(?P<sex>Males|Females|Persons)\s*-\s*85\s*years\s*and\s*over\s*\(no\.?\)", re.I)
        results = []
        for idx, label in enumerate(row7):
            if label is None:
                continue
            s = str(label)
            m = age_band_re.search(s)
            if m:
                results.append((idx, m.group("sex"), int(m.group("lo")), int(m.group("hi")), s))
                continue
            m85 = age_band_85_re.search(s)
            if m85:
                results.append((idx, m85.group("sex"), 85, 199, s))

        # Sort by sex then age
        results.sort(key=lambda r: (r[1], r[2]))
        # Highlight the parent-cohort 25-44 columns
        target_lo, target_hi = 25, 44
        print(f"\n     All age-band 'n' columns (sex, lo-hi, col_idx):")
        for idx, sex, lo, hi, lbl in results:
            mark = "  <-- PARENT" if (sex == "Persons" and target_lo <= lo and hi <= target_hi) else ""
            print(f"       [{idx:>3}] {sex} {lo}-{hi:<3}  {lbl[:80]}{mark}")

        # Also surface total_persons column for parent-cohort denominator pairing
        for idx, label in enumerate(row7):
            if label and re.match(r"^Estimated resident population\s*-\s*(persons|total)", str(label), re.I):
                print(f"     Persons total col: [{idx}] {label}")
        # For belt-and-braces, find col 84 region (we expect Persons 0-4 starts here)
        print(f"     Row 6 col[84] (Persons spanning header): {row6[84] if len(row6) > 84 else 'NA'}")
        print(f"     Row 7 col[84]: {row7[84] if len(row7) > 84 else 'NA'}")
        break  # only first matching sheet


def main():
    if not ZIP_PATH.exists():
        print(f"ERROR: {ZIP_PATH} not found.")
        sys.exit(1)

    print(f"Zip:  {ZIP_PATH}  ({ZIP_PATH.stat().st_size / (1024 * 1024):.1f} MB)")
    print(f"ERP:  {ERP_WORKBOOK}  ({ERP_WORKBOOK.stat().st_size / (1024 * 1024):.1f} MB)")

    with zipfile.ZipFile(ZIP_PATH, "r") as z:
        probe_t05(z)
        probe_t07(z)
        sample_verify_t05(z)
        sample_verify_t07(z)

    probe_erp_age_columns()

    section("DONE")
    print("Read-only probe. No writes performed.")


if __name__ == "__main__":
    main()
