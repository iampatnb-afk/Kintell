"""
Layer 2 Step 5b DIAGNOSTIC — Census 2021 subset layout discovery
─────────────────────────────────────────────────────────────────
READ-ONLY. No DB mutations. No file writes outside recon/.

Probes BOTH Census workbooks to discover the target columns:
  - Income Database.xlsx -> median household income (weekly)
  - Education and employment database.xlsx -> female LFP %

Per Standard 19 expected structure:
  row 6 = spanning headers
  row 7 = sub-headers
  row 8+ = data rows

Per Decision 50: auto-detected columns must validate by sample value
range, not just header text matching.

Outputs:
  - Console summary
  - recon/layer2_step5b_diag.md
  - recon/layer2_step5b_column_map.json
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
DB = PROJECT_ROOT / "data" / "kintell.db"
ABS = PROJECT_ROOT / "abs_data"
OUT_DIR = PROJECT_ROOT / "recon"
OUT_MD = OUT_DIR / "layer2_step5b_diag.md"
OUT_JSON = OUT_DIR / "layer2_step5b_column_map.json"

INCOME_FILE = "Income Database.xlsx"
EMPL_FILE = "Education and employment database.xlsx"

HEADER_ROW = 6


def banner(title: str) -> str:
    return f"\n{'-' * 70}\n{title}\n{'-' * 70}"


def looks_like_sa2(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return s.isdigit() and len(s) == 9


def header_text(rows: list, col_idx: int) -> str:
    """Concatenate non-null header text from rows 6 and 7 for a column."""
    parts = []
    for r_idx in (6, 7):
        if r_idx <= len(rows):
            v = rows[r_idx - 1][col_idx] if col_idx < len(rows[r_idx - 1]) else None
            if v is not None:
                parts.append(str(v))
    return " | ".join(parts)


def find_columns_matching(rows: list, n_cols: int, patterns: list[str]) -> list:
    """Search rows 1-9 of all columns for header text matching any pattern.
    Returns list of (col_idx, joined_header) sorted by relevance."""
    out = []
    for c_idx in range(n_cols):
        full_header = []
        for r_idx in range(min(9, len(rows))):
            v = rows[r_idx][c_idx] if c_idx < len(rows[r_idx]) else None
            if v is not None:
                full_header.append(str(v).lower())
        joined = " | ".join(full_header)
        for pat in patterns:
            if pat in joined:
                out.append((c_idx, joined))
                break
    return out


def sample_column_values(wb_path: Path, sheet: str, col_idx: int,
                         data_start_row: int, n: int = 30) -> list:
    """Return up to n non-null sample values from data rows of one column."""
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[sheet]
    values = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < data_start_row:
            continue
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None:
            continue
        s = str(v).strip()
        if s in ("-", "..", "np", "n.a.", "na", "NA", ""):
            continue
        try:
            values.append(float(s))
        except (ValueError, TypeError):
            pass
        if len(values) >= n:
            break
    wb.close()
    return values


def probe_workbook(wb_path: Path, label: str,
                   target_patterns: list[str]) -> dict:
    """Read first 12 rows of Table 1 and identify target columns."""
    print(banner(f"PROBE: {label}"))
    print(f"  File: {wb_path.name}")

    if not wb_path.exists():
        print(f"  FAIL: file not found")
        return {}

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheets: {sheet_names}")

    # Table 1 should be the SA2 sheet
    table1_name = None
    for sn in sheet_names:
        norm = sn.lower().strip().replace(" ", "")
        if norm.startswith("table1"):
            table1_name = sn
            break
    if table1_name is None:
        # Fall back to any sheet starting with 'Table'
        for sn in sheet_names:
            if sn.lower().startswith("table"):
                table1_name = sn
                break
    if table1_name is None:
        print("  FAIL: no Table 1 sheet found")
        wb.close()
        return {}

    ws = wb[table1_name]
    print(f"  Using sheet: '{table1_name}'")

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        if i >= 12:
            break

    n_cols = max(len(r) for r in rows) if rows else 0
    print(f"  Total cols: {n_cols}")

    # Print rows 1-9 col 0-3 (region code/label area)
    print("\n  First 9 rows, cols 0-3:")
    for r_idx, row in enumerate(rows[:9], start=1):
        cells = []
        for c in row[:4]:
            cells.append("·" if c is None else str(c)[:25])
        print(f"    Row {r_idx}: " + " | ".join(cells))

    # Find columns matching target patterns
    print(f"\n  Columns matching target patterns:")
    print(f"    Patterns: {target_patterns}")
    matches = find_columns_matching(rows, n_cols, target_patterns)
    print(f"    Found {len(matches)} matching column(s):")
    for c_idx, header in matches[:20]:
        print(f"      col[{c_idx:3d}]: {header[:100]}")

    # For each match, sample data values for sanity check
    print(f"\n  Sample values per matching column (data starting row 8):")
    column_samples = {}
    for c_idx, header in matches[:10]:
        samples = sample_column_values(wb_path, table1_name, c_idx, 8, n=10)
        column_samples[c_idx] = samples
        if samples:
            avg = sum(samples) / len(samples)
            print(f"    col[{c_idx:3d}]: n={len(samples)} mean={avg:.2f} "
                  f"min={min(samples):.2f} max={max(samples):.2f}")
            print(f"            samples: {samples[:5]}")
        else:
            print(f"    col[{c_idx:3d}]: no numeric samples")

    wb.close()
    return {
        "file": wb_path.name,
        "sheet": table1_name,
        "n_cols": n_cols,
        "matches": [{"col": c, "header": h, "samples":
                     column_samples.get(c, [])}
                    for c, h in matches],
    }


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    findings = ["# Layer 2 Step 5b Diagnostic — Census 2021 subset", ""]

    income_path = ABS / INCOME_FILE
    empl_path = ABS / EMPL_FILE

    # Probe Income workbook for median household income
    income_info = probe_workbook(
        income_path,
        "INCOME (target: median household income, weekly)",
        target_patterns=[
            "median total household income",
            "median household income",
            "median weekly household income",
            "median tot hhd",
        ],
    )

    # Probe Education and Employment workbook for female LFP
    empl_info = probe_workbook(
        empl_path,
        "EDUCATION & EMPLOYMENT (target: female labour force participation)",
        target_patterns=[
            "labour force participation",
            "participation rate",
            "in the labour force",
        ],
    )

    # audit_log + target table check
    print(banner("AUDIT_LOG + TARGET TABLE"))
    if not DB.exists():
        print(f"FAIL: DB not found at {DB}")
        return 1
    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    cur = conn.cursor()
    cur.execute("SELECT COALESCE(MAX(audit_id), 0) FROM audit_log")
    last_audit = cur.fetchone()[0]
    print(f"Last audit_id: {last_audit} (next: {last_audit + 1})")
    cur.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='census2021_sa2'
    """)
    if cur.fetchone():
        cur.execute("SELECT COUNT(*) FROM census2021_sa2")
        n = cur.fetchone()[0]
        print(f"  census2021_sa2: EXISTS ({n:,} rows)")
    else:
        print("  census2021_sa2: does not exist (apply will create)")
    conn.close()

    # Write column map
    column_map = {
        "header_row": HEADER_ROW,
        "data_start_row": HEADER_ROW + 2,  # row 8 per Standard 19
        "income": income_info,
        "empl": empl_info,
        "null_markers": ["-", "..", "np", "n.a.", "na", "NA", "NP", ""],
        "notes": [
            "Apply will pick the BEST matching column per workbook.",
            "Selection rule: header-text match + sample-value range check.",
            "  Income: expect weekly $ values 500-3000 typically.",
            "  Female LFP: expect % values 30-80 typically.",
            "Decision 50: validate by sample value range, not header alone.",
        ],
    }
    OUT_JSON.write_text(json.dumps(column_map, indent=2, default=str),
                        encoding="utf-8")

    findings.append("## Probe results\n")
    findings.append("```json")
    findings.append(json.dumps(column_map, indent=2, default=str))
    findings.append("```\n")
    OUT_MD.write_text("\n".join(findings), encoding="utf-8")

    print(banner("DONE"))
    print(f"  Markdown: {OUT_MD}")
    print(f"  JSON:     {OUT_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
