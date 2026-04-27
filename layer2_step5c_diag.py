"""
Layer 2 Step 5c DIAGNOSTIC — JSA IVI layout discovery
─────────────────────────────────────────────────────────
READ-ONLY. No DB mutations. No file writes outside recon/.

Probes BOTH JSA IVI workbooks (state and remoteness) to discover:
  - Sheet names and structure
  - Header row + first data row
  - ANZSCO code column + region column
  - Monthly time columns (date format)
  - Suppression/null markers
  - audit_log next id
  - Existence check on target tables

Target codes: ANZSCO 4211 (Child Carers), 2411 (Early Childhood
Pre-primary School Teachers).

Outputs:
  - Console summary
  - recon/layer2_step5c_diag.md
  - recon/layer2_step5c_column_map.json
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
DB = PROJECT_ROOT / "data" / "kintell.db"
ABS = PROJECT_ROOT / "abs_data"
OUT_DIR = PROJECT_ROOT / "recon"
OUT_MD = OUT_DIR / "layer2_step5c_diag.md"
OUT_JSON = OUT_DIR / "layer2_step5c_column_map.json"

STATE_PREFIX = "internet_vacancies_anzsco4_occupations_states"
REMOTE_PREFIX = "internet_vacancies_anzsco4_occupations_jsa_remoteness"

TARGET_ANZSCO = ["4211", "2411"]


def banner(title: str) -> str:
    return f"\n{'-' * 70}\n{title}\n{'-' * 70}"


def find_workbook(prefix: str) -> Path | None:
    matches = [p for p in ABS.glob("*.xlsx")
               if p.name.lower().startswith(prefix.lower())]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def is_anzsco_4digit(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return s.isdigit() and len(s) == 4


def looks_like_month_header(v) -> tuple[bool, str | None]:
    """Detect month/date headers. Return (matches, normalised 'YYYY-MM')."""
    if v is None:
        return (False, None)
    if isinstance(v, datetime):
        return (True, v.strftime("%Y-%m"))
    if isinstance(v, date):
        return (True, v.strftime("%Y-%m"))
    s = str(v).strip()
    # YYYY-MM
    m = re.match(r"^(19|20)\d{2}-(0[1-9]|1[0-2])$", s)
    if m:
        return (True, s)
    # YYYY-MM-DD
    m = re.match(r"^((19|20)\d{2})-(0[1-9]|1[0-2])-(0[1-9]|[12]\d|3[01])", s)
    if m:
        return (True, s[:7])
    # MMM-YY or MMM-YYYY (e.g. Mar-26, Mar-2026)
    m = re.match(r"^([A-Za-z]{3})[-\s](\d{2}|\d{4})$", s)
    if m:
        mon_map = {"jan": "01", "feb": "02", "mar": "03",
                   "apr": "04", "may": "05", "jun": "06",
                   "jul": "07", "aug": "08", "sep": "09",
                   "oct": "10", "nov": "11", "dec": "12"}
        mon = mon_map.get(m.group(1).lower())
        yr = m.group(2)
        if mon and yr:
            if len(yr) == 2:
                yr = "20" + yr  # assume 21st century
            return (True, f"{yr}-{mon}")
    return (False, None)


def probe_workbook(wb_path: Path, label: str) -> dict:
    """Return a structure summary for one workbook."""
    print(banner(f"PROBE: {label} — {wb_path.name}"))
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheets: {sheet_names}")

    sheet_info: dict[str, dict] = {}

    for sn in sheet_names:
        print(f"\n  === Sheet: '{sn}' ===")
        ws = wb[sn]

        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if i >= 15:
                break

        n_cols = max(len(r) for r in rows) if rows else 0
        print(f"    Total cols (first 15 rows): {n_cols}")

        # Print first 8 rows × first 12 cols
        print(f"    First 8 rows, cols 0-11:")
        for r_idx, row in enumerate(rows[:8], start=1):
            cells = []
            for c in row[:12]:
                cells.append("·" if c is None else str(c)[:18])
            print(f"      Row {r_idx:2d}: " + " | ".join(cells))

        # Detect ANZSCO code column: scan first 50 data rows
        anzsco_col_candidates: dict[int, int] = {}
        ws_full = wb[sn]
        scanned = 0
        for r_idx, row in enumerate(ws_full.iter_rows(values_only=True),
                                     start=1):
            if r_idx > 60:
                break
            for c_idx in range(min(8, len(row))):
                if is_anzsco_4digit(row[c_idx]):
                    anzsco_col_candidates[c_idx] = (
                        anzsco_col_candidates.get(c_idx, 0) + 1)
            scanned += 1
        if anzsco_col_candidates:
            anzsco_col = max(anzsco_col_candidates,
                             key=lambda k: anzsco_col_candidates[k])
            print(f"    ANZSCO code col: {anzsco_col} "
                  f"(hits in first 60 rows: "
                  f"{anzsco_col_candidates[anzsco_col]})")
        else:
            anzsco_col = None
            print("    ANZSCO code col: NOT FOUND")

        # Detect header row: row with most month-shaped cells
        header_row = None
        max_m = 0
        for r_idx, row in enumerate(rows, start=1):
            m_count = 0
            for c in row:
                is_m, _ = looks_like_month_header(c)
                if is_m:
                    m_count += 1
            if m_count > max_m:
                max_m = m_count
                header_row = r_idx
        print(f"    Header row: {header_row} (had {max_m} month-shaped "
              f"cells)")

        # Find first/last month columns
        month_cols: dict[int, str] = {}
        first_month_col = None
        if header_row:
            hdr = rows[header_row - 1]
            for c_idx, val in enumerate(hdr):
                is_m, normalised = looks_like_month_header(val)
                if is_m:
                    month_cols[c_idx] = normalised
                    if first_month_col is None:
                        first_month_col = c_idx

        if month_cols:
            sorted_cols = sorted(month_cols.keys())
            first_col = sorted_cols[0]
            last_col = sorted_cols[-1]
            print(f"    Month cols: {len(month_cols)} "
                  f"({month_cols[first_col]} -> {month_cols[last_col]})")

            # Print sample of 5 ANZSCO=4211 rows
            print(f"    Sample rows for ANZSCO 4211:")
            ws3 = wb[sn]
            samples = 0
            for r_idx, row in enumerate(ws3.iter_rows(values_only=True),
                                         start=1):
                if r_idx <= (header_row or 0):
                    continue
                if anzsco_col is None or anzsco_col >= len(row):
                    continue
                code = row[anzsco_col]
                if str(code).strip() != "4211":
                    continue
                first_val = row[first_col] if first_col < len(row) else None
                last_val = row[last_col] if last_col < len(row) else None
                # Show columns 0 to anzsco_col + 1 to see all metadata
                metadata = [str(c)[:20] if c is not None else "·"
                            for c in row[:anzsco_col + 2]]
                print(f"      r{r_idx}: {' | '.join(metadata)}  ... "
                      f"{month_cols[first_col]}={first_val}  "
                      f"{month_cols[last_col]}={last_val}")
                samples += 1
                if samples >= 5:
                    break
            if samples == 0:
                print(f"      (no rows found with ANZSCO 4211)")
        else:
            print(f"    Month cols: 0")
            first_col, last_col = None, None

        sheet_info[sn] = {
            "header_row": header_row,
            "data_start_row": (header_row or 0) + 1,
            "anzsco_col": anzsco_col,
            "month_cols": month_cols,
            "n_cols": n_cols,
        }

    wb.close()
    return {
        "file": wb_path.name,
        "sheets": sheet_info,
    }


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    findings: list[str] = ["# Layer 2 Step 5c — JSA IVI Diagnostic", ""]

    # Locate both workbooks
    state_path = find_workbook(STATE_PREFIX)
    remote_path = find_workbook(REMOTE_PREFIX)

    print(banner("WORKBOOK LOCATIONS"))
    if state_path is None:
        print(f"FAIL: state IVI workbook not found "
              f"(prefix: '{STATE_PREFIX}')")
        return 1
    if remote_path is None:
        print(f"WARN: remoteness IVI workbook not found "
              f"(prefix: '{REMOTE_PREFIX}')")
        print("      Will probe state file only.")
    print(f"State file:      {state_path.name}")
    if remote_path:
        print(f"Remoteness file: {remote_path.name}")

    findings.append("## Workbooks\n")
    findings.append(f"- State: `{state_path.name}`")
    if remote_path:
        findings.append(f"- Remoteness: `{remote_path.name}`")
    findings.append("")

    if not DB.exists():
        print(f"FAIL: DB not found at {DB}")
        return 1

    # Probe both
    state_info = probe_workbook(state_path, "STATE")
    remote_info = probe_workbook(remote_path, "REMOTENESS") if remote_path else None

    # audit_log probe
    print(banner("AUDIT_LOG + TARGET TABLES"))
    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    cur = conn.cursor()

    cur.execute("SELECT COALESCE(MAX(audit_id), 0) FROM audit_log")
    last_audit = cur.fetchone()[0]
    print(f"Last audit_id: {last_audit} (next will be {last_audit + 1})")

    for table in ("jsa_ivi_state_monthly", "jsa_ivi_remoteness_monthly"):
        cur.execute("""
            SELECT name FROM sqlite_master
            WHERE type='table' AND name=?
        """, (table,))
        if cur.fetchone():
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            n = cur.fetchone()[0]
            print(f"  {table}: EXISTS ({n:,} rows). Apply will refuse "
                  f"if non-empty.")
        else:
            print(f"  {table}: does not exist (apply will create)")
    conn.close()

    # Write JSON column map
    column_map = {
        "state_workbook": str(state_path.relative_to(PROJECT_ROOT)),
        "remote_workbook": (str(remote_path.relative_to(PROJECT_ROOT))
                            if remote_path else None),
        "state_sheets": state_info["sheets"],
        "remote_sheets": (remote_info["sheets"] if remote_info else None),
        "target_anzsco": TARGET_ANZSCO,
        "null_markers": ["np", "n.a.", "NA", "NP", "-", "..", "0", ""],
        "notes": [
            "Filter to anzsco_code IN ('4211', '2411') only.",
            "Each metric file has multiple sheets (likely original / "
            "trend / seasonally adjusted). Diagnostic shows all; apply "
            "will pick the most useful one.",
            "Months are stored as Excel datetime in headers.",
        ],
    }
    OUT_JSON.write_text(json.dumps(column_map, indent=2, default=str),
                        encoding="utf-8")

    findings.append("## Probe results\n")
    findings.append("```json")
    findings.append(json.dumps(column_map, indent=2, default=str))
    findings.append("```\n")
    OUT_MD.write_text("\n".join(findings), encoding="utf-8")

    print(banner("DONE"))
    print(f"  Markdown: {OUT_MD}")
    print(f"  JSON:     {OUT_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
