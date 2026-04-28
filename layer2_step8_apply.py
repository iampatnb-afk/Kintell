"""
layer2_step8_apply.py — ABS Births SA2 ingest.

Modes:
  (default)  Extract + validate + write dry-run report; NO DB mutation.
  --apply    Same + DB mutation inside transaction with invariants.

Source:
  abs_data/Births_SA2_2011_2024.xlsx
  Sheets: Contents, Table 1..8 (per-state), Further information
  Header layout per Standard 19:
    row 5 — year span header (2011..2024 at cols 2, 5, 8, ..., 41)
    row 6 — sub-header (ERP | Births | TFR repeated per year)
    row 7 — unit row (persons | no. | rate)
    row 8+ — data; SA2 rows have 9-digit code in col 0
  Births = year_col + 1 (verified from sub-header at runtime, not assumed).

Outputs:
  recon/layer2_step8_apply_dryrun.md   (dry-run only)
  recon/layer2_step8_apply.md          (apply only)
  data/kintell.db.backup_pre_step8_<ts>  (apply only)
  audit_log row, action='abs_sa2_births_ingest_v1'   (apply only)

Schema:
  CREATE TABLE abs_sa2_births_annual (
    sa2_code     TEXT NOT NULL,
    year         INTEGER NOT NULL,
    births_count INTEGER,           -- NULL for 'np' or non-numeric
    PRIMARY KEY (sa2_code, year)
  );

Invariants (apply mode; ROLLBACK on violation):
  - 2,000 <= distinct SA2 count <= 2,500
  - year coverage exactly 2011..2024
  - per-year national sum within +/-5% of ABS Cat 3301.0 published baseline
  - 0 SA2 codes outside abs_sa2_erp_annual canonical universe
  - all births_count values are NULL or non-negative integer
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

DB_PATH = "data/kintell.db"
XLSX_PATH = "abs_data/Births_SA2_2011_2024.xlsx"
TARGET_TABLE = "abs_sa2_births_annual"
AUDIT_ACTION = "abs_sa2_births_ingest_v1"
ACTOR = "layer2_step8_apply"
DRYRUN_REPORT = "recon/layer2_step8_apply_dryrun.md"
APPLY_REPORT = "recon/layer2_step8_apply.md"

DATA_HEADER_ROW_YEAR = 5      # 1-indexed
DATA_HEADER_ROW_SUB = 6       # 1-indexed
DATA_FIRST_ROW = 8            # 1-indexed
TABLE_SHEETS = [f"Table {i}" for i in range(1, 9)]

SA2_CODE_RE = re.compile(r"^\d{9}$")

ABS_NATIONAL_BIRTHS = {
    2011: 297900, 2012: 309600, 2013: 308100, 2014: 299700,
    2015: 305400, 2016: 311100, 2017: 309100, 2018: 315100,
    2019: 305800, 2020: 294400, 2021: 309900, 2022: 300700,
    2023: 286900, 2024: 292500,
}
TOLERANCE_PCT = 5.0
SA2_MIN, SA2_MAX = 2000, 2500
EXPECTED_YEARS = set(range(2011, 2025))


# ---------------------------------------------------------------------------
# Preflight
# ---------------------------------------------------------------------------
for p in (DB_PATH, XLSX_PATH):
    if not Path(p).exists():
        print(f"ERROR: not found: {p}", file=sys.stderr)
        sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    print("Run: pip install --break-system-packages openpyxl", file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def is_sa2(v) -> bool:
    if v is None:
        return False
    return bool(SA2_CODE_RE.match(str(v).strip()))


def looks_like_year(v) -> bool:
    try:
        n = int(v)
        return 2000 <= n <= 2030
    except (TypeError, ValueError):
        return False


def coerce_births(v):
    """Return (int_value_or_None, status_tag).
    status_tag in {'ok', 'np', 'null', 'bad'}."""
    if v is None:
        return None, "null"
    if isinstance(v, (int, float)):
        if v != v:  # NaN
            return None, "null"
        try:
            iv = int(v)
            return iv, "ok"
        except (ValueError, OverflowError):
            return None, "bad"
    s = str(v).strip().lower()
    if s in ("", "np", "n.p.", "..", "n/a", "na"):
        return None, "np"
    try:
        return int(float(s)), "ok"
    except ValueError:
        return None, "bad"


# ---------------------------------------------------------------------------
# Detect (year -> births_col_idx) mapping for a sheet
# ---------------------------------------------------------------------------
def detect_birth_columns(ws) -> dict[int, int]:
    """For one state sheet, return {year: 0-based-col-idx-of-births}.
    Reads row 5 (year header) and row 6 (sub-header), confirms by matching
    'Birth' (case-insensitive) in row 6 within +/- 2 cols of the year col.
    """
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=DATA_HEADER_ROW_SUB,
                                          values_only=True)):
        rows.append(list(row))
    if len(rows) < DATA_HEADER_ROW_SUB:
        return {}

    year_row = rows[DATA_HEADER_ROW_YEAR - 1]  # 0-indexed access
    sub_row = rows[DATA_HEADER_ROW_SUB - 1]

    out: dict[int, int] = {}
    for c_idx, val in enumerate(year_row):
        if not looks_like_year(val):
            continue
        year = int(val)
        # Look at sub_row in cols [c_idx, c_idx+1, c_idx+2]
        for offset in (0, 1, 2):
            sc = c_idx + offset
            if sc >= len(sub_row):
                break
            sv = sub_row[sc]
            if sv is None:
                continue
            if "birth" in str(sv).strip().lower():
                out[year] = sc
                break
    return out


# ---------------------------------------------------------------------------
# Extract all (sa2_code, year, births) tuples
# ---------------------------------------------------------------------------
def extract_all(wb):
    """Returns dict with keys:
      records: list of (sa2_code:str, year:int, births:int|None)
      per_table_stats: dict[sheet] -> dict
      np_count, null_count, bad_count: ints
      detection: dict[sheet] -> {year: birth_col_idx}
    """
    records = []
    per_table = {}
    detection = {}
    np_count = 0
    null_count = 0
    bad_count = 0

    for sheet in TABLE_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"WARN: sheet '{sheet}' not found, skipping", file=sys.stderr)
            continue
        ws = wb[sheet]
        col_map = detect_birth_columns(ws)
        detection[sheet] = col_map
        if not col_map:
            print(f"WARN: no Births columns detected in '{sheet}'", file=sys.stderr)
            per_table[sheet] = {"sa2_rows": 0, "records": 0, "years": []}
            continue

        sa2_rows_seen = 0
        records_in_sheet = 0
        for row in ws.iter_rows(min_row=DATA_FIRST_ROW, values_only=True):
            if not row:
                continue
            code_cell = row[0] if len(row) > 0 else None
            if not is_sa2(code_cell):
                continue
            sa2_code = str(code_cell).strip()
            sa2_rows_seen += 1
            for year, c_idx in col_map.items():
                if c_idx >= len(row):
                    continue
                raw = row[c_idx]
                val, tag = coerce_births(raw)
                if tag == "np":
                    np_count += 1
                elif tag == "null":
                    null_count += 1
                elif tag == "bad":
                    bad_count += 1
                records.append((sa2_code, year, val))
                records_in_sheet += 1
        per_table[sheet] = {
            "sa2_rows": sa2_rows_seen,
            "records": records_in_sheet,
            "years": sorted(col_map.keys()),
            "year_count": len(col_map),
        }

    return {
        "records": records,
        "per_table": per_table,
        "detection": detection,
        "np_count": np_count,
        "null_count": null_count,
        "bad_count": bad_count,
    }


# ---------------------------------------------------------------------------
# Sanity / invariants
# ---------------------------------------------------------------------------
def evaluate(extracted: dict, conn) -> dict:
    records = extracted["records"]
    distinct_sa2 = set(r[0] for r in records)
    years_seen = set(r[1] for r in records)

    # National sum by year (only ok values)
    sum_by_year: dict[int, int] = {}
    cnt_by_year: dict[int, int] = {}
    for sa2, year, val in records:
        cnt_by_year[year] = cnt_by_year.get(year, 0) + 1
        if val is not None:
            sum_by_year[year] = sum_by_year.get(year, 0) + val

    # Canonical universe check
    canonical = set()
    cur = conn.cursor()
    try:
        rows = cur.execute(
            "SELECT DISTINCT sa2_code FROM abs_sa2_erp_annual "
            "WHERE sa2_code IS NOT NULL"
        ).fetchall()
        canonical = set(str(r[0]) for r in rows)
    except sqlite3.Error as e:
        print(f"WARN: canonical universe lookup failed: {e}", file=sys.stderr)
    orphans = distinct_sa2 - canonical if canonical else set()

    # Negative births
    n_negative = sum(1 for _, _, v in records if v is not None and v < 0)

    # Per-year sanity
    per_year_check = []
    sanity_violations = []
    for year in sorted(EXPECTED_YEARS):
        observed = sum_by_year.get(year)
        expected = ABS_NATIONAL_BIRTHS.get(year)
        if observed is None or expected is None:
            ok = False
            pct = None
        else:
            pct = (observed - expected) / expected * 100
            ok = abs(pct) <= TOLERANCE_PCT
        per_year_check.append({
            "year": year, "observed": observed, "expected": expected,
            "pct_delta": pct, "ok": ok,
        })
        if not ok:
            label = (f"{year}: sum={observed} vs ABS {expected} "
                     f"(delta {pct:+.1f}%)" if observed is not None
                     else f"{year}: missing")
            sanity_violations.append(label)

    violations = []
    if not (SA2_MIN <= len(distinct_sa2) <= SA2_MAX):
        violations.append(
            f"distinct SA2 count {len(distinct_sa2)} outside [{SA2_MIN}, {SA2_MAX}]"
        )
    if years_seen != EXPECTED_YEARS:
        missing = EXPECTED_YEARS - years_seen
        extra = years_seen - EXPECTED_YEARS
        if missing:
            violations.append(f"missing years: {sorted(missing)}")
        if extra:
            violations.append(f"unexpected years: {sorted(extra)}")
    if orphans:
        violations.append(
            f"{len(orphans)} SA2 codes outside abs_sa2_erp_annual"
        )
    if n_negative:
        violations.append(f"{n_negative} negative births_count values")
    if sanity_violations:
        violations.append("per-year national-sum tolerance failures: "
                          + "; ".join(sanity_violations))

    return {
        "distinct_sa2": distinct_sa2,
        "years_seen": years_seen,
        "sum_by_year": sum_by_year,
        "cnt_by_year": cnt_by_year,
        "canonical_size": len(canonical),
        "orphans": sorted(orphans),
        "n_negative": n_negative,
        "per_year_check": per_year_check,
        "violations": violations,
    }


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------
def write_report(out_path: str, mode: str, extracted: dict, results: dict,
                 backup_path: str | None, audit_id: int | None,
                 audit_sql: str, audit_params: list) -> None:
    lines: list[str] = []

    def w(s: str = "") -> None:
        lines.append(s)

    title = "DRY RUN" if mode == "dry-run" else "APPLIED"
    w(f"# Layer 2 Step 8 — Apply ({title})")
    w("")
    w(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    w(f"Mode: **{mode}**")
    if backup_path:
        w(f"Backup: `{backup_path}`")
    w("")

    records = extracted["records"]
    n_records = len(records)
    n_with_value = sum(1 for _, _, v in records if v is not None)

    w("## Summary")
    w("")
    w(f"- Source: `{XLSX_PATH}`")
    w(f"- Source sheets read: {len([s for s in TABLE_SHEETS if s in extracted['per_table']])}")
    w(f"- Total records extracted: **{n_records:,}**")
    w(f"- With numeric births_count: **{n_with_value:,}**")
    w(f"- 'np' (confidentialized): {extracted['np_count']:,}")
    w(f"- NULL: {extracted['null_count']:,}")
    w(f"- Unparseable: {extracted['bad_count']:,}")
    w(f"- Distinct SA2 codes: **{len(results['distinct_sa2']):,}** "
      f"(invariant: {SA2_MIN:,}–{SA2_MAX:,})")
    w(f"- Years covered: **{sorted(results['years_seen'])}**")
    w(f"- Negative births_count values: {results['n_negative']}")
    w(f"- Canonical universe (abs_sa2_erp_annual): "
      f"{results['canonical_size']:,} SA2s")
    w(f"- Orphan SA2s (not in canonical): {len(results['orphans'])}")
    w("")

    w("## Invariants")
    w("")
    if results["violations"]:
        w("**✗ VIOLATIONS — apply will be refused:**")
        w("")
        for v in results["violations"]:
            w(f"- {v}")
    else:
        w("✓ All invariants pass.")
        w(f"- distinct SA2 count: {len(results['distinct_sa2'])} ∈ "
          f"[{SA2_MIN}, {SA2_MAX}]")
        w(f"- years exactly cover 2011..2024")
        w(f"- 0 orphan SA2s")
        w(f"- 0 negative births")
        w(f"- per-year national sum within ±{TOLERANCE_PCT:.0f}% of ABS")
    w("")

    # Per-year national-sum table
    w("## Per-year national sum vs ABS Cat 3301.0")
    w("")
    w("| year | SA2-sum | ABS published | delta | within ±5%? |")
    w("|---:|---:|---:|---:|:-:|")
    for entry in results["per_year_check"]:
        obs = entry["observed"]
        exp = entry["expected"]
        pct = entry["pct_delta"]
        mark = "✓" if entry["ok"] else "✗"
        obs_s = f"{obs:,}" if obs is not None else "—"
        exp_s = f"{exp:,}" if exp is not None else "—"
        delta_s = f"{(obs-exp):+,} ({pct:+.1f}%)" if obs is not None and exp is not None else "—"
        w(f"| {entry['year']} | {obs_s} | {exp_s} | {delta_s} | {mark} |")
    w("")

    # Per-table breakdown
    w("## Per-state table breakdown")
    w("")
    w("| sheet | SA2 rows | records | year columns detected |")
    w("|---|---:|---:|---:|")
    for sheet in TABLE_SHEETS:
        info = extracted["per_table"].get(sheet)
        if not info:
            w(f"| `{sheet}` | (skipped) | — | — |")
            continue
        w(f"| `{sheet}` | {info['sa2_rows']:,} | {info['records']:,} "
          f"| {info.get('year_count', 0)} |")
    w("")

    # Detection map detail (one sheet)
    sample_sheet = next((s for s in TABLE_SHEETS
                         if s in extracted["detection"]
                         and extracted["detection"][s]), None)
    if sample_sheet:
        w(f"### Births column detection (sample: `{sample_sheet}`)")
        w("")
        w(f"Year → 0-based column index of Births")
        w("")
        w("| year | births_col | year_col |")
        w("|---:|---:|---:|")
        for year in sorted(extracted["detection"][sample_sheet]):
            bc = extracted["detection"][sample_sheet][year]
            w(f"| {year} | {bc} | {bc - 1} |")
        w("")

    # Sample records
    w("## Sample records (first 10 by SA2 / year)")
    w("")
    sample = sorted(records, key=lambda r: (r[0], r[1]))[:10]
    if sample:
        w("| sa2_code | year | births_count |")
        w("|---|---:|---:|")
        for sa2, year, v in sample:
            w(f"| {sa2} | {year} | {v if v is not None else 'NULL'} |")
    w("")

    # Audit
    w("## Audit log")
    w("")
    if mode == "dry-run":
        w(f"Would insert: action=`{AUDIT_ACTION}`, actor=`{ACTOR}`, "
          f"subject_type=`{TARGET_TABLE}`")
        w("")
        w("Simulated INSERT:")
        w("")
        w("```sql")
        w(audit_sql)
        w("```")
        w("")
        w("Params:")
        w("")
        w("```")
        for i, p in enumerate(audit_params):
            display = str(p)
            if len(display) > 300:
                display = display[:297] + "..."
            w(f"  [{i}] {display}")
        w("```")
    else:
        w(f"Inserted: action=`{AUDIT_ACTION}`, audit_id=**{audit_id}**")
    w("")

    w("---")
    w("End of report.")
    w("")

    Path("recon").mkdir(exist_ok=True)
    Path(out_path).write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Audit log payload
# ---------------------------------------------------------------------------
def build_audit_insert(extracted: dict, results: dict, table_existed: bool):
    n_records = len(extracted["records"])
    n_with_value = sum(1 for _, _, v in extracted["records"] if v is not None)
    sums_str = {str(k): v for k, v in sorted(results["sum_by_year"].items())}

    before = {
        "rows": 0 if not table_existed else None,
        "payload": {
            "table": TARGET_TABLE,
            "table_existed_pre": table_existed,
        },
    }
    after = {
        "rows": n_records,
        "payload": {
            "table": TARGET_TABLE,
            "rows_inserted": n_records,
            "rows_with_value": n_with_value,
            "rows_null": n_records - n_with_value,
            "distinct_sa2": len(results["distinct_sa2"]),
            "year_min": min(results["years_seen"]) if results["years_seen"] else None,
            "year_max": max(results["years_seen"]) if results["years_seen"] else None,
            "year_count": len(results["years_seen"]),
            "national_sum_by_year": sums_str,
            "source_file": XLSX_PATH,
            "source_tables": list(extracted["per_table"].keys()),
            "np_skipped": extracted["np_count"],
            "method": "openpyxl iter_rows; year-col + sub-header 'Births' match",
        },
    }
    reason = (
        f"Layer 2 Step 8: SA2 births annual ingest. "
        f"{n_records:,} records ({n_with_value:,} with numeric values, "
        f"{extracted['np_count']:,} confidentialized) across "
        f"{len(results['distinct_sa2']):,} SA2s and "
        f"{len(results['years_seen'])} years "
        f"({min(results['years_seen'])}-{max(results['years_seen'])}). "
        f"Source: {XLSX_PATH} (Tables 1-8, per-state). "
        f"Per-year national sums within ±{TOLERANCE_PCT:.0f}% of ABS Cat 3301.0."
    )

    sql = (
        "INSERT INTO audit_log "
        "(actor, action, subject_type, subject_id, before_json, after_json, reason) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)"
    )
    params = [
        ACTOR,
        AUDIT_ACTION,
        TARGET_TABLE,
        0,
        json.dumps(before, sort_keys=True),
        json.dumps(after, sort_keys=True),
        reason,
    ]
    return sql, params


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------
def take_backup(backup_path: str) -> None:
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(backup_path)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="Mutate DB. Default is dry-run.")
    ap.add_argument("--dry-run", action="store_true",
                    help="(default) Compute only; no DB mutation.")
    args = ap.parse_args()
    apply_mode = bool(args.apply)
    mode = "apply" if apply_mode else "dry-run"

    print(f"Mode: {mode.upper()}")
    print(f"Loading {XLSX_PATH} ...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

    print("Extracting records from per-state sheets ...")
    extracted = extract_all(wb)
    wb.close()
    print(f"  records extracted: {len(extracted['records']):,}")
    print(f"  np-skipped: {extracted['np_count']:,}, "
          f"null: {extracted['null_count']:,}, "
          f"bad: {extracted['bad_count']:,}")

    print("Connecting to DB (read-only for compute) ...")
    try:
        conn = sqlite3.connect(f"file:{DB_PATH}?mode=ro", uri=True)
    except sqlite3.Error:
        conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Check whether target table already exists (for audit before/after)
    cur = conn.cursor()
    table_existed = bool(cur.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (TARGET_TABLE,)
    ).fetchone())

    print("Evaluating sanity / invariants ...")
    results = evaluate(extracted, conn)
    print(f"  distinct SA2: {len(results['distinct_sa2']):,}")
    print(f"  years: {sorted(results['years_seen'])}")
    if results["violations"]:
        print("  ⚠ violations:")
        for v in results["violations"]:
            print(f"    - {v}")
    else:
        print("  ✓ all invariants pass")

    audit_sql, audit_params = build_audit_insert(extracted, results, table_existed)
    conn.close()

    # ----- DRY RUN -----
    if not apply_mode:
        write_report(DRYRUN_REPORT, "dry-run", extracted, results,
                     None, None, audit_sql, audit_params)
        print()
        print(f"DRY-RUN OK  wrote {DRYRUN_REPORT}")
        if results["violations"]:
            print("⚠ Invariant violations present — apply would be refused.")
        else:
            print("✓ All invariants pass. To execute:")
            print("    python layer2_step8_apply.py --apply")
        return 0

    # ----- APPLY -----
    if results["violations"]:
        print("\n✗ INVARIANT VIOLATIONS — refusing to apply:", file=sys.stderr)
        for v in results["violations"]:
            print(f"  - {v}", file=sys.stderr)
        return 2

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"data/kintell.db.backup_pre_step8_{ts}"
    print(f"\nTaking backup: {backup_path}")
    take_backup(backup_path)
    backup_size_mb = os.path.getsize(backup_path) / (1024 * 1024)
    print(f"  backup size: {backup_size_mb:,.1f} MB")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    audit_id = None
    try:
        cur.execute("BEGIN")

        cur.execute(f'DROP TABLE IF EXISTS "{TARGET_TABLE}"')
        cur.execute(f'''
            CREATE TABLE "{TARGET_TABLE}" (
              sa2_code     TEXT NOT NULL,
              year         INTEGER NOT NULL,
              births_count INTEGER,
              PRIMARY KEY (sa2_code, year)
            )
        ''')
        cur.execute(f'CREATE INDEX idx_{TARGET_TABLE}_sa2  '
                    f'ON "{TARGET_TABLE}"(sa2_code)')
        cur.execute(f'CREATE INDEX idx_{TARGET_TABLE}_year '
                    f'ON "{TARGET_TABLE}"(year)')

        # Bulk insert
        cur.executemany(
            f'INSERT INTO "{TARGET_TABLE}" (sa2_code, year, births_count) '
            f'VALUES (?, ?, ?)',
            extracted["records"]
        )
        n_inserted = len(extracted["records"])
        print(f"  INSERTed {n_inserted:,} rows")

        # Re-verify post-insert
        n_rows_db = cur.execute(
            f'SELECT COUNT(*) FROM "{TARGET_TABLE}"'
        ).fetchone()[0]
        if n_rows_db != n_inserted:
            raise RuntimeError(
                f"Post-insert row count {n_rows_db} != extracted {n_inserted}"
            )
        print(f"  post-insert row count: {n_rows_db:,} (matches)")

        # Re-verify orphans (defensive — same check that passed in dry run)
        bad = cur.execute(
            f'SELECT COUNT(DISTINCT sa2_code) FROM "{TARGET_TABLE}" '
            f'WHERE sa2_code NOT IN '
            f'(SELECT sa2_code FROM abs_sa2_erp_annual WHERE sa2_code IS NOT NULL)'
        ).fetchone()[0]
        if bad > 0:
            raise RuntimeError(
                f"{bad} SA2 codes in ingested table outside canonical universe"
            )
        print(f"  canonical-universe re-check: OK (0 orphans)")

        # Audit log
        cur.execute(audit_sql, audit_params)
        audit_id = cur.lastrowid
        print(f"  audit_log row inserted: audit_id={audit_id}")

        cur.execute("COMMIT")
        print("✓ COMMITTED")
    except Exception as e:
        try:
            cur.execute("ROLLBACK")
        except sqlite3.Error:
            pass
        print(f"\n✗ ROLLED BACK: {e}", file=sys.stderr)
        conn.close()
        return 3
    finally:
        try:
            conn.close()
        except Exception:
            pass

    write_report(APPLY_REPORT, "apply", extracted, results,
                 backup_path, audit_id, audit_sql, audit_params)
    print()
    print(f"APPLY OK  wrote {APPLY_REPORT}")
    print(f"          backup: {backup_path}")
    print(f"          audit_id: {audit_id}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
