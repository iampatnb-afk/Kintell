"""
Layer 2 Step 5b-prime — Structure probe (read-only).
=====================================================
Quick diagnostic to locate WHERE SA2 codes actually live in
the Education and Employment Database workbook.

Prints:
  - First 15 rows x first 8 columns of Table 1 and Table 2
  - The column index containing 9-digit codes (if any)
  - A few rows of SA2-level data from that column outward

No file output. Console only.
"""

import re
from pathlib import Path
import openpyxl

WORKBOOK = Path("abs_data/Education and employment database.xlsx")
SA2_CODE_RE = re.compile(r"^\s*\d{9}\s*$")


def trunc(v, n=22):
    if v is None:
        return ""
    s = str(v).replace("\n", " ").strip()
    return s[:n] + ("…" if len(s) > n else "")


def dump_top(ws, rows=15, cols=8):
    print(f"\n--- {ws.title}: first {rows} rows x {cols} cols ---")
    for r in range(1, rows + 1):
        cells = []
        for c in range(1, cols + 1):
            cells.append(f"{trunc(ws.cell(row=r, column=c).value):<24}")
        print(f"  r{r:>3}: " + " | ".join(cells))


def find_code_column(ws, max_scan_rows=300, max_cols=10):
    """Find which column index holds 9-digit SA2 codes."""
    hits = {}  # col -> count
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if SA2_CODE_RE.match(str(v)):
                hits[c] = hits.get(c, 0) + 1
    return hits


def dump_sa2_sample(ws, code_col, n=5, cols=8):
    print(f"\n--- {ws.title}: first {n} SA2-level rows (code col = {code_col}) ---")
    found = 0
    for r in range(1, min(ws.max_row, 4000) + 1):
        v = ws.cell(row=r, column=code_col).value
        if v is None or not SA2_CODE_RE.match(str(v)):
            continue
        cells = [f"{trunc(ws.cell(row=r, column=c).value):<22}"
                 for c in range(1, cols + 1)]
        print(f"  r{r:>4}: " + " | ".join(cells))
        found += 1
        if found >= n:
            break


def main():
    print(f"Opening: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, read_only=False, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    print(f"Note: read_only=False so we can address cells freely.")

    for name in wb.sheetnames:
        if name.lower() == "contents":
            continue
        ws = wb[name]
        print(f"\n========== {name} ==========")
        print(f"max_row={ws.max_row}, max_col={ws.max_column}")
        dump_top(ws, rows=15, cols=8)

        hits = find_code_column(ws)
        if not hits:
            print(f"\n  [!] No 9-digit codes found in cols 1-10 of first 300 rows.")
            print(f"      Will scan wider...")
            # Wider scan
            hits_wide = {}
            for r in range(1, min(ws.max_row, 300) + 1):
                for c in range(1, min(ws.max_column, 30) + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    if SA2_CODE_RE.match(str(v)):
                        hits_wide[c] = hits_wide.get(c, 0) + 1
            if hits_wide:
                print(f"      Wider scan hits: {hits_wide}")
                hits = hits_wide
            else:
                # Maybe codes are integers, not strings — look for 9-digit ints
                int_hits = {}
                for r in range(1, min(ws.max_row, 300) + 1):
                    for c in range(1, min(ws.max_column, 30) + 1):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, int) and 100000000 <= v <= 999999999:
                            int_hits[c] = int_hits.get(c, 0) + 1
                        elif isinstance(v, float) and v == int(v) and 100000000 <= v <= 999999999:
                            int_hits[c] = int_hits.get(c, 0) + 1
                print(f"      Integer 9-digit hits: {int_hits}")
                hits = int_hits

        if hits:
            print(f"\n  9-digit code hits per column: {hits}")
            best_col = max(hits, key=hits.get)
            print(f"  Best guess: code column = {best_col} (letter {openpyxl.utils.get_column_letter(best_col)})")
            dump_sa2_sample(ws, best_col, n=5, cols=10)

    wb.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
