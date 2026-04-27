# layer1_files_inspect.py
# Phase 2.5 Layer 1 Part C — Raw file inventory in abs_data\
# Read-only. Writes recon\layer1_files_findings.md
import os, csv, datetime, sys

ABS = "abs_data"
OUT = os.path.join("recon", "layer1_files_findings.md")
lines = []
def emit(s=""):
    print(s, flush=True)
    lines.append(s)

emit("# Phase 2.5 Layer 1 — Raw File Inventory (abs_data\\)")
emit(f"_Generated: {datetime.datetime.now().isoformat(timespec='seconds')}_")
emit("")

try:
    from openpyxl import load_workbook
except ImportError:
    emit("FATAL: openpyxl not installed. pip install openpyxl")
    sys.exit(1)

def fsize(p):
    return f"{os.path.getsize(p):,} bytes"

# --- File listing ---
emit("## File listing")
for f in sorted(os.listdir(ABS)):
    p = os.path.join(ABS, f)
    if os.path.isfile(p):
        emit(f"  - {f:<70} {os.path.getsize(p):>14,} bytes")
emit("")

# --- NQAITS (the big unlock) ---
nqaits = os.path.join(ABS, "NQAITS Quarterly Data Splits (Q3 2013 - Q4 2025).xlsx")
emit("## NQAITS Quarterly Data Splits — schema reconnaissance")
if os.path.exists(nqaits):
    emit(f"`{nqaits}`  ({fsize(nqaits)})")
    print(f"  [opening NQAITS — 138MB, may take 30-60s ...]", flush=True)
    try:
        wb = load_workbook(nqaits, read_only=True, data_only=True)
        sheets = wb.sheetnames
        emit(f"sheet count: {len(sheets)}")
        emit("sheets (verbatim names):")
        for s in sheets:
            emit(f"  - {s}")
        emit("")
        # First / middle / last
        if len(sheets) >= 3:
            picks = [sheets[0], sheets[len(sheets)//2], sheets[-1]]
        else:
            picks = sheets
        emit(f"### Schema dump — first / middle / last sheet")
        emit(f"Picked: {picks}")
        emit("")
        headers = {}
        for s in picks:
            print(f"  [reading header from sheet: {s}]", flush=True)
            ws = wb[s]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                h = next(rows_iter)
                headers[s] = list(h)
                try:
                    sample = next(rows_iter)
                except StopIteration:
                    sample = None
                emit(f"#### `{s}`  cols={len(h)}")
                for i, c in enumerate(h):
                    emit(f"  {i:3d}. {c}")
                if sample:
                    emit(f"sample row: {list(sample)}")
                emit("")
            except StopIteration:
                emit(f"  [{s} empty]")
        # Cross-sheet stability
        emit("### Column stability across sampled sheets")
        if len(headers) > 1:
            sets = {s: set(str(c) for c in h if c is not None) for s, h in headers.items()}
            common = set.intersection(*sets.values())
            emit(f"COMMON across sampled: {len(common)}")
            for c in sorted(common):
                emit(f"  + {c}")
            emit("")
            for s, cs in sets.items():
                only = cs - common
                emit(f"UNIQUE to `{s}`: {len(only)}")
                for c in sorted(only):
                    emit(f"  - {c}")
            emit("")
        # Service Approval column check
        emit("### Service Approval column presence")
        for s, h in headers.items():
            sa = [str(c) for c in h if c and "service" in str(c).lower() and "approval" in str(c).lower()]
            emit(f"  {s}: {sa}")
        emit("")
        wb.close()
    except Exception as e:
        emit(f"NQAITS ERROR: {e}")
        emit("")
else:
    emit("  MISSING")
    emit("")

# --- SALM ---
salm = os.path.join(ABS, "SALM Smoothed SA2 Datafiles (ASGS 2021) - December quarter 2025.xlsx")
emit("## SALM Smoothed SA2")
if os.path.exists(salm):
    emit(f"`{salm}`  ({fsize(salm)})")
    print(f"  [opening SALM ...]", flush=True)
    try:
        wb = load_workbook(salm, read_only=True, data_only=True)
        emit(f"sheet count: {len(wb.sheetnames)}")
        for s in wb.sheetnames:
            emit(f"  - {s}")
        emit("")
        for s in wb.sheetnames:
            ws = wb[s]
            emit(f"### sheet: `{s}`")
            rows_iter = ws.iter_rows(values_only=True)
            for i, r in enumerate(rows_iter):
                if i >= 6:
                    break
                emit(f"  row {i}: {list(r)}")
            emit("")
        wb.close()
    except Exception as e:
        emit(f"SALM ERROR: {e}")
        emit("")
else:
    emit("  MISSING")
    emit("")

# --- Census 2021 ABS databases ---
census_files = [
    ("Population and People Database.xlsx", "ABS Population and People"),
    ("Family and Community Database.xlsx",  "ABS Family and Community"),
    ("Income Database.xlsx",                "ABS Income"),
    ("Education and employment database.xlsx", "ABS Education and Employment"),
]
for fn, label in census_files:
    p = os.path.join(ABS, fn)
    emit(f"## {label}")
    if not os.path.exists(p):
        emit(f"  MISSING: {fn}")
        emit("")
        continue
    print(f"  [opening {fn} ...]", flush=True)
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        emit(f"`{p}`  ({fsize(p)})")
        emit(f"sheet count: {len(wb.sheetnames)}")
        for s in wb.sheetnames:
            emit(f"  - {s}")
        ws = wb[wb.sheetnames[0]]
        emit(f"### first sheet `{wb.sheetnames[0]}` — first 5 rows (truncated to 20 cols)")
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= 5:
                break
            emit(f"  row {i}: {list(r)[:20]}{'...' if len(r)>20 else ''}")
        emit("")
        wb.close()
    except Exception as e:
        emit(f"ERROR: {e}")
        emit("")

# --- Economic and Industry Database (suspiciously small CSV) ---
ei = os.path.join(ABS, "Economic and Industry Database.csv")
emit("## ABS Economic and Industry Database (CSV — only 2.7KB)")
if os.path.exists(ei):
    emit(f"`{ei}`  ({fsize(ei)})")
    emit("Full contents:")
    emit("```")
    try:
        with open(ei, "r", encoding="utf-8-sig", errors="replace") as f:
            emit(f.read())
    except Exception as e:
        emit(f"ERROR: {e}")
    emit("```")
emit("")

# --- Boundary / concordance files ---
for fn, label in [
    ("SA2_2021_AUST.xlsx", "SA2 boundaries (ABS ASGS 2021)"),
    ("POA_2021_AUST.xlsx", "Postcode boundaries (ABS ASGS 2021)"),
    ("meshblock-correspondence-file-asgs-edn3.xlsx", "Meshblock correspondence (ASGS edn3)"),
]:
    p = os.path.join(ABS, fn)
    emit(f"## {label}")
    if not os.path.exists(p):
        emit(f"  MISSING: {fn}")
        emit("")
        continue
    print(f"  [opening {fn} ...]", flush=True)
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        emit(f"`{p}`  ({fsize(p)})")
        emit(f"sheet count: {len(wb.sheetnames)}")
        for s in wb.sheetnames:
            emit(f"  - {s}")
        ws = wb[wb.sheetnames[0]]
        emit(f"### first sheet `{wb.sheetnames[0]}` — first 3 rows")
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i >= 3:
                break
            emit(f"  row {i}: {list(r)}")
        emit("")
        wb.close()
    except Exception as e:
        emit(f"ERROR: {e}")
        emit("")

# --- postcode_to_sa2_concordance ---
pc = os.path.join(ABS, "postcode_to_sa2_concordance.csv")
emit("## postcode_to_sa2_concordance.csv")
if os.path.exists(pc):
    emit(f"`{pc}`  ({fsize(pc)})")
    try:
        with open(pc, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 6:
                    break
                emit(f"  row {i}: {row}")
    except Exception as e:
        emit(f"ERROR: {e}")
emit("")

# --- Births file search (Layer 2 dependency) ---
emit("## Births data — file search")
emit("Plan: 'Patrick has the file already' (ABS Cat 3301.0).")
emit("Searching for any file containing 'birth' (case-insensitive) at top-level of:")
search_paths = [
    ABS,
    ".",
    "data",
    os.path.join(os.path.expanduser("~"), "Downloads"),
]
found = []
for sp in search_paths:
    emit(f"  - {sp}")
    if not os.path.isdir(sp):
        continue
    try:
        for f in os.listdir(sp):
            if "birth" in f.lower():
                found.append(os.path.join(sp, f))
    except Exception:
        pass
emit("")
if found:
    emit("MATCHES:")
    for f in found:
        emit(f"  + {f}")
else:
    emit("NO MATCHES — births file not present in expected locations.")
emit("")

# --- module2b_catchment.py anomaly ---
mp = os.path.join(ABS, "module2b_catchment.py")
if os.path.exists(mp):
    emit("## Anomaly: module2b_catchment.py in abs_data\\")
    emit(f"  `{mp}`  ({fsize(mp)})")
    emit("  Python script in data folder. Likely misplaced.")
    emit("  First 15 lines:")
    try:
        with open(mp, "r", encoding="utf-8", errors="replace") as f:
            for i in range(15):
                line = f.readline()
                if not line:
                    break
                emit(f"  | {line.rstrip()}")
    except Exception as e:
        emit(f"  ERROR: {e}")
    emit("")

# --- write file ---
os.makedirs("recon", exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"\n[wrote {OUT}]", flush=True)
