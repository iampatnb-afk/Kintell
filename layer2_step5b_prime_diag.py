"""
Layer 2 Step 5b-prime — Diagnostic v4 (multi-year sampling).
=============================================================
v4 fix: prior version sampled only the latest year (2023), which made
all Census-only metrics (educational attainment, LFP, occupation,
employment) appear as 'no_data'. This version samples ALL years and
classifies per (column, year) so we see cadence + value type clearly.

Standard 25 still applies: SA2-level rows only.
Same long-format structure as v3.

Outputs:
  recon/layer2_step5b_prime_diag.md
  recon/layer2_step5b_prime_column_map.json
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime
from collections import OrderedDict, Counter, defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

# --- Config ---------------------------------------------------------------

WORKBOOK = Path("abs_data/Education and employment database.xlsx")
RECON_DIR = Path("recon")
DIAG_MD = RECON_DIR / "layer2_step5b_prime_diag.md"
COLUMN_MAP_JSON = RECON_DIR / "layer2_step5b_prime_column_map.json"

SEARCH_PATTERNS = OrderedDict([
    ("preschool",           ["preschool", "4 year old", "four year old",
                             "5 year old", "five year old", "enrolled"]),
    ("labour_force_partic", ["labour force participation", "participation rate",
                             "participation"]),
    ("female",              ["female", "females", "women"]),
    ("male",                ["male ", "males"]),
    ("educational_attain",  ["highest educational", "bachelor", "year 12",
                             "diploma", "certificate", "postgraduate",
                             "non-school qualification", "qualification",
                             "degree", "advanced diploma"]),
    ("industry",            ["industry"]),
    ("occupation",          ["occupation"]),
    ("employment",          ["employed", "employment"]),
    ("unemployment",        ["unemployment", "unemployed", "jobless"]),
])

SA2_CODE_RE   = re.compile(r"^\s*\d{9}\s*$")
SA3_CODE_RE   = re.compile(r"^\s*\d{5}\s*$")
SA4_CODE_RE   = re.compile(r"^\s*\d{3}\s*$")
LGA_CODE_RE   = re.compile(r"^\s*\d{5}\s*$")
STATE_CODE_RE = re.compile(r"^\s*[1-9]\s*$")
GCCSA_RE      = re.compile(r"^\s*\d[A-Z]+$")

SAMPLE_N = 12   # SA2 samples per (col, year)
PROGRESS_EVERY = 5000


# --- Helpers --------------------------------------------------------------

def classify_geo_code(code, table_kind):
    if code is None:
        return "OTHER"
    s = str(code).strip()
    if s == "AUS":               return "AUS"
    if STATE_CODE_RE.match(s):   return "STATE"
    if GCCSA_RE.match(s):        return "GCCSA"
    if SA4_CODE_RE.match(s):     return "SA4"
    if SA2_CODE_RE.match(s):     return "SA2"
    if SA3_CODE_RE.match(s):     return "LGA" if table_kind == "lga" else "SA3"
    return "OTHER"


def classify_samples(values):
    numeric = [v for v in values if isinstance(v, (int, float))]
    n = len(numeric)
    if n == 0:
        return {"kind": "no_data", "n": 0, "mean": None,
                "min": None, "max": None, "has_decimals": False}
    mn, mx = min(numeric), max(numeric)
    mean = sum(numeric) / n
    has_decimals = any(isinstance(v, float) and v != int(v) for v in numeric)

    if mx <= 100 and mn >= 0 and has_decimals:
        kind = "pct_likely"
    elif mx <= 100 and mn >= 0:
        kind = "pct_or_small_count_AMBIG"
    elif mx > 1000:
        kind = "count_likely"
    elif 0 <= mn and mx <= 1:
        kind = "ratio_likely"
    else:
        kind = "ambiguous"

    return {"kind": kind, "n": n, "mean": round(mean, 2),
            "min": mn, "max": mx, "has_decimals": has_decimals}


def header_combo(h6, h7):
    parts = []
    if h6 not in (None, ""): parts.append(str(h6).strip())
    if h7 not in (None, ""): parts.append(str(h7).strip())
    return " | ".join(parts)


def match_patterns(text):
    h = text.lower()
    matched = []
    for key, terms in SEARCH_PATTERNS.items():
        for t in terms:
            if t.lower() in h:
                matched.append(key)
                break
    return matched


def coerce_numeric(v):
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


# --- Sheet probe (single streaming pass, multi-year samples) -------------

def probe_sheet(ws, table_kind):
    geo_counts = Counter()
    years_seen = set()
    sa2_rows_total = 0
    headers_r6 = {}
    headers_r7 = {}
    candidate_cols = None  # list of dicts
    # samples[col][year] = list of raw values, capped at SAMPLE_N
    samples = defaultdict(lambda: defaultdict(list))

    print(f"  scanning {ws.max_row} rows (streaming, multi-year)...", flush=True)

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if r_idx == 6:
            for c_idx, v in enumerate(row, start=1):
                if v is not None:
                    headers_r6[c_idx] = v
            continue
        if r_idx == 7:
            for c_idx, v in enumerate(row, start=1):
                if v is not None:
                    headers_r7[c_idx] = v
            candidate_cols = []
            for col in range(4, ws.max_column + 1):
                h6 = headers_r6.get(col)
                h7 = headers_r7.get(col)
                text = header_combo(h6, h7)
                if not text:
                    continue
                matched = match_patterns(text)
                if matched:
                    candidate_cols.append({
                        "col": col, "h6": h6, "h7": h7, "matched": matched,
                    })
            print(f"    row 7 read. {len(candidate_cols)} candidate cols out of "
                  f"{ws.max_column - 3}.", flush=True)
            continue
        if r_idx < 8:
            continue

        code = row[0] if len(row) > 0 else None
        year = row[2] if len(row) > 2 else None
        kind = classify_geo_code(code, table_kind)
        geo_counts[kind] += 1
        if year is not None:
            years_seen.add(year)

        if kind == "SA2" and candidate_cols:
            sa2_rows_total += 1
            for cc in candidate_cols:
                col = cc["col"]
                bucket = samples[col][year]
                if len(bucket) < SAMPLE_N and col - 1 < len(row):
                    bucket.append(row[col - 1])

        if r_idx % PROGRESS_EVERY == 0:
            print(f"    row {r_idx}/{ws.max_row} | SA2 so far: {sa2_rows_total}",
                  flush=True)

    print(f"    DONE. SA2 rows total: {sa2_rows_total}, geo: {dict(geo_counts)}",
          flush=True)

    # Build candidate output: per col, per year, classified
    int_years = sorted([y for y in years_seen if isinstance(y, int)])
    candidates_out = []
    for cc in (candidate_cols or []):
        col = cc["col"]
        per_year = []
        for y in int_years:
            raw = samples[col].get(y, [])
            clean = [coerce_numeric(v) for v in raw]
            clean = [v for v in clean if v is not None]
            cls = classify_samples(clean)
            per_year.append({
                "year": y,
                "kind": cls["kind"],
                "n": cls["n"],
                "mean": cls["mean"],
                "min": cls["min"],
                "max": cls["max"],
                "sample_raw": [v for v in raw if v is not None][:5],
            })
        # Determine cadence
        years_with_data = [py["year"] for py in per_year if py["n"] > 0]
        candidates_out.append({
            "col_index": col,
            "col_letter": get_column_letter(col),
            "header_row6": cc["h6"],
            "header_row7": cc["h7"],
            "matched_patterns": cc["matched"],
            "years_with_data": years_with_data,
            "per_year": per_year,
        })

    return {
        "sheet": ws.title,
        "table_kind": table_kind,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "geo_counts": dict(geo_counts),
        "years_seen": int_years,
        "sa2_rows_total": sa2_rows_total,
        "candidate_columns": candidates_out,
    }


# --- Output helpers -------------------------------------------------------

def cadence_label(years_with_data, all_years):
    """Return a short label: 'census', 'annual', 'sparse', or 'none'."""
    if not years_with_data:
        return "none"
    census_years = {2011, 2016, 2021}
    annual_years = set(all_years) - census_years
    has = set(years_with_data)
    only_census = has and has.issubset(census_years)
    has_most_annual = annual_years and len(has & annual_years) >= max(1, len(annual_years) // 2)
    if only_census:
        return "census"
    if has_most_annual:
        return "annual"
    return "sparse"


def overall_kind(per_year):
    """Most common 'kind' across years where n > 0."""
    kinds = [py["kind"] for py in per_year if py["n"] > 0]
    if not kinds:
        return "no_data"
    counts = Counter(kinds)
    return counts.most_common(1)[0][0]


# --- Main -----------------------------------------------------------------

def main():
    if not WORKBOOK.exists():
        print(f"ERROR: workbook not found at {WORKBOOK}", file=sys.stderr)
        sys.exit(1)

    RECON_DIR.mkdir(exist_ok=True)

    print(f"Opening workbook: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    sheet_kinds = {"Table 1": "sa2", "Table 2": "lga"}

    results = []
    for name in wb.sheetnames:
        if name.lower() == "contents":
            print(f"  [skip] {name}")
            continue
        kind = sheet_kinds.get(name)
        if kind is None:
            print(f"  [skip] {name} (unknown)")
            continue
        print(f"  [probe {kind}] {name}")
        results.append(probe_sheet(wb[name], kind))

    wb.close()

    with open(COLUMN_MAP_JSON, "w", encoding="utf-8") as f:
        json.dump({
            "workbook": str(WORKBOOK),
            "generated": datetime.now().isoformat(timespec="seconds"),
            "format": "long (Code | Label | Year | metric_1 | metric_2 ...)",
            "sheets": results,
        }, f, indent=2, default=str)
    print(f"Wrote {COLUMN_MAP_JSON}")

    L = []
    L.append("# Layer 2 Step 5b-prime — Diagnostic v4 (multi-year)")
    L.append("")
    L.append(f"**Workbook:** `{WORKBOOK}`  ")
    L.append(f"**Generated:** {datetime.now().isoformat(timespec='seconds')}")
    L.append("")
    L.append("Long-format workbook (Code | Label | Year | metric cols).")
    L.append("Samples drawn from SA2-level rows for **every available year** so")
    L.append("Census-only metrics (publish 2011/2016/2021 only) are visible.")
    L.append("")

    for r in results:
        if r["table_kind"] != "sa2":
            L.append(f"## Sheet: `{r['sheet']}` (kind: {r['table_kind']}) — skipped from detail (no SA2 rows)")
            L.append("")
            continue

        L.append(f"## Sheet: `{r['sheet']}` (kind: {r['table_kind']})")
        L.append("")
        L.append(f"- max_row: {r['max_row']}, max_col: {r['max_col']}")
        L.append(f"- geography breakdown: {r['geo_counts']}")
        L.append(f"- years seen: {r['years_seen']}")
        L.append(f"- SA2 rows (total): {r['sa2_rows_total']}")
        L.append("")

        # Compact summary table: one row per col
        L.append("### Summary — one row per candidate column")
        L.append("")
        L.append("| Col | Letter | Group (row 6) | Metric (row 7) | Matched | Cadence | Overall kind | Years w/ data |")
        L.append("|----:|:-------|:--------------|:---------------|:--------|:--------|:-------------|:--------------|")
        for c in r["candidate_columns"]:
            r6 = str(c["header_row6"] or "").replace("|", "\\|").replace("\n", " ")[:46]
            r7 = str(c["header_row7"] or "").replace("|", "\\|").replace("\n", " ")[:46]
            cad = cadence_label(c["years_with_data"], r["years_seen"])
            ok = overall_kind(c["per_year"])
            yws = ",".join(str(y) for y in c["years_with_data"]) or "-"
            L.append(f"| {c['col_index']} | {c['col_letter']} | {r6} | {r7} | "
                     f"{','.join(c['matched_patterns'])} | {cad} | {ok} | {yws} |")
        L.append("")

        # Detail table: per (col, year) — only non-empty rows
        L.append("### Detail — per-year classification (n > 0 only)")
        L.append("")
        L.append("| Col | Letter | Metric | Year | n | Kind | Mean | Min | Max | Sample |")
        L.append("|----:|:-------|:-------|-----:|--:|:-----|-----:|----:|----:|:-------|")
        for c in r["candidate_columns"]:
            r7 = str(c["header_row7"] or "").replace("|", "\\|").replace("\n", " ")[:42]
            for py in c["per_year"]:
                if py["n"] == 0:
                    continue
                sample_str = ", ".join(str(v)[:10] for v in py["sample_raw"][:4])
                L.append(f"| {c['col_index']} | {c['col_letter']} | {r7} | "
                         f"{py['year']} | {py['n']} | {py['kind']} | "
                         f"{py['mean']} | {py['min']} | {py['max']} | {sample_str} |")
        L.append("")

    L.append("## Manual review checklist")
    L.append("")
    L.append("1. Pick 6-10 target metrics for ingest. Likely shortlist:")
    L.append("   - Annual preschool: cols 4 (4yo enrolled), 9 (total enrolled),")
    L.append("     11 (15h+/week attendance)")
    L.append("   - Census educational attainment: col 12 (Year 12+%),")
    L.append("     col 22 (Bachelor%)")
    L.append("   - Census labour: col 54 (Unemployment rate %),")
    L.append("     col 55 (Participation rate %)")
    L.append("   - Female-specific: TBD — check if col 60/61 (jobs by sex)")
    L.append("     gives a usable signal at SA2, or whether female LFP is")
    L.append("     not actually published at SA2 in this workbook.")
    L.append("2. Schema: new table `abs_sa2_education_employment_annual` with")
    L.append("   columns (sa2_code, year, metric_name, value).")
    L.append("3. Sanity checks per metric type — Standard 25.")

    with open(DIAG_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    print(f"Wrote {DIAG_MD}")

    total = sum(len(r["candidate_columns"]) for r in results)
    print(f"\nDone. {total} candidate columns across {len(results)} sheets.")


if __name__ == "__main__":
    main()
