"""
Layer 2 Step 6 DIAGNOSTIC — ABS ERP at SA2 layout discovery
─────────────────────────────────────────────────────────────
READ-ONLY. No DB mutations. No file writes outside recon/.

Purpose:
  Probes Table 1 of Population and People Database.xlsx to discover:
    - Which column carries the year value
    - Where total/all-ages persons sits
    - Confirms under-5 columns at indices 12/48/84
    - Probes data/kintell.db.audit_log schema for the apply step
    - Checks whether abs_sa2_erp_annual already exists

Outputs:
  - Console summary
  - recon/layer2_step6_diag.md       (human-readable findings)
  - recon/layer2_step6_column_map.json  (consumed by apply script)
"""

from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
WORKBOOK = PROJECT_ROOT / "abs_data" / "Population and People Database.xlsx"
DB = PROJECT_ROOT / "data" / "kintell.db"
OUT_DIR = PROJECT_ROOT / "recon"
OUT_MD = OUT_DIR / "layer2_step6_diag.md"
OUT_JSON = OUT_DIR / "layer2_step6_column_map.json"

EXPECTED_YEARS = [2011, 2016, 2018, 2019, 2020, 2021, 2022, 2023, 2024]
HEADER_ROW = 6


def banner(title: str) -> str:
    return f"\n{'─' * 70}\n{title}\n{'─' * 70}"


def main() -> int:
    if not WORKBOOK.exists():
        print(f"FAIL: {WORKBOOK} not found")
        return 1
    if not DB.exists():
        print(f"FAIL: {DB} not found")
        return 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    findings: list[str] = ["# Layer 2 Step 6 — Diagnostic", ""]

    # ── 1. Read first 12 rows of Table 1 ─────────────────────────────
    print(banner("1. WORKBOOK STRUCTURE (first 12 rows)"))
    findings.append("## 1. Workbook structure (first 12 rows)\n")

    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Table 1"]

    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        if i >= 12:
            break

    n_cols = max(len(r) for r in rows)
    print(f"Total columns: {n_cols}")
    findings.append(f"- Total columns: **{n_cols}**\n")

    # ── 2. Print rows 1-12, columns 0-15 (metadata block) ────────────
    print(banner("2. ROWS 1-12, COLUMNS 0-15 (likely metadata + year)"))
    findings.append("## 2. Rows 1-12, columns 0-15 (metadata block)\n")
    findings.append("```")
    for r_idx, row in enumerate(rows, start=1):
        cells = []
        for c in row[:16]:
            if c is None:
                cells.append("·")
            else:
                cells.append(str(c)[:20])
        line = f"Row {r_idx:2d}: " + " | ".join(cells)
        print(line)
        findings.append(line)
    findings.append("```\n")

    # ── 3. Year-column detection ─────────────────────────────────────
    print(banner("3. YEAR COLUMN DETECTION"))
    findings.append("## 3. Year column detection\n")

    year_col_candidates: dict[int, set[int]] = {}
    # Scan first 12 rows × first 20 cols for 4-digit year cells
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row[:20]):
            if val is None:
                continue
            s = str(val).strip()
            if s.isdigit() and len(s) == 4:
                yr = int(s)
                if 2010 <= yr <= 2030:
                    year_col_candidates.setdefault(c_idx, set()).add(yr)

    if year_col_candidates:
        for c_idx in sorted(year_col_candidates):
            yrs = sorted(year_col_candidates[c_idx])
            print(f"  col[{c_idx}]: years seen in first 12 rows = {yrs}")
            findings.append(f"- col[{c_idx}]: years seen = `{yrs}`")
        # Pick the one with the most matches
        best = max(year_col_candidates, key=lambda k: len(year_col_candidates[k]))
        print(f"  → likely year column: {best}")
        findings.append(f"- **Best guess year column: `{best}`**\n")
        likely_year_col = best
    else:
        print("  No year cells in first 12 rows × cols 0-19.")
        print("  Year may be encoded later in data, or as a label string.")
        findings.append("- No 4-digit year found in metadata block.")
        findings.append("- Need to scan further into data rows.\n")
        likely_year_col = None

    # ── 4. Wider scan: look in cols 0-15 of first 30 data rows ───────
    if likely_year_col is None:
        print(banner("4. EXTENDED YEAR SCAN (rows 6-50, cols 0-15)"))
        findings.append("## 4. Extended year scan\n")
        wb2 = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
        ws2 = wb2["Table 1"]
        ext_candidates: dict[int, set[int]] = {}
        for r_idx, row in enumerate(ws2.iter_rows(values_only=True), start=1):
            if r_idx <= HEADER_ROW:
                continue
            if r_idx > 50:
                break
            for c_idx in range(min(16, len(row))):
                val = row[c_idx]
                if val is None:
                    continue
                s = str(val).strip()
                if s.isdigit() and len(s) == 4 and 2010 <= int(s) <= 2030:
                    ext_candidates.setdefault(c_idx, set()).add(int(s))
        wb2.close()
        if ext_candidates:
            for c_idx in sorted(ext_candidates):
                print(f"  col[{c_idx}]: {sorted(ext_candidates[c_idx])}")
                findings.append(f"- col[{c_idx}]: `{sorted(ext_candidates[c_idx])}`")
            best = max(ext_candidates, key=lambda k: len(ext_candidates[k]))
            likely_year_col = best
            print(f"  → likely year column: {best}")
            findings.append(f"- **Best guess year column: `{best}`**\n")

    # ── 5. Row 6 spanning headers ────────────────────────────────────
    print(banner("5. ROW 6 SPANNING HEADERS (non-null only)"))
    findings.append("## 5. Row 6 spanning headers (non-null only)\n")
    findings.append("```")
    row6 = rows[5]
    for c_idx in range(n_cols):
        if c_idx < len(row6) and row6[c_idx] is not None:
            line = f"col[{c_idx:3d}]: {str(row6[c_idx])[:80]}"
            print(line)
            findings.append(line)
    findings.append("```\n")

    # ── 6. Total / All-ages search ───────────────────────────────────
    print(banner("6. TOTAL / ALL-AGES COLUMN CANDIDATES"))
    findings.append("## 6. Total / all-ages column candidates\n")

    total_candidates: list[tuple[int, int, str]] = []
    for r_idx in range(min(10, len(rows))):
        for c_idx, v in enumerate(rows[r_idx]):
            if v is None:
                continue
            s = str(v).lower()
            if any(t in s for t in ["all ages", "total persons",
                                    "all age", "total"]):
                total_candidates.append(
                    (r_idx + 1, c_idx, str(v)[:80])
                )

    if total_candidates:
        for r, c, v in total_candidates[:30]:
            print(f"  row {r:2d}, col[{c:3d}]: {v}")
            findings.append(f"- row {r}, col[{c}]: `{v}`")
        # Heuristic: pick the col that appears in the Persons block
        # (col >= 84) and has 'all ages' or 'total'
        persons_block_totals = [
            c for r, c, v in total_candidates
            if c >= 84 and any(t in v.lower()
                               for t in ["all ages", "total"])
        ]
        likely_total_col = persons_block_totals[0] if persons_block_totals else None
    else:
        print("  No 'all ages' / 'total persons' text in first 10 rows.")
        findings.append("- **No total-persons column header found.**")
        findings.append("- Will need to derive total by summing age bands "
                        "at apply time.")
        likely_total_col = None

    if likely_total_col is not None:
        print(f"  → likely total_persons column: {likely_total_col}")
        findings.append(f"- **Best guess total_persons column: "
                        f"`{likely_total_col}`**\n")
    else:
        findings.append("")

    # ── 7. Sample 5 SA2 rows showing year col + under-5 cols ─────────
    print(banner("7. FIRST 5 SA2-SHAPED DATA ROWS"))
    findings.append("## 7. First 5 SA2-shaped data rows\n")

    wb3 = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws3 = wb3["Table 1"]
    sa2_samples = []
    for r_idx, row in enumerate(ws3.iter_rows(values_only=True), start=1):
        if r_idx <= HEADER_ROW:
            continue
        first = row[0] if len(row) > 0 else None
        if first is None:
            continue
        s = str(first).strip()
        if s.isdigit() and len(s) == 9:
            sa2_samples.append((r_idx, list(row)))
            if len(sa2_samples) >= 5:
                break
    wb3.close()

    findings.append("```")
    for r_idx, row in sa2_samples:
        sa2 = row[0]
        label = row[1] if len(row) > 1 else None
        year_val = (row[likely_year_col]
                    if likely_year_col is not None
                    and likely_year_col < len(row)
                    else "?")
        u5m = row[12] if 12 < len(row) else "?"
        u5f = row[48] if 48 < len(row) else "?"
        u5p = row[84] if 84 < len(row) else "?"
        tot = (row[likely_total_col]
               if likely_total_col is not None
               and likely_total_col < len(row)
               else "?")
        line = (f"row {r_idx:5d}: sa2={sa2} label={str(label)[:25]:25s} "
                f"year={year_val} u5m={u5m} u5f={u5f} u5p={u5p} tot={tot}")
        print(line)
        findings.append(line)
    findings.append("```\n")

    wb.close()

    # ── 8. Probe audit_log schema ────────────────────────────────────
    print(banner("8. AUDIT_LOG SCHEMA"))
    findings.append("## 8. audit_log schema (for apply step)\n")

    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(audit_log)")
    audit_cols = cur.fetchall()
    print("audit_log columns:")
    findings.append("| cid | name | type | notnull | dflt | pk |")
    findings.append("|---:|---|---|:-:|---|:-:|")
    audit_col_names: list[str] = []
    for col in audit_cols:
        line = (f"  cid={col[0]} name={col[1]:20s} type={col[2]:15s} "
                f"notnull={col[3]} dflt={col[4]} pk={col[5]}")
        print(line)
        findings.append(f"| {col[0]} | `{col[1]}` | `{col[2]}` | "
                        f"{col[3]} | `{col[4]}` | {col[5]} |")
        audit_col_names.append(col[1])
    findings.append("")

    # Check for sa2_erp_annual
    cur.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='abs_sa2_erp_annual'
    """)
    exists = cur.fetchone()
    if exists:
        cur.execute("SELECT COUNT(*) FROM abs_sa2_erp_annual")
        n = cur.fetchone()[0]
        print(f"\nabs_sa2_erp_annual already exists: {n:,} rows")
        findings.append(f"- ⚠️ `abs_sa2_erp_annual` already exists "
                        f"with {n:,} rows.")
        findings.append("  Apply will refuse if non-empty.")
    else:
        print("\nabs_sa2_erp_annual does not exist (apply will create).")
        findings.append("- `abs_sa2_erp_annual` does not exist (apply "
                        "will create).")
    findings.append("")

    # Last audit_id
    cur.execute("SELECT COALESCE(MAX(audit_id), 0) FROM audit_log")
    last_audit = cur.fetchone()[0]
    print(f"Last audit_id: {last_audit} (next will be {last_audit + 1})")
    findings.append(f"- Last `audit_id`: **{last_audit}** (next: "
                    f"**{last_audit + 1}**)\n")

    conn.close()

    # ── 9. Write column_map.json ─────────────────────────────────────
    print(banner("9. COLUMN MAP"))
    column_map = {
        "table_sheet": "Table 1",
        "header_row": HEADER_ROW,
        "data_start_row": HEADER_ROW + 1,
        "columns": {
            "sa2_code": 0,
            "region_label": 1,
            "year_column": likely_year_col,
            "under_5_males": 12,
            "under_5_females": 48,
            "under_5_persons": 84,
            "total_persons": likely_total_col,
        },
        "null_markers": ["-", "..", "np", "n.a.", "na", "—", ""],
        "expected_years": EXPECTED_YEARS,
        "audit_log_columns": audit_col_names,
        "notes": [
            "Verify year_column and total_persons before running apply.",
            "If year_column is null, check the diag markdown to find it.",
            "If total_persons is null, the apply will skip that age_group.",
        ],
    }
    OUT_JSON.write_text(json.dumps(column_map, indent=2), encoding="utf-8")
    print(f"Column map written to: {OUT_JSON}")
    print("\nResolved column map:")
    print(json.dumps(column_map["columns"], indent=2))

    findings.append("## 9. Column map (written to JSON)\n")
    findings.append("```json")
    findings.append(json.dumps(column_map["columns"], indent=2))
    findings.append("```\n")
    findings.append("⚠️ **Verify the column map before running apply.** "
                    "If `year_column` or `total_persons` shows `null`, "
                    "review sections 3–6 above and edit the JSON manually.")

    OUT_MD.write_text("\n".join(findings), encoding="utf-8")
    print(banner("DONE"))
    print(f"  Markdown: {OUT_MD}")
    print(f"  JSON:     {OUT_JSON}")
    print()
    print("NEXT: open the markdown, confirm year_column and total_persons,")
    print("      edit the JSON if needed, then run the apply script.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
