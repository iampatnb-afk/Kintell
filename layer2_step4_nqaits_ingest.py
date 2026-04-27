#!/usr/bin/env python3
"""
layer2_step4_nqaits_ingest.py v1 — Phase 2.5 Layer 2 Step 4

NQAITS Quarterly Data Splits ingest (Q3 2013 - Q4 2025).
Creates nqs_history table; populates ~807K rows across 50 quarterly snapshots
covering ~23,439 services (incl ~5,514 closed centres not in services table).

Pattern (Decision 10):
  1. Backup
  2. Pre-state invariants
  3. Open NQAITS, resolve canonical columns per sheet
  4. BEGIN TRANSACTION
  5. CREATE TABLE nqs_history + indexes
  6. INSERT chunked across all 50 sheets (executemany ~1000 rows per chunk)
  7. INSERT audit_log row
  8. Post-state invariants -> assert
  9. COMMIT (or ROLLBACK + raise)
 10. Post-commit re-verify
"""
import os
import re
import sys
import csv
import json
import shutil
import sqlite3
import datetime
from collections import defaultdict, Counter

NQAITS = os.path.join("abs_data", "NQAITS Quarterly Data Splits (Q3 2013 - Q4 2025).xlsx")
DB = os.path.join("data", "kintell.db")
TS = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
BACKUP = f"{DB}.backup_pre_nqaits_ingest_{TS}"
ACTOR = "layer2_step4_nqaits_ingest_v1"
ACTION = "nqaits_ingest_v1"

EXPECTED_SERVICES_TOTAL = 18223
EXPECTED_ENTITIES_TOTAL = 7143
EXPECTED_GROUPS_TOTAL = 6507  # post-step-2

# Canonical column variant map (validated 50/50 in preflight v2)
COL_VARIANTS = {
    "service_approval_number":  ["Service ID", "Service Approval Number"],
    "service_name":             ["Service Name"],
    "provider_id":              ["Provider ID"],
    "provider_name":            ["Provider Name"],
    "provider_management_type": ["Provider Management Type"],
    "managing_jurisdiction":    ["Managing Jurisdiction", "Managing Jurisdiction "],
    "service_type":             ["Service Type"],
    "service_sub_type":         ["Service sub-type (ordered counting method)", "Service Sub Type"],
    "approval_date":            ["Approval Date", "ApprovalDate"],
    "seifa":                    ["SEIFA"],
    "aria":                     ["ARIA", "ARIA+"],
    "max_places":               ["Maximum total places"],
    "nqs_version":              ["NQS Version"],
    "final_report_sent_date":   ["Final Report Sent Date"],
    "overall_rating":           ["Overall Rating", "OverallRating"],
    "quality_area_1":           ["Quality Area 1", "Q1"],
    "quality_area_2":           ["Quality Area 2", "Q2"],
    "quality_area_3":           ["Quality Area 3", "Q3"],
    "quality_area_4":           ["Quality Area 4", "Q4"],
    "quality_area_5":           ["Quality Area 5", "Q5"],
    "quality_area_6":           ["Quality Area 6", "Q6"],
    "quality_area_7":           ["Quality Area 7", "Q7"],
    "long_day_care":            ["Long Day Care"],
    "preschool_standalone":     [
        "PreschoolKindergarten Stand Alone",
        "Preschool/Kindergarten Stand Alone",
        "Preschool/\nKindergarten Stand Alone",
    ],
    "preschool_in_school":      [
        "PreschoolKindergarten Part of a School",
        "Preschool/\nKindergarten Part of a School",
    ],
    "oshc_before_school":       ["OSHC BeforeSchool", "OSHC Before School"],
    "oshc_after_school":        ["OSHC After School"],
    "oshc_vacation_care":       ["OSHC Vacation Care"],
    "nature_care_other":        ["Nature Care Other"],
    "postcode":                 ["Postcode"],
    "latitude":                 ["Latitude"],
    "longitude":                ["Longitude"],
}

# Insert column order — matches CREATE TABLE
INSERT_COLS = [
    "service_approval_number", "quarter", "quarter_end_date", "source_sheet",
    "service_name", "service_type", "service_sub_type",
    "provider_id", "provider_name", "provider_management_type",
    "managing_jurisdiction", "postcode", "latitude", "longitude",
    "aria", "seifa",
    "nqs_version", "overall_rating",
    "quality_area_1", "quality_area_2", "quality_area_3", "quality_area_4",
    "quality_area_5", "quality_area_6", "quality_area_7",
    "approval_date", "final_report_sent_date",
    "max_places",
    "long_day_care", "preschool_standalone", "preschool_in_school",
    "oshc_before_school", "oshc_after_school", "oshc_vacation_care",
    "nature_care_other",
]

CREATE_TABLE_SQL = """
CREATE TABLE nqs_history (
    nqs_history_id            INTEGER PRIMARY KEY AUTOINCREMENT,
    service_approval_number   TEXT NOT NULL,
    quarter                   TEXT NOT NULL,
    quarter_end_date          TEXT NOT NULL,
    source_sheet              TEXT NOT NULL,
    service_name              TEXT,
    service_type              TEXT,
    service_sub_type          TEXT,
    provider_id               TEXT,
    provider_name             TEXT,
    provider_management_type  TEXT,
    managing_jurisdiction     TEXT,
    postcode                  TEXT,
    latitude                  REAL,
    longitude                 REAL,
    aria                      TEXT,
    seifa                     TEXT,
    nqs_version               TEXT,
    overall_rating            TEXT,
    quality_area_1            TEXT,
    quality_area_2            TEXT,
    quality_area_3            TEXT,
    quality_area_4            TEXT,
    quality_area_5            TEXT,
    quality_area_6            TEXT,
    quality_area_7            TEXT,
    approval_date             TEXT,
    final_report_sent_date    TEXT,
    max_places                INTEGER,
    long_day_care             TEXT,
    preschool_standalone      TEXT,
    preschool_in_school       TEXT,
    oshc_before_school        TEXT,
    oshc_after_school         TEXT,
    oshc_vacation_care        TEXT,
    nature_care_other         TEXT,
    ingested_at               TEXT NOT NULL DEFAULT (datetime('now')),
    UNIQUE (service_approval_number, quarter)
);
"""

CREATE_INDEX_SQL = [
    "CREATE INDEX ix_nqs_history_san ON nqs_history (service_approval_number);",
    "CREATE INDEX ix_nqs_history_quarter ON nqs_history (quarter_end_date);",
    "CREATE INDEX ix_nqs_history_provider ON nqs_history (provider_id);",
]


def banner(msg):
    print("\n" + "=" * 60)
    print(msg)
    print("=" * 60, flush=True)


def parse_sheet_to_quarter(sheet_name):
    """ 'Q42025data' -> ('Q42025', '2025-12-31') """
    m = re.match(r"^Q(\d)(\d{4})data$", sheet_name)
    if not m:
        return None, None
    qnum, year = int(m.group(1)), int(m.group(2))
    end_dates = {1: "03-31", 2: "06-30", 3: "09-30", 4: "12-31"}
    return f"Q{qnum}{year}", f"{year}-{end_dates[qnum]}"


def resolve_columns(header_row):
    h = [str(c).strip() if c is not None else None for c in header_row]
    out = {}
    for canon, variants in COL_VARIANTS.items():
        for v in variants:
            if v in h:
                out[canon] = h.index(v)
                break
    return out


def cell_to_text(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S") if (v.hour or v.minute or v.second) else v.strftime("%d/%m/%Y")
    if isinstance(v, datetime.date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    return s if s else None


def cell_to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def cell_to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def main():
    if not os.path.exists(NQAITS):
        sys.exit(f"FATAL: {NQAITS} not found")
    if not os.path.exists(DB):
        sys.exit(f"FATAL: {DB} not found")
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("FATAL: pip install openpyxl")

    # --- 1. Backup ---
    banner("1. Backup")
    shutil.copy2(DB, BACKUP)
    print(f"  copied -> {BACKUP}")
    print(f"  size: {os.path.getsize(BACKUP):,} bytes")
    if os.path.getsize(BACKUP) != os.path.getsize(DB):
        sys.exit("FATAL: backup size mismatch")

    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    def cnt(table, where=""):
        sql = f"SELECT COUNT(*) FROM {table}"
        if where:
            sql += f" WHERE {where}"
        cur.execute(sql)
        return cur.fetchone()[0]

    # --- 2. Pre-state invariants ---
    banner("2. Pre-state invariants")
    pre = {
        "services_total": cnt("services"),
        "entities_total": cnt("entities"),
        "groups_total":   cnt("groups"),
        "audit_log_count": cnt("audit_log"),
    }
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='nqs_history';")
    pre["nqs_history_exists"] = cur.fetchone() is not None
    for k, v in pre.items():
        print(f"  {k:<28} {v}")

    if pre["services_total"] != EXPECTED_SERVICES_TOTAL:
        sys.exit(f"FATAL: services_total {pre['services_total']} != {EXPECTED_SERVICES_TOTAL}")
    if pre["entities_total"] != EXPECTED_ENTITIES_TOTAL:
        sys.exit(f"FATAL: entities_total {pre['entities_total']} != {EXPECTED_ENTITIES_TOTAL}")
    if pre["groups_total"] != EXPECTED_GROUPS_TOTAL:
        sys.exit(f"FATAL: groups_total {pre['groups_total']} != {EXPECTED_GROUPS_TOTAL}")
    if pre["nqs_history_exists"]:
        sys.exit("FATAL: nqs_history table already exists; this script creates it fresh")

    # --- 3. Open workbook + resolve columns ---
    banner("3. Open workbook + resolve columns")
    print("[opening NQAITS — 138MB, ~30s ...]", flush=True)
    wb = load_workbook(NQAITS, read_only=True, data_only=True)
    data_sheets = [s for s in wb.sheetnames if re.match(r"Q\d20\d\ddata", s)]
    print(f"  data sheets: {len(data_sheets)}")

    sheet_cols = {}
    for s in data_sheets:
        ws = wb[s]
        h = next(ws.iter_rows(values_only=True))
        sheet_cols[s] = resolve_columns(h)
    unresolved = {s: [c for c in COL_VARIANTS if c not in cols] for s, cols in sheet_cols.items()}
    unresolved = {s: m for s, m in unresolved.items() if m}
    if unresolved:
        for s, m in list(unresolved.items())[:5]:
            print(f"  {s}: missing canonical {m}")
        if any(["service_approval_number" in m for m in unresolved.values()]):
            sys.exit("FATAL: service_approval_number unresolved in some sheets")
    print(f"  fully resolved: {len(data_sheets) - len(unresolved)} / {len(data_sheets)}")

    # --- 4-9. Transaction ---
    banner("4. Apply transaction (CREATE TABLE + INSERT all rows)")
    inserted_total = 0
    sheet_counts = {}
    sid_set = set()
    chain_changes = 0  # for audit_log
    sid_provider_per_quarter = defaultdict(dict)

    try:
        cur.execute("BEGIN TRANSACTION;")
        cur.executescript(CREATE_TABLE_SQL)
        for sql in CREATE_INDEX_SQL:
            cur.execute(sql)
        print("  CREATE TABLE + 3 indexes OK")

        BATCH = 1000
        placeholders = ",".join(["?"] * len(INSERT_COLS))
        insert_sql = f"INSERT INTO nqs_history ({','.join(INSERT_COLS)}) VALUES ({placeholders});"

        for sidx, sheet in enumerate(data_sheets):
            quarter, q_end = parse_sheet_to_quarter(sheet)
            cols = sheet_cols[sheet]
            sid_idx = cols["service_approval_number"]
            ws = wb[sheet]
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter)  # header

            batch = []
            sheet_inserted = 0
            for row in rows_iter:
                if not row or row[sid_idx] is None:
                    continue
                san = str(row[sid_idx]).strip()
                if not san:
                    continue
                sid_set.add(san)

                # Build value tuple in INSERT_COLS order
                vals = [
                    san,
                    quarter,
                    q_end,
                    sheet,
                    # service identity
                    cell_to_text(row[cols["service_name"]]) if "service_name" in cols and cols["service_name"] < len(row) else None,
                    cell_to_text(row[cols["service_type"]]) if "service_type" in cols and cols["service_type"] < len(row) else None,
                    cell_to_text(row[cols["service_sub_type"]]) if "service_sub_type" in cols and cols["service_sub_type"] < len(row) else None,
                    # provider
                    cell_to_text(row[cols["provider_id"]]) if "provider_id" in cols and cols["provider_id"] < len(row) else None,
                    cell_to_text(row[cols["provider_name"]]) if "provider_name" in cols and cols["provider_name"] < len(row) else None,
                    cell_to_text(row[cols["provider_management_type"]]) if "provider_management_type" in cols and cols["provider_management_type"] < len(row) else None,
                    # location
                    cell_to_text(row[cols["managing_jurisdiction"]]) if "managing_jurisdiction" in cols and cols["managing_jurisdiction"] < len(row) else None,
                    cell_to_text(row[cols["postcode"]]) if "postcode" in cols and cols["postcode"] < len(row) else None,
                    cell_to_float(row[cols["latitude"]]) if "latitude" in cols and cols["latitude"] < len(row) else None,
                    cell_to_float(row[cols["longitude"]]) if "longitude" in cols and cols["longitude"] < len(row) else None,
                    # demographics
                    cell_to_text(row[cols["aria"]]) if "aria" in cols and cols["aria"] < len(row) else None,
                    cell_to_text(row[cols["seifa"]]) if "seifa" in cols and cols["seifa"] < len(row) else None,
                    # rating
                    cell_to_text(row[cols["nqs_version"]]) if "nqs_version" in cols and cols["nqs_version"] < len(row) else None,
                    cell_to_text(row[cols["overall_rating"]]) if "overall_rating" in cols and cols["overall_rating"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_1"]]) if "quality_area_1" in cols and cols["quality_area_1"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_2"]]) if "quality_area_2" in cols and cols["quality_area_2"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_3"]]) if "quality_area_3" in cols and cols["quality_area_3"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_4"]]) if "quality_area_4" in cols and cols["quality_area_4"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_5"]]) if "quality_area_5" in cols and cols["quality_area_5"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_6"]]) if "quality_area_6" in cols and cols["quality_area_6"] < len(row) else None,
                    cell_to_text(row[cols["quality_area_7"]]) if "quality_area_7" in cols and cols["quality_area_7"] < len(row) else None,
                    # dates
                    cell_to_text(row[cols["approval_date"]]) if "approval_date" in cols and cols["approval_date"] < len(row) else None,
                    cell_to_text(row[cols["final_report_sent_date"]]) if "final_report_sent_date" in cols and cols["final_report_sent_date"] < len(row) else None,
                    # places + composition
                    cell_to_int(row[cols["max_places"]]) if "max_places" in cols and cols["max_places"] < len(row) else None,
                    cell_to_text(row[cols["long_day_care"]]) if "long_day_care" in cols and cols["long_day_care"] < len(row) else None,
                    cell_to_text(row[cols["preschool_standalone"]]) if "preschool_standalone" in cols and cols["preschool_standalone"] < len(row) else None,
                    cell_to_text(row[cols["preschool_in_school"]]) if "preschool_in_school" in cols and cols["preschool_in_school"] < len(row) else None,
                    cell_to_text(row[cols["oshc_before_school"]]) if "oshc_before_school" in cols and cols["oshc_before_school"] < len(row) else None,
                    cell_to_text(row[cols["oshc_after_school"]]) if "oshc_after_school" in cols and cols["oshc_after_school"] < len(row) else None,
                    cell_to_text(row[cols["oshc_vacation_care"]]) if "oshc_vacation_care" in cols and cols["oshc_vacation_care"] < len(row) else None,
                    cell_to_text(row[cols["nature_care_other"]]) if "nature_care_other" in cols and cols["nature_care_other"] < len(row) else None,
                ]

                # Track PA-chain changes for audit summary
                pid = vals[INSERT_COLS.index("provider_id")]
                if pid:
                    sid_provider_per_quarter[san][quarter] = pid

                batch.append(vals)
                if len(batch) >= BATCH:
                    cur.executemany(insert_sql, batch)
                    sheet_inserted += len(batch)
                    batch = []
            if batch:
                cur.executemany(insert_sql, batch)
                sheet_inserted += len(batch)

            inserted_total += sheet_inserted
            sheet_counts[sheet] = sheet_inserted
            if (sidx + 1) % 5 == 0 or sidx == len(data_sheets) - 1:
                print(f"  [{sidx+1:>2}/{len(data_sheets)}] {sheet}: +{sheet_inserted:,}  total {inserted_total:,}", flush=True)

        wb.close()

        # PA chain count
        for sid, qprov in sid_provider_per_quarter.items():
            if len(set(qprov.values())) > 1:
                chain_changes += 1

        # --- 7. audit_log ---
        banner("5. audit_log + invariants")
        before_json = json.dumps({
            "nqs_history_exists": False,
            "expected_volume_estimate": 807526,
        })
        after_json = json.dumps({
            "nqs_history_rowcount": inserted_total,
            "unique_service_ids": len(sid_set),
            "sheets_ingested": len(sheet_counts),
            "pa_chain_changers": chain_changes,
        })
        reason = (
            f"Phase 2.5 Layer 2 Step 4 — NQAITS quarterly historical ingest. "
            f"Created nqs_history table + 3 indexes, ingested {inserted_total:,} rows "
            f"across {len(sheet_counts)} quarterly sheets ({data_sheets[-1]} -> {data_sheets[0]}). "
            f"{len(sid_set):,} unique service_approval_numbers; "
            f"{chain_changes:,} services with PA-chain changes. "
            f"Backup: {BACKUP}."
        )
        cur.execute(
            """INSERT INTO audit_log
                 (actor, action, subject_type, subject_id, before_json, after_json, reason)
               VALUES (?, ?, ?, ?, ?, ?, ?);""",
            (ACTOR, ACTION, "nqs_history", 0, before_json, after_json, reason),
        )
        print("  INSERT audit_log         : 1 row")

        # post-state invariants
        post = {
            "services_total": cnt("services"),
            "entities_total": cnt("entities"),
            "groups_total": cnt("groups"),
            "nqs_history_rowcount": cnt("nqs_history"),
            "nqs_history_unique_sids": cur.execute(
                "SELECT COUNT(DISTINCT service_approval_number) FROM nqs_history;"
            ).fetchone()[0],
        }
        for k, v in post.items():
            print(f"  {k:<28} {v}")

        assert post["services_total"] == pre["services_total"], "services rowcount changed"
        assert post["entities_total"] == pre["entities_total"], "entities rowcount changed"
        assert post["groups_total"] == pre["groups_total"], "groups rowcount changed"
        assert post["nqs_history_rowcount"] == inserted_total, (
            f"nqs_history rowcount {post['nqs_history_rowcount']} != inserted {inserted_total}"
        )

        conn.commit()
        print("  COMMIT OK")

    except Exception as e:
        conn.rollback()
        print(f"  ROLLBACK: {e}")
        raise

    # --- 10. Re-verify ---
    banner("6. Post-commit re-verify")
    final = {
        "nqs_history_rowcount": cnt("nqs_history"),
        "nqs_history_unique_sids": cur.execute(
            "SELECT COUNT(DISTINCT service_approval_number) FROM nqs_history;"
        ).fetchone()[0],
        "audit_log_count": cnt("audit_log"),
    }
    for k, v in final.items():
        print(f"  {k:<28} {v}")
    cur.execute(
        "SELECT audit_id, action, subject_type, occurred_at FROM audit_log ORDER BY audit_id DESC LIMIT 1;"
    )
    print(f"  last audit row: {cur.fetchone()}")

    assert final["audit_log_count"] == pre["audit_log_count"] + 1
    assert final["nqs_history_rowcount"] == inserted_total

    conn.close()
    banner("DONE")
    print(f"  backup        : {BACKUP}")
    print(f"  rows ingested : {inserted_total:,}")
    print(f"  unique sids   : {len(sid_set):,}")
    print(f"  PA chain ops  : {chain_changes:,}")


if __name__ == "__main__":
    main()
