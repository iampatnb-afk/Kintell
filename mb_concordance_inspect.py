"""
mb_concordance_inspect.py

Read-only probe of abs_data/meshblock-correspondence-file-asgs-edn3.xlsx
to determine whether Remoteness Area (RA) columns are present.

If RA is present, we can derive SA2 -> RA from the existing repo file
(no new ABS download needed).

Output: stdout. No file writes.

Usage:
  cd <repo root>
  python mb_concordance_inspect.py
"""
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
XLSX_PATH = REPO_ROOT / "abs_data" / "meshblock-correspondence-file-asgs-edn3.xlsx"


def is_ra_keyword(s):
    if not isinstance(s, str):
        return False
    up = s.upper()
    return (
        "REMOTE" in up or
        up == "RA" or
        up.startswith("RA_") or
        up.startswith("RA ") or
        "REMOTENESS" in up
    )


def main():
    print("Meshblock correspondence inspect - read-only")
    print(f"File:   {XLSX_PATH}")
    if not XLSX_PATH.exists():
        print("ERROR: file not found.")
        return 1
    sz = XLSX_PATH.stat().st_size
    print(f"Size:   {sz:,} bytes ({sz / (1024 * 1024):.1f} MB)")
    print()

    print("Loading workbook (read-only)...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"Sheets ({len(wb.sheetnames)}):")
    for sn in wb.sheetnames:
        print(f"  - {sn}")
    print()

    overall_ra_hits = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"=== Sheet: {sn} (dims={ws.calculate_dimension()}) ===")

        # Read up to 8 rows to capture title + header rows
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 7:
                break

        for ridx, r in enumerate(rows):
            # Truncate display to first 25 cells
            shown = r[:25] if r else r
            tail = "..." if r and len(r) > 25 else ""
            print(f"  Row {ridx}: {shown}{tail}")

        # Scan rows 0..7 for RA-keyword cells
        sheet_ra_hits = []
        for ridx, r in enumerate(rows):
            if not r:
                continue
            for cidx, cell in enumerate(r):
                if is_ra_keyword(cell):
                    sheet_ra_hits.append((ridx, cidx, cell))

        if sheet_ra_hits:
            print(f"  >>> RA keyword found in this sheet:")
            for ridx, cidx, cell in sheet_ra_hits:
                print(f"      row {ridx} col {cidx}: {cell!r}")
            overall_ra_hits.extend(
                [(sn, ridx, cidx, cell)
                 for ridx, cidx, cell in sheet_ra_hits]
            )
        else:
            print(f"  (no RA keyword in first 8 rows of this sheet)")
        print()

    wb.close()

    print("=" * 64)
    if overall_ra_hits:
        print("OUTCOME: Remoteness Area columns DETECTED in correspondence file.")
        print("Path A unblocked using existing repo data. No ABS download.")
        print()
        print("RA column locations:")
        for sn, ridx, cidx, cell in overall_ra_hits:
            print(f"  Sheet '{sn}' row {ridx} col {cidx}: {cell!r}")
    else:
        print("OUTCOME: No RA columns detected in correspondence file.")
        print("Need separate ABS SA2 -> Remoteness Area concordance file.")
        print("Suggested ABS source:")
        print("  https://www.abs.gov.au/statistics/standards/")
        print("  australian-statistical-geography-standard-asgs-edition-3/")
        print("  jul2021-jun2026/access-and-downloads/")
        print("  correspondences")
    print("=" * 64)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
