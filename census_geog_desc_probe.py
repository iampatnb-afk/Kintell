"""
census_geog_desc_probe.py

Read the Metadata/2021Census_geog_desc_1st_and_2nd_release.xlsx from
inside the Census TSP zip (no extraction) and look for SA2 + Remoteness
columns.

Output: stdout. No file writes.
"""
import io
import zipfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.")
    raise SystemExit(1)


REPO_ROOT = Path(__file__).resolve().parent
ZIP_PATH = REPO_ROOT / "abs_data" / "2021_TSP_SA2_for_AUS_short-header.zip"
TARGET = "Metadata/2021Census_geog_desc_1st_and_2nd_release.xlsx"

KEYWORDS = ["aria", "remote", "remoteness", "ra_code", "ra_name", "ra_2021"]


def is_match(s):
    if s is None or not isinstance(s, str):
        return False
    return any(kw in s.lower() for kw in KEYWORDS)


def main():
    print("Census geog_desc probe - read-only")
    print(f"Zip: {ZIP_PATH}")
    print(f"Target inside zip: {TARGET}")
    print()

    if not ZIP_PATH.exists():
        print("ERROR: zip not found")
        return 1

    with zipfile.ZipFile(ZIP_PATH, "r") as zf:
        if TARGET not in zf.namelist():
            print(f"ERROR: {TARGET} not in zip.")
            print("Available metadata files:")
            for n in zf.namelist():
                if n.startswith("Metadata/"):
                    print(f"  - {n}")
            return 1
        info = zf.getinfo(TARGET)
        print(f"Embedded size: {info.file_size:,} bytes")
        with zf.open(TARGET) as fh:
            data = fh.read()
        print(f"Bytes read: {len(data):,}")
        print()

    # Open with openpyxl from BytesIO
    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True, data_only=True)
    print(f"Sheets ({len(wb.sheetnames)}):")
    for sn in wb.sheetnames:
        print(f"  - {sn}")
    print()

    found = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            dims = ws.calculate_dimension()
        except Exception:
            dims = "?"
        print(f"=== Sheet: {sn} (dims={dims}) ===")

        rows_buf = []
        for ridx, row in enumerate(ws.iter_rows(values_only=True)):
            rows_buf.append(row)
            if ridx >= 9:
                break

        # Print first 5 rows for context
        for ridx, row in enumerate(rows_buf[:5]):
            shown = row[:15] if row else row
            tail = "..." if row and len(row) > 15 else ""
            print(f"  row {ridx}: {shown}{tail}")

        # Find any cell matching keywords in first 10 rows
        sheet_hits = []
        for ridx, row in enumerate(rows_buf):
            if not row:
                continue
            for cidx, cell in enumerate(row):
                if is_match(cell):
                    sheet_hits.append((ridx, cidx, cell))

        if sheet_hits:
            print(f"  >>> {len(sheet_hits)} REMOTENESS keyword hit(s):")
            for ridx, cidx, cell in sheet_hits[:15]:
                print(f"      row {ridx} col {cidx}: {cell!r}")
            found.append((sn, sheet_hits))

            # If a header row carries RA, sample a couple data rows past
            # the headers to see actual values
            header_row_idxs = sorted({r for r, _, _ in sheet_hits})
            sample_target = max(header_row_idxs) + 1 if header_row_idxs else 1
            cols_to_sample = list({c for _, c, _ in sheet_hits})[:5]
            print(f"  Sampling data rows {sample_target}..{sample_target+2} "
                  f"at cols {cols_to_sample}:")
            try:
                # Re-iter (openpyxl read-only forward-only)
                ws2 = wb[sn]
                for ridx, row in enumerate(ws2.iter_rows(values_only=True)):
                    if ridx < sample_target:
                        continue
                    if ridx > sample_target + 2:
                        break
                    snapshot = {c: row[c] if c < len(row) else None
                                for c in cols_to_sample}
                    print(f"    row {ridx}: {snapshot}")
            except Exception as e:
                print(f"    (sample failed: {e})")
        else:
            print("  (no RA/remoteness keyword in first 10 rows)")
        print()

    wb.close()

    print("=" * 64)
    if found:
        print("OUTCOME: SA2/RA lookup likely PRESENT in TSP geog_desc xlsx.")
        for sn, hits in found:
            print(f"  Sheet '{sn}': {len(hits)} keyword hit(s)")
    else:
        print("OUTCOME: no remoteness data in TSP geog_desc.")
        print("Need separate ABS Remoteness Areas source.")
    print("=" * 64)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
