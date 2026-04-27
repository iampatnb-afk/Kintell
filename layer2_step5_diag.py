"""
Layer 2 Step 5 DIAGNOSTIC — SALM SA2 unemployment layout discovery
─────────────────────────────────────────────────────────────────
READ-ONLY. No DB mutations. No file writes outside recon/.

Per project doc 2026-04-27 — STEP 5 IS SA2-ONLY for V1. JSA does not
publish SALM at SA4 level (Decision 52). Step 5d (SA4 via ABS Cat
6202.0) deferred to P2.

Probes the SA2 SALM workbook to discover:
  - Sheet names and which one carries which metric (rate, count, LF)
  - Header row + first data row
  - SA2 code column + label column
  - Quarterly time columns and their date format
  - Suppression/null markers used for low-population SA2s
  - audit_log next id
  - Existence check on target table

Outputs:
  - Console summary
  - recon/layer2_step5_diag.md
  - recon/layer2_step5_column_map.json
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path.cwd()
DB = PROJECT_ROOT / "data" / "kintell.db"
ABS = PROJECT_ROOT / "abs_data"
OUT_DIR = PROJECT_ROOT / "recon"
OUT_MD = OUT_DIR / "layer2_step5_diag.md"
OUT_JSON = OUT_DIR / "layer2_step5_column_map.json"

SA2_PREFIX = "SALM Smoothed SA2"


def banner(title: str) -> str:
    return f"\n{'-' * 70}\n{title}\n{'-' * 70}"


def find_workbook(prefix: str) -> Path | None:
    """Return newest .xlsx in abs_data/ whose name starts with prefix."""
    matches = [p for p in ABS.glob("*.xlsx")
               if p.name.startswith(prefix)]
    if not matches:
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def classify_sheet(sheet_name: str) -> str | None:
    """Map sheet name to one of: rate, count, labour_force, unknown."""
    s = sheet_name.lower()
    if "rate" in s:
        return "rate"
    if "labour" in s and "force" in s:
        return "labour_force"
    if "count" in s or "unemploy" in s:
        return "count"
    if "smoothed" in s and "rate" not in s and "labour" not in s:
        return "count"
    return None


def looks_like_sa2(v) -> bool:
    """True if value is a 9-digit string."""
    if v is None:
        return False
    s = str(v).strip()
    return s.isdigit() and len(s) == 9


def looks_like_quarter(v) -> tuple[bool, str | None]:
    """Heuristic: is this header cell a quarter label?
    Returns (is_quarter, normalised) where normalised is e.g. '2025-Q4'."""
    if v is None:
        return (False, None)
    s = str(v).strip()
    # Date-like: year+month
    m = re.match(r"^(19|20)\d{2}[-/]?(0[1-9]|1[0-2])$", s)
    if m:
        year, month = s[:4], s[-2:]
        q = (int(month) - 1) // 3 + 1
        return (True, f"{year}-Q{q}")
    # MMM-YYYY (e.g. Dec-2025, Mar 2024)
    m = re.match(r"^([A-Za-z]{3})[-\s](19|20)\d{2}$", s)
    if m:
        mon = m.group(1).lower()
        mon_q = {"jan": 1, "feb": 1, "mar": 1,
                 "apr": 2, "may": 2, "jun": 2,
                 "jul": 3, "aug": 3, "sep": 3,
                 "oct": 4, "nov": 4, "dec": 4}
        q = mon_q.get(mon)
        if q:
            year = s[-4:]
            return (True, f"{year}-Q{q}")
    # YYYY-Qn or YYYYQn already
    m = re.match(r"^(19|20)\d{2}[-_\s]?[Qq][1-4]$", s)
    if m:
        return (True, s.upper().replace(" ", "-").replace("_", "-"))
    return (False, None)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    findings: list[str] = ["# Layer 2 Step 5 — SALM SA2 Diagnostic", ""]

    # ── 1. Locate workbook ───────────────────────────────────────────
    print(banner("1. LOCATE SA2 SALM WORKBOOK"))
    wb_path = find_workbook(SA2_PREFIX)
    if wb_path is None:
        print(f"FAIL: no workbook in abs_data/ starting with '{SA2_PREFIX}'")
        return 1
    size_mb = wb_path.stat().st_size / (1024 * 1024)
    print(f"OK: {wb_path.name} ({size_mb:.2f} MB)")
    findings.append(f"## 1. Workbook\n")
    findings.append(f"- File: `{wb_path.name}`")
    findings.append(f"- Size: {size_mb:.2f} MB\n")

    if not DB.exists():
        print(f"FAIL: DB not found at {DB}")
        return 1

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")
    findings.append(f"## 2. Sheets in workbook\n")
    findings.append(f"- All sheets: `{sheet_names}`\n")

    # ── 2. Classify sheets ───────────────────────────────────────────
    print(banner("2. CLASSIFY SHEETS"))
    classified: dict[str, str] = {}  # sheet_name -> kind
    for sn in sheet_names:
        kind = classify_sheet(sn)
        marker = kind if kind else "(unclassified)"
        print(f"  '{sn}' -> {marker}")
        if kind:
            classified[sn] = kind
        findings.append(f"- `{sn}` -> **{marker}**")
    findings.append("")

    # If multiple sheets classify to the same kind, prefer the longer name
    dedup: dict[str, str] = {}  # kind -> sheet_name
    for sn, kind in classified.items():
        if kind not in dedup or len(sn) > len(dedup[kind]):
            dedup[kind] = sn
    sheet_for_kind = dedup
    print(f"\nResolved sheet for each metric:")
    for kind, sn in sheet_for_kind.items():
        print(f"  {kind:15s} -> '{sn}'")
    findings.append(f"\n## 3. Resolved sheet per metric\n")
    for kind, sn in sheet_for_kind.items():
        findings.append(f"- {kind}: `{sn}`")
    findings.append("")

    # ── 3. Per-sheet structure probe ─────────────────────────────────
    sheet_structure: dict[str, dict] = {}
    for kind in ("rate", "count", "labour_force"):
        if kind not in sheet_for_kind:
            print(f"\nWARN: no sheet found for kind={kind}")
            findings.append(f"- WARN: no sheet found for {kind}\n")
            continue
        sn = sheet_for_kind[kind]
        print(banner(f"3.{kind.upper()} STRUCTURE: '{sn}'"))
        findings.append(f"## 4.{kind} — sheet `{sn}`\n")

        ws = wb[sn]
        rows: list[list] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if i >= 12:
                break

        n_cols = max(len(r) for r in rows)
        print(f"Total columns: {n_cols}")
        findings.append(f"- Total columns: {n_cols}")

        # Print first 12 rows, first 8 cols
        print("First 12 rows, cols 0-7:")
        findings.append("\n```")
        for r_idx, row in enumerate(rows, start=1):
            cells = []
            for c in row[:8]:
                cells.append("·" if c is None else str(c)[:18])
            line = f"Row {r_idx:2d}: " + " | ".join(cells)
            print(f"  {line}")
            findings.append(line)
        findings.append("```\n")

        # Find SA2 code column: scan first 30 data rows for 9-digit cells
        print("\nSearching for SA2 code column (9-digit values)...")
        sa2_col_candidates: dict[int, int] = {}
        for r in rows:
            for c_idx in range(min(8, len(r))):
                if looks_like_sa2(r[c_idx]):
                    sa2_col_candidates[c_idx] = sa2_col_candidates.get(c_idx, 0) + 1

        # Stream further if no hits in first 12
        if not sa2_col_candidates:
            ws_full = wb[sn]
            checked = 0
            for r_idx, row in enumerate(ws_full.iter_rows(values_only=True),
                                         start=1):
                if r_idx <= 12:
                    continue
                for c_idx in range(min(8, len(row))):
                    if looks_like_sa2(row[c_idx]):
                        sa2_col_candidates[c_idx] = (
                            sa2_col_candidates.get(c_idx, 0) + 1)
                checked += 1
                if checked >= 30:
                    break

        if sa2_col_candidates:
            sa2_col = max(sa2_col_candidates,
                          key=lambda k: sa2_col_candidates[k])
            print(f"  SA2 code column: {sa2_col} "
                  f"(seen in {sa2_col_candidates[sa2_col]} rows)")
            findings.append(f"- SA2 code column: **{sa2_col}** "
                            f"(hits: {sa2_col_candidates[sa2_col]})")
        else:
            sa2_col = None
            print("  SA2 code column NOT FOUND")
            findings.append("- SA2 code column: **NOT FOUND**")

        # Find header row: row whose cells are mostly strings/quarter-labels
        # Scan rows 1-10 for the row with most quarter-shaped cells
        header_row = None
        max_q = 0
        for r_idx, row in enumerate(rows, start=1):
            q_count = 0
            for c in row:
                is_q, _ = looks_like_quarter(c)
                if is_q:
                    q_count += 1
            if q_count > max_q:
                max_q = q_count
                header_row = r_idx

        print(f"  Header row: {header_row} (had {max_q} quarter-shaped cells)")
        findings.append(f"- Header row: **{header_row}** "
                        f"({max_q} quarter cells)")

        # Identify quarterly columns and their first/last quarters
        quarter_cols: dict[int, str] = {}
        if header_row:
            hdr = rows[header_row - 1]
            for c_idx, val in enumerate(hdr):
                is_q, normalised = looks_like_quarter(val)
                if is_q:
                    quarter_cols[c_idx] = normalised

        if quarter_cols:
            sorted_cols = sorted(quarter_cols.keys())
            first_col = sorted_cols[0]
            last_col = sorted_cols[-1]
            print(f"  Quarter columns: {len(quarter_cols)} "
                  f"({quarter_cols[first_col]} -> {quarter_cols[last_col]})")
            findings.append(f"- Quarter columns: **{len(quarter_cols)}** "
                            f"(first: {quarter_cols[first_col]} at col "
                            f"{first_col}, last: {quarter_cols[last_col]} "
                            f"at col {last_col})")
        else:
            print("  Quarter columns: none detected")
            findings.append("- Quarter columns: **none detected**")

        # Sample 3 SA2 data rows showing values
        if sa2_col is not None and quarter_cols:
            print("\n  Sample SA2 data rows:")
            ws3 = wb[sn]
            samples_shown = 0
            for r_idx, row in enumerate(ws3.iter_rows(values_only=True),
                                         start=1):
                if r_idx <= (header_row or 0):
                    continue
                if sa2_col >= len(row):
                    continue
                sa2 = row[sa2_col]
                if not looks_like_sa2(sa2):
                    continue
                first_q = row[first_col] if first_col < len(row) else None
                last_q = row[last_col] if last_col < len(row) else None
                line = (f"    sa2={sa2}  "
                        f"{quarter_cols[first_col]}={first_q}  "
                        f"{quarter_cols[last_col]}={last_q}")
                print(line)
                findings.append(f"  - {line.strip()}")
                samples_shown += 1
                if samples_shown >= 5:
                    break

        sheet_structure[kind] = {
            "sheet_name": sn,
            "header_row": header_row,
            "data_start_row": (header_row or 0) + 1,
            "sa2_code_col": sa2_col,
            "quarter_cols": quarter_cols,
        }
        findings.append("")

    wb.close()

    # ── 4. audit_log + target table check ────────────────────────────
    print(banner("4. AUDIT_LOG + TARGET TABLE"))
    findings.append("## 5. audit_log + target table\n")

    conn = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(audit_log)")
    audit_cols = [row[1] for row in cur.fetchall()]
    print(f"audit_log columns: {audit_cols}")
    findings.append(f"- audit_log columns: `{audit_cols}`")

    cur.execute("SELECT COALESCE(MAX(audit_id), 0) FROM audit_log")
    last_audit = cur.fetchone()[0]
    print(f"Last audit_id: {last_audit} (next will be {last_audit + 1})")
    findings.append(f"- Last audit_id: {last_audit} "
                    f"(next: {last_audit + 1})")

    cur.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='abs_sa2_unemployment_quarterly'
    """)
    if cur.fetchone():
        cur.execute("SELECT COUNT(*) FROM abs_sa2_unemployment_quarterly")
        n = cur.fetchone()[0]
        print(f"abs_sa2_unemployment_quarterly already exists ({n:,} rows)")
        findings.append(f"- ⚠️ `abs_sa2_unemployment_quarterly` already "
                        f"exists ({n:,} rows). Apply will refuse if "
                        f"non-empty.")
    else:
        print("abs_sa2_unemployment_quarterly: does not exist (apply "
              "will create)")
        findings.append("- `abs_sa2_unemployment_quarterly`: does not "
                        "exist (apply will create)\n")
    conn.close()

    # ── 5. Build column_map.json ─────────────────────────────────────
    print(banner("5. COLUMN MAP"))
    column_map = {
        "workbook": str(wb_path.relative_to(PROJECT_ROOT)),
        "metrics": {
            kind: {
                "sheet_name": s["sheet_name"],
                "header_row": s["header_row"],
                "data_start_row": s["data_start_row"],
                "sa2_code_col": s["sa2_code_col"],
                "quarter_cols": s["quarter_cols"],
            }
            for kind, s in sheet_structure.items()
        },
        "null_markers": ["-", "..", "np", "n.a.", "na", "NA", "NP", ""],
        "audit_log_columns": audit_cols,
        "target_table": "abs_sa2_unemployment_quarterly",
        "notes": [
            "Each metric (rate, count, labour_force) is on its own sheet.",
            "Each row is one SA2; each column is one quarter (wide format).",
            "Apply will pivot to long: (sa2_code, year_qtr, metric, value).",
        ],
    }
    OUT_JSON.write_text(json.dumps(column_map, indent=2), encoding="utf-8")
    print(f"Column map written to: {OUT_JSON}")
    print()
    print("Resolved column map summary:")
    for kind, s in sheet_structure.items():
        n_q = len(s["quarter_cols"])
        print(f"  {kind:15s}: sheet='{s['sheet_name']}' "
              f"sa2_col={s['sa2_code_col']} "
              f"data_start_row={s['data_start_row']} "
              f"quarter_cols={n_q}")

    findings.append("\n## 6. Column map (written to JSON)\n")
    findings.append("```json")
    findings.append(json.dumps(column_map, indent=2))
    findings.append("```\n")

    OUT_MD.write_text("\n".join(findings), encoding="utf-8")
    print(banner("DONE"))
    print(f"  Markdown: {OUT_MD}")
    print(f"  JSON:     {OUT_JSON}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
