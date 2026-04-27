"""
Layer 2 Step 5b-prime — Header dump (read-only).
================================================
Print row 6 (forward-filled across merged cells) and row 7 for ALL
columns of Table 1. Surfaces any sex-disaggregated or otherwise
missed columns that pattern-matching with raw row-6 values misses
when the row-6 header is on a merged cell.
"""

from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

WORKBOOK = Path("abs_data/Education and employment database.xlsx")

def main():
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Table 1"]

    # Read first 7 rows
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=7, values_only=True), 1):
        rows.append(row)

    row6 = list(rows[5])
    row7 = list(rows[6])

    # Forward-fill row 6
    last = None
    filled6 = []
    for v in row6:
        if v not in (None, ""):
            last = v
        filled6.append(last)

    print(f"{'Col':>3} {'L':<4} {'Group (row 6 forward-filled)':<58} {'Metric (row 7)':<58}")
    print("-" * 128)
    for i in range(1, ws.max_column + 1):
        g = filled6[i - 1] if i - 1 < len(filled6) else None
        m = row7[i - 1] if i - 1 < len(row7) else None
        g_s = (str(g) if g else "")[:56]
        m_s = (str(m) if m else "")[:56]
        print(f"{i:>3} {get_column_letter(i):<4} {g_s:<58} {m_s:<58}")

    wb.close()

if __name__ == "__main__":
    main()
