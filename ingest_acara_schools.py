"""
ingest_acara_schools.py — load ACARA per-school row-level data into kintell.db.

Closes the substrate gap that blocked:
  - OSHC school-attached detection (proximity match between OSHC service
    lat/lng and the nearest school lat/lng).
  - Bundle 2 schools-in-catchment drawer (per-school names + sectors +
    enrolments by SA2 — previously only aggregates were available).

Source files (committed to .gitignore in `abs_data/`, public ACARA data):
  - School Location 2025.xlsx (per-school SA2 + lat/lng + sector/type)
  - School Profile 2025.xlsx  (per-school enrolment + ICSEA)

Schema (one row per school):
  acara_schools (
    acara_sml_id INTEGER PRIMARY KEY,
    school_name  TEXT NOT NULL,
    suburb       TEXT,
    state        TEXT,
    postcode     TEXT,
    sector       TEXT,             -- Government | Catholic | Independent
    school_type  TEXT,             -- Primary | Secondary | Combined | Special
    sa2_code     TEXT,             -- 9-digit ABS SA2 code
    sa2_name     TEXT,
    lat          REAL,
    lng          REAL,
    enrolments_total INTEGER,
    icsea        INTEGER,
    fetched_at   TEXT
  )

STD-08 backup discipline: takes a pre-mutation DB backup. audit_log gets a
single row tagged `acara_schools_ingest_v1`.

Idempotent: drops + recreates the table on each run.

Run manually:
    python ingest_acara_schools.py
"""

import datetime
import shutil
import sqlite3
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent
DB_PATH = REPO_ROOT / "data" / "kintell.db"
SRC_LOCATION = REPO_ROOT / "abs_data" / "School Location 2025.xlsx"
SRC_PROFILE = REPO_ROOT / "abs_data" / "School Profile 2025.xlsx"


def _backup_db() -> Path:
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup = REPO_ROOT / "data" / f"kintell.db.backup_pre_acara_schools_{ts}"
    shutil.copy2(DB_PATH, backup)
    return backup


def _load_location_rows():
    import openpyxl
    wb = openpyxl.load_workbook(SRC_LOCATION, read_only=True, data_only=True)
    ws = wb["SchoolLocations 2025"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sml_id = row[idx["ACARA SML ID"]]
        if sml_id is None:
            continue
        out.append({
            "acara_sml_id": int(sml_id),
            "school_name":  row[idx["School Name"]],
            "suburb":       row[idx["Suburb"]],
            "state":        row[idx["State"]],
            "postcode":     str(row[idx["Postcode"]]) if row[idx["Postcode"]] is not None else None,
            "sector":       row[idx["School Sector"]],
            "school_type":  row[idx["School Type"]],
            "sa2_code":     str(row[idx["Statistical Area 2"]]) if row[idx["Statistical Area 2"]] is not None else None,
            "sa2_name":     row[idx["Statistical Area 2 Name"]],
            "lat":          float(row[idx["Latitude"]]) if row[idx["Latitude"]] is not None else None,
            "lng":          float(row[idx["Longitude"]]) if row[idx["Longitude"]] is not None else None,
        })
    return out


def _load_profile_rows():
    import openpyxl
    wb = openpyxl.load_workbook(SRC_PROFILE, read_only=True, data_only=True)
    ws = wb["SchoolProfile 2025"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sml_id = row[idx["ACARA SML ID"]]
        if sml_id is None:
            continue
        out[int(sml_id)] = {
            "enrolments_total": int(row[idx["Total Enrolments"]]) if row[idx["Total Enrolments"]] not in (None, "") else None,
            "icsea":            int(row[idx["ICSEA"]]) if row[idx["ICSEA"]] not in (None, "") else None,
        }
    return out


def main() -> int:
    if not DB_PATH.exists():
        print(f"[FAIL] DB not found: {DB_PATH}", file=sys.stderr)
        return 1
    if not SRC_LOCATION.exists() or not SRC_PROFILE.exists():
        print(f"[FAIL] Source xlsx files missing in {REPO_ROOT / 'abs_data'}", file=sys.stderr)
        return 1
    print("=== ACARA schools ingest ===")
    print(f"  DB: {DB_PATH}")
    print(f"  Location: {SRC_LOCATION.name}")
    print(f"  Profile:  {SRC_PROFILE.name}")
    backup = _backup_db()
    print(f"  [STD-08] Pre-ingest backup: {backup.name}")

    print("  Loading School Location 2025…")
    loc_rows = _load_location_rows()
    print(f"    {len(loc_rows)} location rows")
    print("  Loading School Profile 2025…")
    prof_by_id = _load_profile_rows()
    print(f"    {len(prof_by_id)} profile rows")

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        cur.execute("DROP TABLE IF EXISTS acara_schools")
        cur.execute("""
            CREATE TABLE acara_schools (
                acara_sml_id     INTEGER PRIMARY KEY,
                school_name      TEXT NOT NULL,
                suburb           TEXT,
                state            TEXT,
                postcode         TEXT,
                sector           TEXT,
                school_type      TEXT,
                sa2_code         TEXT,
                sa2_name         TEXT,
                lat              REAL,
                lng              REAL,
                enrolments_total INTEGER,
                icsea            INTEGER,
                fetched_at       TEXT
            )
        """)
        cur.execute("CREATE INDEX idx_acara_schools_sa2 ON acara_schools(sa2_code)")
        cur.execute("CREATE INDEX idx_acara_schools_lat ON acara_schools(lat)")
        cur.execute("CREATE INDEX idx_acara_schools_lng ON acara_schools(lng)")
        now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat()
        ins = 0
        for r in loc_rows:
            prof = prof_by_id.get(r["acara_sml_id"], {})
            cur.execute("""
                INSERT INTO acara_schools (
                    acara_sml_id, school_name, suburb, state, postcode,
                    sector, school_type, sa2_code, sa2_name, lat, lng,
                    enrolments_total, icsea, fetched_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                r["acara_sml_id"], r["school_name"], r["suburb"], r["state"], r["postcode"],
                r["sector"], r["school_type"], r["sa2_code"], r["sa2_name"], r["lat"], r["lng"],
                prof.get("enrolments_total"), prof.get("icsea"), now_iso,
            ))
            ins += 1
        # audit_log row — schema: actor, action, subject_type, subject_id,
        # before_json, after_json, reason, occurred_at
        cur.execute("""
            INSERT INTO audit_log (actor, action, subject_type, subject_id, after_json, reason, occurred_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "ingest_acara_schools.py",
            "acara_schools_ingest_v1",
            "table", "acara_schools",
            f'{{"rows": {ins}, "backup": "{backup.name}"}}',
            "First per-school row-level ingest (Location 2025 + Profile 2025) — unlocks OSHC school-attached detection + Bundle 2 schools-in-catchment drawer.",
            now_iso,
        ))
        conn.commit()
        print(f"[ok] Inserted {ins} schools into acara_schools.")
        # Quick stats
        cur.execute("SELECT COUNT(*), COUNT(lat), COUNT(enrolments_total), COUNT(icsea) FROM acara_schools")
        n, n_lat, n_enr, n_ics = cur.fetchone()
        print(f"     coverage: {n_lat}/{n} lat-lng · {n_enr}/{n} enrolments · {n_ics}/{n} ICSEA")
        cur.execute("SELECT sector, COUNT(*) FROM acara_schools GROUP BY sector ORDER BY 2 DESC")
        for sec, c in cur.fetchall():
            print(f"     {sec or '(no sector)'}: {c}")
    finally:
        conn.close()
    print("=== Done ===")
    return 0


if __name__ == "__main__":
    sys.exit(main())
