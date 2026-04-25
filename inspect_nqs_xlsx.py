import openpyxl, json
from pathlib import Path

xlsx = Path("abs_data/NQS Data Q4 2025.XLSX")
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

summary = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    first_data = next(rows, None)
    second_data = next(rows, None)
    # Count rows cheaply
    row_count = ws.max_row
    summary[sheet_name] = {
        "row_count": row_count,
        "headers": [str(h) if h is not None else None for h in (headers or [])],
        "sample_row_1": [str(v) if v is not None else None for v in (first_data or [])],
        "sample_row_2": [str(v) if v is not None else None for v in (second_data or [])],
    }

Path("nqs_xlsx_inspect.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
print(f"Sheets: {list(summary.keys())}")
for name, info in summary.items():
    print(f"  {name}: {info['row_count']} rows, {len(info['headers'])} columns")
print("wrote nqs_xlsx_inspect.json")
