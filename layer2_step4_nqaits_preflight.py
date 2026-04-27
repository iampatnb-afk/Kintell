# layer2_step4_nqaits_preflight.py — read-only
# Verifies NQAITS join key, cross-quarter stability, PA-chain capture, header drift.
import os, sys, sqlite3, datetime, re
from collections import Counter, defaultdict

NQAITS = os.path.join("abs_data", "NQAITS Quarterly Data Splits (Q3 2013 - Q4 2025).xlsx")
DB = os.path.join("data", "kintell.db")
OUT = os.path.join("recon", "layer2_step4_nqaits_preflight.md")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("FATAL: pip install openpyxl")

if not os.path.exists(NQAITS):
    sys.exit(f"FATAL: {NQAITS} not found")
if not os.path.exists(DB):
    sys.exit(f"FATAL: {DB} not found")

lines = []
def emit(s=""):
    print(s, flush=True)
    lines.append(s)

emit("# Phase 2.5 Layer 2 Step 4 — NQAITS Preflight")
emit(f"_Generated: {datetime.datetime.now().isoformat(timespec='seconds')}_")
emit("")

print("[opening NQAITS — 138MB, ~30s ...]", flush=True)
wb = load_workbook(NQAITS, read_only=True, data_only=True)
data_sheets = [s for s in wb.sheetnames if re.match(r"Q\d20\d\ddata", s)]
emit(f"## Data sheets")
emit(f"  total: {len(data_sheets)}")
emit(f"  range: {data_sheets[-1]} -> {data_sheets[0]}")  # oldest -> newest
emit("")

# Build header map across ALL data sheets
print(f"[scanning headers across {len(data_sheets)} sheets ...]", flush=True)
sheet_headers = {}
for s in data_sheets:
    ws = wb[s]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        h = next(rows_iter)
        sheet_headers[s] = [str(c).strip() if c is not None else None for c in h]
    except StopIteration:
        sheet_headers[s] = []

# Header drift analysis
emit("## Header drift across all 50 data sheets")
all_col_names = set()
for h in sheet_headers.values():
    all_col_names.update(c for c in h if c)
emit(f"  unique column names across all sheets: {len(all_col_names)}")
emit("")

# Group cols by which sheets have them
col_sheet_count = Counter()
for s, h in sheet_headers.items():
    for c in h:
        if c:
            col_sheet_count[c] += 1

emit("  column -> sheet count (cols not in all 50 are drifters):")
for c, n in sorted(col_sheet_count.items(), key=lambda x: (-x[1], x[0])):
    flag = " <-- DRIFT" if n != len(data_sheets) else ""
    emit(f"    {n:>3}/{len(data_sheets)}  {c!r}{flag}")
emit("")

# Service ID column position per sheet (must be column 0 every time)
emit("## Service ID column position consistency")
positions = Counter()
for s, h in sheet_headers.items():
    for i, c in enumerate(h):
        if c == "Service ID":
            positions[i] += 1
            break
emit(f"  Service ID position distribution: {dict(positions)}")
emit("")

# Build service_id -> quarter set across all sheets
print("[scanning Service IDs across all sheets — this takes ~60s ...]", flush=True)
sid_quarters = defaultdict(set)
sid_provider_per_quarter = defaultdict(dict)  # sid -> {quarter: provider_id}
for idx, s in enumerate(data_sheets):
    if idx % 10 == 0:
        print(f"  sheet {idx+1}/{len(data_sheets)}: {s}", flush=True)
    ws = wb[s]
    h = sheet_headers[s]
    try:
        sid_idx = h.index("Service ID")
    except ValueError:
        emit(f"  WARN: sheet {s} missing Service ID")
        continue
    try:
        pid_idx = h.index("Provider ID")
    except ValueError:
        pid_idx = None
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter)  # header
    for row in rows_iter:
        if not row or row[sid_idx] is None:
            continue
        sid = str(row[sid_idx]).strip()
        if not sid:
            continue
        sid_quarters[sid].add(s)
        if pid_idx is not None and row[pid_idx] is not None:
            pid = str(row[pid_idx]).strip()
            sid_provider_per_quarter[sid][s] = pid
wb.close()

emit("## Cross-quarter coverage of Service IDs")
emit(f"  unique Service IDs across all 50 quarters: {len(sid_quarters):,}")
qcounts = Counter(len(qs) for qs in sid_quarters.values())
emit(f"  distribution of services by quarter-count present:")
for q, n in sorted(qcounts.items(), reverse=True)[:15]:
    emit(f"    {q:>3} quarters: {n:>6} services")
emit("")

# Services in ALL 50 quarters (the rock-solid centre-following set)
all_q = set(data_sheets)
in_all = [sid for sid, qs in sid_quarters.items() if qs == all_q]
emit(f"  Service IDs present in ALL {len(data_sheets)} quarters: {len(in_all):,}")
emit("")

# PA chain analysis — services with Provider ID change
emit("## Provider ID change analysis (PA chain reconstruction)")
chain_changes = 0
sample_changers = []
for sid, qprov in sid_provider_per_quarter.items():
    distinct_pids = set(qprov.values())
    if len(distinct_pids) > 1:
        chain_changes += 1
        if len(sample_changers) < 5:
            # show first 4 quarter->pid pairs sorted by quarter (newest first)
            sample = sorted(qprov.items(), reverse=True)[:6]
            sample_changers.append((sid, distinct_pids, sample))
emit(f"  services with >1 distinct Provider ID across quarters: {chain_changes:,}")
emit(f"  ({chain_changes / len(sid_quarters) * 100:.1f}% of all services)")
emit("")
emit("  sample PA-chain changers:")
for sid, pids, sample in sample_changers:
    emit(f"    {sid} touched {len(pids)} provider IDs")
    for q, p in sample:
        emit(f"      {q} -> {p}")
emit("")

# Join check vs services.service_approval_number
emit("## Join check: NQAITS Service ID vs services.service_approval_number")
conn = sqlite3.connect(DB)
cur = conn.cursor()
cur.execute("SELECT service_approval_number FROM services WHERE service_approval_number IS NOT NULL;")
db_sans = {row[0] for row in cur.fetchall()}
conn.close()

nqaits_sids = set(sid_quarters.keys())
overlap = db_sans & nqaits_sids
db_only = db_sans - nqaits_sids
nq_only = nqaits_sids - db_sans

emit(f"  services in DB           : {len(db_sans):,}")
emit(f"  unique sids in NQAITS     : {len(nqaits_sids):,}")
emit(f"  matched (DB ∩ NQAITS)    : {len(overlap):,}  ({len(overlap)/len(db_sans)*100:.2f}% of DB)")
emit(f"  in DB but not NQAITS      : {len(db_only):,}")
emit(f"  in NQAITS but not DB      : {len(nq_only):,}")
emit("")
emit("  sample DB-only (5):")
for s in list(db_only)[:5]:
    emit(f"    {s!r}")
emit("  sample NQAITS-only (5):")
for s in list(nq_only)[:5]:
    emit(f"    {s!r}")
emit("")

# Format check on a sample sid
emit("## Sample value checks")
sample_sid = next(iter(overlap))
emit(f"  sampling sid: {sample_sid!r}")
emit(f"  in NQAITS quarters: {len(sid_quarters[sample_sid])} of {len(data_sheets)}")
if sample_sid in sid_provider_per_quarter:
    pids = sid_provider_per_quarter[sample_sid]
    emit(f"  provider history: {len(set(pids.values()))} distinct provider(s)")
emit("")

# Volume estimate
emit("## Ingest volume estimate")
total_rowq = sum(len(qs) for qs in sid_quarters.values())
emit(f"  total (sid, quarter) pairs across all sheets: {total_rowq:,}")
emit(f"  avg quarters per service: {total_rowq / len(sid_quarters):.1f}")
emit("")

os.makedirs("recon", exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"\n[wrote {OUT}]")
