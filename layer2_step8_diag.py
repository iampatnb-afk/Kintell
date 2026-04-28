"""
layer2_step8_diag.py — Read-only diagnostic for ABS Births SA2 ingest.

Input:
  abs_data/Births_SA2_2011_2024.xlsx  (ABS Cat 3301.0 Table 2)

Output:
  recon/layer2_step8_diag.md
  terminal summary

Probes:
  1. Workbook structure (sheets, dimensions)
  2. Header rows 1-8 raw dump
  3. Format detection (WIDE per Standard 19 vs LONG per Standard 26)
  4. SA2-level row detection (rows below national/state aggregates per Std 25)
  5. Year column / metric detection
  6. SA2-level sample rows
  7. National aggregate sanity check vs ABS published births

No mutations.
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path

XLSX_PATH = "abs_data/Births_SA2_2011_2024.xlsx"
OUT_PATH = "recon/layer2_step8_diag.md"
HEADER_ROWS = 8
SAMPLE_SA2_ROWS = 8
DEEP_SCAN_ROW_LIMIT = 200  # how far down to look for SA2-level start
MAX_COLS_TO_REPORT = 25  # cap on per-row dumps for readability

# ABS published national births (Cat 3301.0 Table 1) — sanity baseline
# Approx values; tolerant comparison only.
ABS_NATIONAL_BIRTHS = {
    2011: 297900, 2012: 309600, 2013: 308100, 2014: 299700,
    2015: 305400, 2016: 311100, 2017: 309100, 2018: 315100,
    2019: 305800, 2020: 294400, 2021: 309900, 2022: 300700,
    2023: 286900, 2024: 292500,
}
TOLERANCE_PCT = 5.0  # +/- 5% considered consistent

# ---------------------------------------------------------------------------
# Preflight
# ---------------------------------------------------------------------------
if not Path(XLSX_PATH).exists():
    print(f"ERROR: not found: {XLSX_PATH}", file=sys.stderr)
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    print("Run: pip install --break-system-packages openpyxl", file=sys.stderr)
    sys.exit(2)

# ---------------------------------------------------------------------------
# Load workbook
# ---------------------------------------------------------------------------
print(f"Loading {XLSX_PATH} ...")
try:
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
except Exception as e:
    print(f"ERROR loading workbook: {e}", file=sys.stderr)
    print("If the error mentions externalLinks .rels, see Standard 15 "
          "(zip-surgery repair).", file=sys.stderr)
    sys.exit(3)

# ---------------------------------------------------------------------------
# Output accumulator
# ---------------------------------------------------------------------------
lines: list[str] = []


def w(s: str = "") -> None:
    lines.append(s)


def fmt_cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    if len(s) > 60:
        s = s[:57] + "..."
    return s.replace("|", "\\|").replace("\n", " ")


# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
size_kb = os.path.getsize(XLSX_PATH) / 1024
w("# Layer 2 Step 8 — Births SA2 Diagnostic")
w("")
w(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
w(f"Source: `{XLSX_PATH}` ({size_kb:,.0f} KB)")
w(f"Sheets: {len(wb.sheetnames)} — {', '.join(repr(s) for s in wb.sheetnames)}")
w("")
w("Read-only diagnostic. No mutations.")
w("")

# ---------------------------------------------------------------------------
# Per-sheet probe
# ---------------------------------------------------------------------------
def stream_rows(ws, max_rows: int):
    """Yield first max_rows rows as lists of cell values."""
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        out.append(list(row))
    return out


# Heuristics
SA2_CODE_RE = re.compile(r"^\d{9}$")  # 2021 SA2 codes are 9 digits


def is_sa2_code(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return bool(SA2_CODE_RE.match(s))


def looks_like_year(v) -> bool:
    try:
        n = int(v)
        return 2000 <= n <= 2030
    except (TypeError, ValueError):
        return False


def detect_format(rows: list[list]) -> dict:
    """Detect WIDE (Standard 19) vs LONG (Standard 26) layout.

    Heuristic: scan rows 1..HEADER_ROWS for column headers:
      - LONG if a header cell named exactly 'Year' or 'YEAR' (case-insensitive)
        appears in the first ~5 columns AND values below it are int years.
      - WIDE if multiple distinct year-like values appear across one header row
        (typically row 6 or 7) as column headers.
      - Unknown otherwise.
    """
    n_rows = len(rows)
    candidate = {"format": "unknown", "header_row": None,
                 "year_cols": [], "year_col_idx": None,
                 "sa2_code_col_idx": None, "label_col_idx": None}

    for r_idx, row in enumerate(rows):
        if not row:
            continue
        # LONG: explicit 'year' header cell
        for c_idx, val in enumerate(row[:8]):
            if val is None:
                continue
            sval = str(val).strip().lower()
            if sval in ("year", "yr"):
                # confirm by checking col below has year-like ints
                # we only have HEADER_ROWS rows here; defer confirmation
                candidate["format"] = "long"
                candidate["header_row"] = r_idx + 1  # 1-indexed
                candidate["year_col_idx"] = c_idx
                # mark Code/Label cols if neighbors look right
                for cc, vv in enumerate(row[:8]):
                    if vv is None:
                        continue
                    svv = str(vv).strip().lower()
                    if svv in ("code", "sa2_code", "sa2 code"):
                        candidate["sa2_code_col_idx"] = cc
                    elif svv in ("label", "name", "sa2_name", "sa2 name"):
                        candidate["label_col_idx"] = cc
                return candidate

        # WIDE: row contains many year-like values
        year_cols = [(c_idx, v) for c_idx, v in enumerate(row)
                     if looks_like_year(v)]
        if len(year_cols) >= 3:  # at least 3 distinct years across cols
            candidate["format"] = "wide"
            candidate["header_row"] = r_idx + 1
            candidate["year_cols"] = year_cols
            return candidate

    return candidate


def find_first_sa2_row(ws, start_row: int = 1, limit: int = DEEP_SCAN_ROW_LIMIT) -> int | None:
    """Return 1-indexed row number of the first row whose first ~5 cells contain
    a 9-digit SA2 code. Returns None if not found within `limit` rows."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i + 1 < start_row:
            continue
        if i + 1 > limit:
            break
        for c in row[:6]:
            if is_sa2_code(c):
                return i + 1
    return None


# Run probe for every sheet (births workbook usually has Contents + Table sheets)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    w(f"## Sheet: `{sheet_name}`")
    w("")

    # Dimensions
    try:
        max_row = ws.max_row
        max_col = ws.max_column
    except Exception:
        max_row = max_col = "?"
    w(f"- max_row: **{max_row}**, max_col: **{max_col}**")

    # Read first HEADER_ROWS rows
    header = stream_rows(ws, HEADER_ROWS)
    w(f"- read first {len(header)} rows for header analysis")
    w("")

    # Header dump (first MAX_COLS_TO_REPORT cols)
    w("### Raw header rows")
    w("")
    if header:
        # Build column index header
        max_c = min(max(len(r) for r in header) if header else 0,
                    MAX_COLS_TO_REPORT)
        col_hdr = "| row | " + " | ".join(
            f"c{i+1}" for i in range(max_c)
        ) + " |"
        sep = "|---|" + "|".join("---" for _ in range(max_c)) + "|"
        w(col_hdr)
        w(sep)
        for r_idx, row in enumerate(header, start=1):
            cells = [fmt_cell(v) for v in row[:max_c]]
            w("| " + str(r_idx) + " | " + " | ".join(cells) + " |")
    w("")

    # Format detection
    fmt = detect_format(header)
    w("### Format detection")
    w("")
    w(f"- Detected format: **{fmt['format']}** "
      f"(WIDE per Std 19 vs LONG per Std 26)")
    if fmt["header_row"] is not None:
        w(f"- Probable header row: **{fmt['header_row']}**")
    if fmt["format"] == "long":
        w(f"- Year column index (0-based): {fmt['year_col_idx']}")
        w(f"- SA2 code column index: {fmt['sa2_code_col_idx']}")
        w(f"- Label column index: {fmt['label_col_idx']}")
    elif fmt["format"] == "wide":
        ycs = fmt["year_cols"]
        w(f"- Year-like column count: {len(ycs)}")
        if ycs:
            yrs = [v for _, v in ycs]
            w(f"- Year span: {min(yrs)} → {max(yrs)}")
            w(f"- Year columns (0-based, value): "
              f"{', '.join(f'{i}={v}' for i, v in ycs[:25])}"
              f"{' ...' if len(ycs) > 25 else ''}")
    w("")

    # Find first SA2-level row
    first_sa2 = find_first_sa2_row(ws, start_row=1)
    w("### SA2-level row detection")
    w("")
    if first_sa2 is None:
        w(f"- (No 9-digit SA2 code found in first {DEEP_SCAN_ROW_LIMIT} rows)")
    else:
        w(f"- First row with 9-digit SA2 code: **row {first_sa2}**")
        w("")
        # Read 8 SA2-level rows starting there
        w(f"### Sample SA2-level rows (rows {first_sa2}–{first_sa2 + SAMPLE_SA2_ROWS - 1})")
        w("")
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=first_sa2,
                                              max_row=first_sa2 + SAMPLE_SA2_ROWS - 1,
                                              values_only=True)):
            sample_rows.append(list(row))

        if sample_rows:
            max_c = min(max(len(r) for r in sample_rows), MAX_COLS_TO_REPORT)
            col_hdr = "| row | " + " | ".join(
                f"c{i+1}" for i in range(max_c)
            ) + " |"
            sep = "|---|" + "|".join("---" for _ in range(max_c)) + "|"
            w(col_hdr)
            w(sep)
            for r_idx, row in enumerate(sample_rows, start=first_sa2):
                cells = [fmt_cell(v) for v in row[:max_c]]
                w("| " + str(r_idx) + " | " + " | ".join(cells) + " |")
        w("")

        # National aggregate sanity check
        # Strategy: if WIDE, sum each year column across all SA2 rows; compare
        #           to ABS published.
        #           If LONG, group by year column.
        w("### National aggregate sanity check")
        w("")
        if fmt["format"] == "wide" and fmt["year_cols"]:
            # Sum each year column across SA2-level rows
            year_col_idx_map = {v: i for i, v in fmt["year_cols"]}
            sums: dict[int, int] = {y: 0 for y in year_col_idx_map}
            counts: dict[int, int] = {y: 0 for y in year_col_idx_map}
            sa2_count = 0
            for row in ws.iter_rows(min_row=first_sa2, values_only=True):
                if not row:
                    continue
                # Determine if this is an SA2 row
                first_cell = row[0] if len(row) > 0 else None
                if not is_sa2_code(first_cell):
                    # tolerate name/code reordering: scan first 6 cells
                    if not any(is_sa2_code(c) for c in row[:6]):
                        continue
                sa2_count += 1
                for year, c_idx in year_col_idx_map.items():
                    if c_idx >= len(row):
                        continue
                    v = row[c_idx]
                    if v is None:
                        continue
                    try:
                        sums[year] += int(v)
                        counts[year] += 1
                    except (ValueError, TypeError):
                        # skip 'np' (not published) or text suppressors
                        pass

            w(f"- SA2 rows aggregated: **{sa2_count:,}**")
            w("")
            w("| year | SA2-sum | ABS published | delta | within ±5%? |")
            w("|---:|---:|---:|---:|:-:|")
            for year in sorted(year_col_idx_map):
                obs = sums[year]
                exp = ABS_NATIONAL_BIRTHS.get(year)
                if exp is None:
                    w(f"| {year} | {obs:,} | (no baseline) | — | — |")
                else:
                    delta = obs - exp
                    pct = (delta / exp * 100) if exp else 0
                    ok = abs(pct) <= TOLERANCE_PCT
                    mark = "✓" if ok else "⚠"
                    w(f"| {year} | {obs:,} | {exp:,} "
                      f"| {delta:+,} ({pct:+.1f}%) | {mark} |")
            w("")
            w(f"Tolerance: ±{TOLERANCE_PCT:.0f}% vs ABS Cat 3301.0 Table 1.")
            w("Misses larger than tolerance suggest column mis-detection or "
              "double-counting (state/national rows leaking into SA2 sum).")

        elif fmt["format"] == "long" and fmt["year_col_idx"] is not None:
            w("(LONG format detected — national aggregate by year would require "
              "scanning all rows; not yet implemented in diag. Apply phase will "
              "compute.)")
        else:
            w("(Format unknown — skipping aggregate.)")
        w("")

    w("---")
    w("")

# ---------------------------------------------------------------------------
# Recommended apply schema
# ---------------------------------------------------------------------------
w("## Recommended apply phase")
w("")
w("**Target table** (proposed):")
w("")
w("```sql")
w("CREATE TABLE abs_sa2_births_annual (")
w("  sa2_code     TEXT NOT NULL,")
w("  year         INTEGER NOT NULL,")
w("  births_count INTEGER,")
w("  PRIMARY KEY (sa2_code, year)")
w(");")
w("CREATE INDEX idx_abs_sa2_births_annual_sa2  ON abs_sa2_births_annual(sa2_code);")
w("CREATE INDEX idx_abs_sa2_births_annual_year ON abs_sa2_births_annual(year);")
w("```")
w("")
w("**Audit log row** (action='abs_sa2_births_ingest_v1', subject_type="
  "'abs_sa2_births_annual').")
w("")
w("**Backup**: `data/kintell.db.backup_pre_step8_<ts>` before mutation (Std 8).")
w("")
w("**Invariants** (apply phase):")
w("")
w("- distinct SA2 count between 2,000 and 2,500")
w("- year coverage 2011-2024 (or whatever the workbook contains)")
w("- per-year national sum within ±5% of ABS Cat 3301.0 Table 1")
w("- no births_count < 0; no SA2 codes outside abs_sa2_erp_annual")
w("")
w("---")
w("End of diagnostic.")

# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------
Path("recon").mkdir(exist_ok=True)
Path(OUT_PATH).write_text("\n".join(lines), encoding="utf-8")
wb.close()

print()
print(f"OK  wrote {OUT_PATH}")
print(f"    sheets: {len(wb.sheetnames)}")
